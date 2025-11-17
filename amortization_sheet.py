"""
시트 생성 및 설정 모듈
- 전대 계약별 시트 생성
- 헤더 설정
- 데이터 입력
"""

from excel_utils import add_sheet, set_cell_value, get_cell_value


# ============================================================================
# 헤더 설정 (수정이 쉽도록 상수로 관리)
# ============================================================================

# 헤더 리스트 (A열부터 순서대로, 총 28개)
# 이미지에서 보여준 순서대로 정렬
HEADERS = [
    "계약번호(Ref no.)",
    "리스명",
    "상위리스_계약번호(Ref no.)",
    "리스개시일",
    "리스종료일",
    "지급일자",
    "지급연도",
    "유효일자",
    "회계연도",
    "결산월",
    "기간(월)",
    "감가상각기간(월)",
    "수령회차",
    "리스료",
    "할인율(연)",
    "임대보증금할인율(연)",
    "리스채권",
    "이자수익",
    "상각액",
    "리스채권_유동",
    "전대비율",
    "차감_사용권자산",
    "차감_감가상각비",
    "차감_감가상각누계액",
    "차감_사용권자산(순)",
    "임대보증금",
    "임대보증금_현할차",
    "임대보증금이자",
]

# databody에서 직접 가져올 헤더들 (계약 데이터에서 그대로 사용, 10개)
# 이전 이미지에서 지정한 9개 헤더 + 전대비율 (실제 데이터 컬럼명과 매칭 필요)
HEADERS_FROM_DATA = [
    "계약번호(Ref no.)",
    "리스명",
    "상위리스_계약번호(Ref no.)",  # 또는 "상위리스 계약번호(Ref no.)", "상위리스"
    "리스개시일",
    "리스종료일",
    "기간(월)",
    "감가상각기간(월)",
    "할인율(연)",
    "임대보증금할인율(연)",  # 또는 "보증금할인율(연)", "보증금할인"
    "전대비율",
]


# ============================================================================
# 유형별 행수 계산
# ============================================================================

def calculate_data_rows(payment_type, start_basis, depreciation_period):
    """
    유형별 데이터 행수 계산
    
    Args:
        payment_type (str): '선급' 또는 '후급'
        start_basis (str): '월초' 또는 '월말'
        depreciation_period (int): 감가상각기간 값
        
    Returns:
        int: 데이터 행수
    """
    if start_basis == '월초':
        # 월초 case: 감가상각기간 + 2
        return depreciation_period + 2
    elif start_basis == '월말':
        # 월말 case: 감가상각기간 + 1
        return depreciation_period + 1
    else:
        # 기본값
        return depreciation_period + 1


# ============================================================================
# 시트 생성 및 설정
# ============================================================================

def create_contract_sheet(wb, contract_no, contract_data):
    """
    계약번호를 시트명으로 하는 새 시트 생성 및 기본 설정
    
    Args:
        wb: Workbook 객체
        contract_no: 계약번호 (시트명으로 사용)
        contract_data: 계약 데이터 (딕셔너리)
        
    Returns:
        Worksheet: 생성된 시트 객체
    """
    # 시트명은 계약번호로 (Excel 시트명 제한 고려)
    # Excel 시트명에 사용할 수 없는 문자 제거: \ / ? * [ ]
    sheet_name = str(contract_no)
    invalid_chars = ['\\', '/', '?', '*', '[', ']']
    for char in invalid_chars:
        sheet_name = sheet_name.replace(char, '_')
    
    # Excel 시트명 최대 31자
    sheet_name = sheet_name[:31]
    
    # 시트 생성
    ws = add_sheet(wb, sheet_name)
    
    print(f"  ✓ 시트 생성: {sheet_name}")
    
    return ws


def set_headers(ws, headers=None, start_row=1, start_col=1):
    """
    헤더를 시트에 설정
    
    Args:
        ws: Worksheet 객체
        headers (list): 헤더 리스트 (None이면 기본 HEADERS 사용)
        start_row (int): 시작 행 (기본값: 1)
        start_col (int): 시작 열 (기본값: 1, A열)
        
    Returns:
        int: 헤더가 설정된 마지막 열 번호
    """
    if headers is None:
        headers = HEADERS
    
    from openpyxl.utils import get_column_letter
    
    for idx, header in enumerate(headers):
        col_letter = get_column_letter(start_col + idx)
        cell_address = f"{col_letter}{start_row}"
        set_cell_value(ws, cell_address, header)
    
    print(f"  ✓ 헤더 설정 완료: {len(headers)}개 컬럼")
    
    return start_col + len(headers) - 1


def fill_databody_from_contract(ws, contract_data, headers_from_data=None, start_row=2, num_rows=1):
    """
    계약 데이터에서 정보를 가져와 databody 행들에 입력
    
    Args:
        ws: Worksheet 객체
        contract_data: 계약 데이터 (딕셔너리)
        headers_from_data (list): 데이터에서 가져올 헤더 리스트 (None이면 기본값 사용)
        start_row (int): databody 시작 행 (기본값: 2, 헤더 다음 행)
        num_rows (int): 데이터를 입력할 행 수
    """
    if headers_from_data is None:
        headers_from_data = HEADERS_FROM_DATA
    
    from openpyxl.utils import get_column_letter
    
    filled_count = 0
    
    # 각 행에 동일한 데이터 입력
    for row_idx in range(start_row, start_row + num_rows):
        for col_idx, header_name in enumerate(HEADERS, start=1):
            if header_name in headers_from_data:
                # 데이터에서 값 가져오기 (다양한 컬럼명 시도)
                value = None
                
                # 정확한 매칭 시도
                if header_name in contract_data:
                    value = contract_data.get(header_name)
                else:
                    # 유사한 컬럼명 찾기 (줄바꿈, 공백, 언더스코어 차이 무시)
                    value = None
                    for key in contract_data.keys():
                        key_str = str(key).strip()
                        header_str = header_name.strip()
                        
                        # 줄바꿈, 탭, 여러 공백을 단일 공백으로 정규화
                        import re
                        key_normalized = re.sub(r'\s+', ' ', key_str.replace('\n', ' ').replace('\r', ' ').replace('\t', ' '))
                        header_normalized = re.sub(r'\s+', ' ', header_str.replace('\n', ' ').replace('\r', ' ').replace('\t', ' '))
                        
                        # 정확히 일치
                        if key_normalized == header_normalized:
                            value = contract_data.get(key)
                            break
                        
                        # 공백/언더스코어 차이 무시하여 비교
                        key_no_space = key_normalized.replace(' ', '').replace('_', '')
                        header_no_space = header_normalized.replace(' ', '').replace('_', '')
                        if key_no_space == header_no_space:
                            value = contract_data.get(key)
                            break
                        
                        # 부분 일치 (헤더가 키에 포함되거나 그 반대)
                        if header_normalized in key_normalized or key_normalized in header_normalized:
                            value = contract_data.get(key)
                            break
                    
                    # 여전히 못 찾으면 빈 값
                    if value is None:
                        value = ''
                
                col_letter = get_column_letter(col_idx)
                cell_address = f"{col_letter}{row_idx}"
                set_cell_value(ws, cell_address, value)
                
                if row_idx == start_row:  # 첫 번째 행만 출력
                    filled_count += 1
                    if value:
                        print(f"    - {header_name}: {value}")
        
        if row_idx == start_row:
            print(f"  ✓ {len([h for h in HEADERS if h in headers_from_data])}개 필드 입력 완료")
    
    print(f"  ✓ Databody 입력 완료: {num_rows}행에 데이터 입력")


def fill_receipt_count(ws, payment_type, start_basis, num_rows, start_row=2):
    """
    수령회차 열 채우기
    
    Args:
        ws: Worksheet 객체
        payment_type (str): '선급' 또는 '후급'
        start_basis (str): '월초' 또는 '월말'
        num_rows (int): 채울 행 수
        start_row (int): 시작 행 (기본값: 2)
    """
    from openpyxl.utils import get_column_letter
    
    # 수령회차 열 찾기 (HEADERS에서 인덱스)
    receipt_count_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "수령회차":
            receipt_count_col_idx = idx
            break
    
    if receipt_count_col_idx is None:
        print("    경고: '수령회차' 헤더를 찾을 수 없습니다.")
        return
    
    col_letter = get_column_letter(receipt_count_col_idx)
    
    # 수령회차 시작값 결정
    # 월초 유형: 0부터 시작
    # 월말 유형: 선급/월말은 1부터, 후급/월말은 0부터 시작
    if start_basis == '월초':
        start_value = 0
    else:  # 월말
        if payment_type == '후급':
            start_value = 0  # 후급/월말: 0부터 시작
        else:
            start_value = 1  # 선급/월말: 1부터 시작
    
    # 각 행에 값 입력
    for row_idx in range(start_row, start_row + num_rows):
        value = start_value + (row_idx - start_row)
        cell_address = f"{col_letter}{row_idx}"
        set_cell_value(ws, cell_address, value)
    
    print(f"  ✓ 수령회차 입력 완료: {start_value}부터 {num_rows}행 ({payment_type}/{start_basis})")


def fill_lease_fee(ws, contract_data, num_rows, start_row=2):
    """
    리스료 열 채우기
    - 수령회차가 1인 라인부터 감가상각대상기간(월) 값만큼 채우기
    - 값은 고정리스료(월) 컬럼의 값 사용
    
    Args:
        ws: Worksheet 객체
        contract_data: 계약 데이터 (딕셔너리)
        num_rows (int): 전체 행 수
        start_row (int): 시작 행 (기본값: 2)
    """
    from openpyxl.utils import get_column_letter
    
    # 리스료 열 찾기
    lease_fee_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "리스료":
            lease_fee_col_idx = idx
            break
    
    if lease_fee_col_idx is None:
        print("    경고: '리스료' 헤더를 찾을 수 없습니다.")
        return
    
    # 수령회차 열 찾기
    receipt_count_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "수령회차":
            receipt_count_col_idx = idx
            break
    
    if receipt_count_col_idx is None:
        print("    경고: '수령회차' 헤더를 찾을 수 없습니다.")
        return
    
    # 고정리스료(월) 값 가져오기
    fixed_lease_fee = None
    for key in contract_data.keys():
        key_str = str(key).strip()
        # 줄바꿈, 공백 정규화
        import re
        key_normalized = re.sub(r'\s+', ' ', key_str.replace('\n', ' ').replace('\r', ' ').replace('\t', ' '))
        
        if '고정리스료' in key_normalized and ('월' in key_normalized or 'month' in key_normalized.lower()):
            fixed_lease_fee = contract_data.get(key)
            break
    
    if fixed_lease_fee is None:
        # 다른 가능한 컬럼명 시도
        fixed_lease_fee = contract_data.get('고정리스료(월)', 
                          contract_data.get('고정리스료',
                          contract_data.get('리스료', 0)))
    
    # 숫자로 변환
    try:
        if isinstance(fixed_lease_fee, str):
            fixed_lease_fee = float(fixed_lease_fee)
        elif fixed_lease_fee is None:
            fixed_lease_fee = 0
        else:
            fixed_lease_fee = float(fixed_lease_fee)
    except (ValueError, TypeError):
        print(f"    경고: 고정리스료(월) 값을 숫자로 변환할 수 없습니다: {fixed_lease_fee}")
        fixed_lease_fee = 0
    
    # 감가상각대상기간(월) 값 가져오기
    depreciation_period = None
    for key in contract_data.keys():
        key_str = str(key).strip()
        import re
        key_normalized = re.sub(r'\s+', ' ', key_str.replace('\n', ' ').replace('\r', ' ').replace('\t', ' '))
        
        if '감가상각대상기간' in key_normalized or ('감가상각' in key_normalized and '대상' in key_normalized):
            depreciation_period = contract_data.get(key)
            break
    
    if depreciation_period is None:
        # 다른 가능한 컬럼명 시도
        depreciation_period = contract_data.get('감가상각대상기간(월)',
                          contract_data.get('감가상각대상기간',
                          contract_data.get('감가상각기간(월)',
                          contract_data.get('감가상각기간', 0))))
    
    # 숫자로 변환
    try:
        if isinstance(depreciation_period, str):
            depreciation_period = int(float(depreciation_period))
        elif depreciation_period is None:
            depreciation_period = 0
        else:
            depreciation_period = int(depreciation_period)
    except (ValueError, TypeError):
        print(f"    경고: 감가상각대상기간(월) 값을 숫자로 변환할 수 없습니다: {depreciation_period}")
        depreciation_period = 0
    
    lease_fee_col_letter = get_column_letter(lease_fee_col_idx)
    receipt_count_col_letter = get_column_letter(receipt_count_col_idx)
    
    # 수령회차가 1인 라인부터 감가상각대상기간만큼 채우기
    filled_count = 0
    for row_idx in range(start_row, start_row + num_rows):
        # 수령회차 값 확인
        receipt_cell = f"{receipt_count_col_letter}{row_idx}"
        receipt_value = ws[receipt_cell].value
        
        # 수령회차가 1인 경우부터 시작
        if receipt_value == 1:
            # 감가상각대상기간만큼 채우기
            for i in range(depreciation_period):
                if row_idx + i < start_row + num_rows:
                    lease_fee_cell = f"{lease_fee_col_letter}{row_idx + i}"
                    set_cell_value(ws, lease_fee_cell, fixed_lease_fee)
                    filled_count += 1
            break  # 첫 번째 1을 찾으면 채우고 종료
    
    print(f"  ✓ 리스료 입력 완료: {filled_count}행 (고정리스료: {fixed_lease_fee}, 감가상각대상기간: {depreciation_period}개월)")


def fill_effective_date(ws, start_basis, num_rows, start_row=2):
    """
    유효일자 열을 수식으로 채우기
    
    Args:
        ws: Worksheet 객체
        start_basis (str): '월초' 또는 '월말'
        num_rows (int): 채울 행 수
        start_row (int): 시작 행 (기본값: 2)
    """
    from openpyxl.utils import get_column_letter
    from excel_utils import add_formula
    
    # 유효일자 열 찾기
    effective_date_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "유효일자":
            effective_date_col_idx = idx
            break
    
    if effective_date_col_idx is None:
        print("    경고: '유효일자' 헤더를 찾을 수 없습니다.")
        return
    
    # 리스개시일 열 찾기
    lease_start_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "리스개시일":
            lease_start_col_idx = idx
            break
    
    if lease_start_col_idx is None:
        print("    경고: '리스개시일' 헤더를 찾을 수 없습니다.")
        return
    
    effective_date_col_letter = get_column_letter(effective_date_col_idx)
    lease_start_col_letter = get_column_letter(lease_start_col_idx)
    
    if start_basis == '월초':
        # 월초 유형
        # 1. 첫 행: =EOMONTH(리스개시일 동일행, -1)+1
        first_row = start_row
        lease_start_cell = f"{lease_start_col_letter}{first_row}"
        effective_date_cell = f"{effective_date_col_letter}{first_row}"
        formula1 = f"=EOMONTH({lease_start_cell},-1)+1"
        add_formula(ws, effective_date_cell, formula1)
        
        # 2. 두 번째 행: =EOMONTH(유효일자 첫 행, 0)
        if num_rows > 1:
            second_row = start_row + 1
            first_effective_cell = f"{effective_date_col_letter}{first_row}"
            effective_date_cell2 = f"{effective_date_col_letter}{second_row}"
            formula2 = f"=EOMONTH({first_effective_cell},0)"
            add_formula(ws, effective_date_cell2, formula2)
        
        # 3. 세 번째 행 이후: =EOMONTH(유효일자 바로 윗행, 1)
        for row_idx in range(start_row + 2, start_row + num_rows):
            prev_row = row_idx - 1
            prev_effective_cell = f"{effective_date_col_letter}{prev_row}"
            effective_date_cell = f"{effective_date_col_letter}{row_idx}"
            formula = f"=EOMONTH({prev_effective_cell},1)"
            add_formula(ws, effective_date_cell, formula)
        
        print(f"  ✓ 유효일자 입력 완료: 월초 유형, {num_rows}행")
    
    else:  # 월말
        # 월말 유형
        # 1. 첫 행: =EOMONTH(리스개시일 동일행, 0)
        first_row = start_row
        lease_start_cell = f"{lease_start_col_letter}{first_row}"
        effective_date_cell = f"{effective_date_col_letter}{first_row}"
        formula1 = f"=EOMONTH({lease_start_cell},0)"
        add_formula(ws, effective_date_cell, formula1)
        
        # 2. 두 번째 행 이후: =EOMONTH(유효일자 바로 윗행, 1)
        for row_idx in range(start_row + 1, start_row + num_rows):
            prev_row = row_idx - 1
            prev_effective_cell = f"{effective_date_col_letter}{prev_row}"
            effective_date_cell = f"{effective_date_col_letter}{row_idx}"
            formula = f"=EOMONTH({prev_effective_cell},1)"
            add_formula(ws, effective_date_cell, formula)
        
        print(f"  ✓ 유효일자 입력 완료: 월말 유형, {num_rows}행")


def fill_payment_date(ws, payment_type, start_basis, num_rows, start_row=2):
    """
    지급일자 열을 수식으로 채우기
    
    Args:
        ws: Worksheet 객체
        payment_type (str): '선급' 또는 '후급'
        start_basis (str): '월초' 또는 '월말'
        num_rows (int): 채울 행 수
        start_row (int): 시작 행 (기본값: 2)
    """
    from openpyxl.utils import get_column_letter
    from excel_utils import add_formula
    
    # 지급일자 열 찾기
    payment_date_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "지급일자":
            payment_date_col_idx = idx
            break
    
    if payment_date_col_idx is None:
        print("    경고: '지급일자' 헤더를 찾을 수 없습니다.")
        return
    
    # 유효일자 열 찾기
    effective_date_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "유효일자":
            effective_date_col_idx = idx
            break
    
    if effective_date_col_idx is None:
        print("    경고: '유효일자' 헤더를 찾을 수 없습니다.")
        return
    
    payment_date_col_letter = get_column_letter(payment_date_col_idx)
    effective_date_col_letter = get_column_letter(effective_date_col_idx)
    
    if payment_type == '선급' and start_basis == '월초':
        # 선급/월초: 첫행 비우고, 두번째행부터 =EOMONTH(유효일자 바로 윗행+1,-1)+1
        for row_idx in range(start_row + 1, start_row + num_rows):
            prev_row = row_idx - 1
            prev_effective_cell = f"{effective_date_col_letter}{prev_row}"
            payment_date_cell = f"{payment_date_col_letter}{row_idx}"
            formula = f"=EOMONTH({prev_effective_cell}+1,-1)+1"
            add_formula(ws, payment_date_cell, formula)
        
        print(f"  ✓ 지급일자 입력 완료: 선급/월초 유형, {num_rows - 1}행 (첫행 제외)")
    
    elif payment_type == '선급' and start_basis == '월말':
        # 선급/월말: 첫행부터 =유효일자 동일행
        for row_idx in range(start_row, start_row + num_rows):
            effective_cell = f"{effective_date_col_letter}{row_idx}"
            payment_date_cell = f"{payment_date_col_letter}{row_idx}"
            formula = f"={effective_cell}"
            add_formula(ws, payment_date_cell, formula)
        
        print(f"  ✓ 지급일자 입력 완료: 선급/월말 유형, {num_rows}행")
    
    elif payment_type == '후급':
        # 후급/월초, 후급/월말: 첫행 비우고, 두번째행부터 =EOMONTH(유효일자 바로 윗행+1,0)
        for row_idx in range(start_row + 1, start_row + num_rows):
            prev_row = row_idx - 1
            prev_effective_cell = f"{effective_date_col_letter}{prev_row}"
            payment_date_cell = f"{payment_date_col_letter}{row_idx}"
            formula = f"=EOMONTH({prev_effective_cell}+1,0)"
            add_formula(ws, payment_date_cell, formula)
        
        print(f"  ✓ 지급일자 입력 완료: 후급 유형 ({start_basis}), {num_rows - 1}행 (첫행 제외)")


def fill_payment_year(ws, payment_type, start_basis, num_rows, start_row=2):
    """
    지급연도 열을 수식으로 채우기
    - 첫행 비울지 여부는 지급일자 열의 유형별 로직과 동일
    - 모든 유형에서 동일한 수식: =YEAR(지급일자 동일행)
    
    Args:
        ws: Worksheet 객체
        payment_type (str): '선급' 또는 '후급'
        start_basis (str): '월초' 또는 '월말'
        num_rows (int): 채울 행 수
        start_row (int): 시작 행 (기본값: 2)
    """
    from openpyxl.utils import get_column_letter
    from excel_utils import add_formula
    
    # 지급연도 열 찾기
    payment_year_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "지급연도":
            payment_year_col_idx = idx
            break
    
    if payment_year_col_idx is None:
        print("    경고: '지급연도' 헤더를 찾을 수 없습니다.")
        return
    
    # 지급일자 열 찾기
    payment_date_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "지급일자":
            payment_date_col_idx = idx
            break
    
    if payment_date_col_idx is None:
        print("    경고: '지급일자' 헤더를 찾을 수 없습니다.")
        return
    
    payment_year_col_letter = get_column_letter(payment_year_col_idx)
    payment_date_col_letter = get_column_letter(payment_date_col_idx)
    
    # 지급일자 열과 동일한 로직으로 첫행 처리
    if payment_type == '선급' and start_basis == '월초':
        # 선급/월초: 첫행 비우고, 두번째행부터 =YEAR(지급일자 동일행)
        for row_idx in range(start_row + 1, start_row + num_rows):
            payment_date_cell = f"{payment_date_col_letter}{row_idx}"
            payment_year_cell = f"{payment_year_col_letter}{row_idx}"
            formula = f"=YEAR({payment_date_cell})"
            add_formula(ws, payment_year_cell, formula)
        
        print(f"  ✓ 지급연도 입력 완료: 선급/월초 유형, {num_rows - 1}행 (첫행 제외)")
    
    elif payment_type == '선급' and start_basis == '월말':
        # 선급/월말: 첫행부터 =YEAR(지급일자 동일행)
        for row_idx in range(start_row, start_row + num_rows):
            payment_date_cell = f"{payment_date_col_letter}{row_idx}"
            payment_year_cell = f"{payment_year_col_letter}{row_idx}"
            formula = f"=YEAR({payment_date_cell})"
            add_formula(ws, payment_year_cell, formula)
        
        print(f"  ✓ 지급연도 입력 완료: 선급/월말 유형, {num_rows}행")
    
    elif payment_type == '후급':
        # 후급/월초, 후급/월말: 첫행 비우고, 두번째행부터 =YEAR(지급일자 동일행)
        for row_idx in range(start_row + 1, start_row + num_rows):
            payment_date_cell = f"{payment_date_col_letter}{row_idx}"
            payment_year_cell = f"{payment_year_col_letter}{row_idx}"
            formula = f"=YEAR({payment_date_cell})"
            add_formula(ws, payment_year_cell, formula)
        
        print(f"  ✓ 지급연도 입력 완료: 후급 유형 ({start_basis}), {num_rows - 1}행 (첫행 제외)")


def fill_fiscal_year_and_month(ws, num_rows, start_row=2):
    """
    회계연도 열과 결산월 열을 수식으로 채우기
    - 모든 유형에서 동일한 수식
    - 첫행부터 쭉 채움
    
    Args:
        ws: Worksheet 객체
        num_rows (int): 채울 행 수
        start_row (int): 시작 행 (기본값: 2)
    """
    from openpyxl.utils import get_column_letter
    from excel_utils import add_formula
    
    # 회계연도 열 찾기
    fiscal_year_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "회계연도":
            fiscal_year_col_idx = idx
            break
    
    if fiscal_year_col_idx is None:
        print("    경고: '회계연도' 헤더를 찾을 수 없습니다.")
        return
    
    # 결산월 열 찾기
    settlement_month_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "결산월":
            settlement_month_col_idx = idx
            break
    
    if settlement_month_col_idx is None:
        print("    경고: '결산월' 헤더를 찾을 수 없습니다.")
        return
    
    # 유효일자 열 찾기
    effective_date_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "유효일자":
            effective_date_col_idx = idx
            break
    
    if effective_date_col_idx is None:
        print("    경고: '유효일자' 헤더를 찾을 수 없습니다.")
        return
    
    fiscal_year_col_letter = get_column_letter(fiscal_year_col_idx)
    settlement_month_col_letter = get_column_letter(settlement_month_col_idx)
    effective_date_col_letter = get_column_letter(effective_date_col_idx)
    
    # 첫행부터 모든 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        effective_cell = f"{effective_date_col_letter}{row_idx}"
        
        # 회계연도: =YEAR(유효일자 동일행)
        fiscal_year_cell = f"{fiscal_year_col_letter}{row_idx}"
        fiscal_year_formula = f"=YEAR({effective_cell})"
        add_formula(ws, fiscal_year_cell, fiscal_year_formula)
        
        # 결산월: =MONTH(유효일자 동일행)
        settlement_month_cell = f"{settlement_month_col_letter}{row_idx}"
        settlement_month_formula = f"=MONTH({effective_cell})"
        add_formula(ws, settlement_month_cell, settlement_month_formula)
    
    print(f"  ✓ 회계연도 및 결산월 입력 완료: {num_rows}행")


def fill_lease_receivable(ws, payment_type, start_basis, contract_data, num_rows, start_row=2):
    """
    리스채권 열을 수식으로 채우기
    
    Args:
        ws: Worksheet 객체
        payment_type (str): '선급' 또는 '후급'
        start_basis (str): '월초' 또는 '월말'
        contract_data: 계약 데이터 (딕셔너리)
        num_rows (int): 채울 행 수
        start_row (int): 시작 행 (기본값: 2)
    """
    from openpyxl.utils import get_column_letter
    from excel_utils import add_formula
    
    # 리스채권 열 찾기
    lease_receivable_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "리스채권":
            lease_receivable_col_idx = idx
            break
    
    if lease_receivable_col_idx is None:
        print("    경고: '리스채권' 헤더를 찾을 수 없습니다.")
        return
    
    # 리스료 열 찾기
    lease_fee_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "리스료":
            lease_fee_col_idx = idx
            break
    
    if lease_fee_col_idx is None:
        print("    경고: '리스료' 헤더를 찾을 수 없습니다.")
        return
    
    # 할인율 열 찾기
    discount_rate_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "할인율(연)":
            discount_rate_col_idx = idx
            break
    
    if discount_rate_col_idx is None:
        print("    경고: '할인율(연)' 헤더를 찾을 수 없습니다.")
        return
    
    # 수령회차 열 찾기
    receipt_count_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "수령회차":
            receipt_count_col_idx = idx
            break
    
    if receipt_count_col_idx is None:
        print("    경고: '수령회차' 헤더를 찾을 수 없습니다.")
        return
    
    # 기간(월) 값 가져오기
    period_months = None
    for key in contract_data.keys():
        key_str = str(key).strip()
        import re
        key_normalized = re.sub(r'\s+', ' ', key_str.replace('\n', ' ').replace('\r', ' ').replace('\t', ' '))
        
        if '기간' in key_normalized and '월' in key_normalized:
            period_months = contract_data.get(key)
            break
    
    if period_months is None:
        period_months = contract_data.get('기간(월)', 0)
    
    try:
        if isinstance(period_months, str):
            period_months = int(float(period_months))
        elif period_months is None:
            period_months = 0
        else:
            period_months = int(period_months)
    except (ValueError, TypeError):
        print(f"    경고: 기간(월) 값을 숫자로 변환할 수 없습니다: {period_months}")
        period_months = 0
    
    lease_receivable_col_letter = get_column_letter(lease_receivable_col_idx)
    lease_fee_col_letter = get_column_letter(lease_fee_col_idx)
    discount_rate_col_letter = get_column_letter(discount_rate_col_idx)
    receipt_count_col_letter = get_column_letter(receipt_count_col_idx)
    
    if payment_type == '선급' and start_basis == '월초':
        # 선급/월초 유형
        # 1. 첫행: =ROUND(NPV(할인율 동일행/12, 리스료 범위),0)
        #    리스료 범위는 수령회차가 2인 것부터 마지막까지
        first_row = start_row
        
        # 수령회차가 2인 행 찾기
        receipt_2_row = None
        for row_idx in range(start_row, start_row + num_rows):
            receipt_cell = f"{receipt_count_col_letter}{row_idx}"
            receipt_value = ws[receipt_cell].value
            if receipt_value == 2:
                receipt_2_row = row_idx
                break
        
        if receipt_2_row is None:
            print("    경고: 수령회차가 2인 행을 찾을 수 없습니다.")
            return
        
        # 리스료 범위: 수령회차가 2인 행부터 마지막 행까지
        lease_fee_range_start = f"{lease_fee_col_letter}{receipt_2_row}"
        lease_fee_range_end = f"{lease_fee_col_letter}{start_row + num_rows - 1}"
        lease_fee_range = f"{lease_fee_range_start}:{lease_fee_range_end}"
        
        discount_rate_cell = f"{discount_rate_col_letter}{first_row}"
        lease_receivable_cell = f"{lease_receivable_col_letter}{first_row}"
        formula1 = f"=ROUND(NPV({discount_rate_cell}/12,{lease_fee_range}),0)"
        add_formula(ws, lease_receivable_cell, formula1)
        
        # 2. 두번째행: =ROUND(바로 윗행 리스채권*(1+할인율 동일행/12),)
        if num_rows > 1:
            second_row = start_row + 1
            prev_lease_receivable_cell = f"{lease_receivable_col_letter}{first_row}"
            discount_rate_cell2 = f"{discount_rate_col_letter}{second_row}"
            lease_receivable_cell2 = f"{lease_receivable_col_letter}{second_row}"
            formula2 = f"=ROUND({prev_lease_receivable_cell}*(1+{discount_rate_cell2}/12),)"
            add_formula(ws, lease_receivable_cell2, formula2)
        
        # 3. 세번째행부터 마지막까지: =ROUND((바로 윗행 리스채권-동일행 리스료)*(1+동일행 할인율/12),)
        for row_idx in range(start_row + 2, start_row + num_rows):
            prev_row = row_idx - 1
            prev_lease_receivable_cell = f"{lease_receivable_col_letter}{prev_row}"
            lease_fee_cell = f"{lease_fee_col_letter}{row_idx}"
            discount_rate_cell = f"{discount_rate_col_letter}{row_idx}"
            lease_receivable_cell = f"{lease_receivable_col_letter}{row_idx}"
            formula = f"=ROUND(({prev_lease_receivable_cell}-{lease_fee_cell})*(1+{discount_rate_cell}/12),)"
            add_formula(ws, lease_receivable_cell, formula)
        
        # 4. 수령회차가 기간(월)-1인 행: 바로 아래행의 리스료 열 참조
        target_receipt = period_months - 1
        for row_idx in range(start_row, start_row + num_rows):
            receipt_cell = f"{receipt_count_col_letter}{row_idx}"
            receipt_value = ws[receipt_cell].value
            if receipt_value == target_receipt:
                # 바로 아래행의 리스료 열 참조
                next_row = row_idx + 1
                if next_row < start_row + num_rows:
                    next_lease_fee_cell = f"{lease_fee_col_letter}{next_row}"
                    lease_receivable_cell = f"{lease_receivable_col_letter}{row_idx}"
                    formula = f"={next_lease_fee_cell}"
                    add_formula(ws, lease_receivable_cell, formula)
                break
        
        print(f"  ✓ 리스채권 입력 완료: 선급/월초 유형, {num_rows}행")
    
    elif payment_type == '선급' and start_basis == '월말':
        # 선급/월말 유형
        # 1. 첫행: =ROUND(NPV(할인율 동일행/12, 리스료 범위),0)
        #    리스료 범위는 수령회차가 2인 행부터 끝까지
        first_row = start_row
        
        # 수령회차가 2인 행 찾기
        receipt_2_row = None
        for row_idx in range(start_row, start_row + num_rows):
            receipt_cell = f"{receipt_count_col_letter}{row_idx}"
            receipt_value = ws[receipt_cell].value
            if receipt_value == 2:
                receipt_2_row = row_idx
                break
        
        if receipt_2_row is None:
            print("    경고: 수령회차가 2인 행을 찾을 수 없습니다.")
            return
        
        # 리스료 범위: 수령회차가 2인 행부터 마지막 행까지
        lease_fee_range_start = f"{lease_fee_col_letter}{receipt_2_row}"
        lease_fee_range_end = f"{lease_fee_col_letter}{start_row + num_rows - 1}"
        lease_fee_range = f"{lease_fee_range_start}:{lease_fee_range_end}"
        
        discount_rate_cell = f"{discount_rate_col_letter}{first_row}"
        lease_receivable_cell = f"{lease_receivable_col_letter}{first_row}"
        formula1 = f"=ROUND(NPV({discount_rate_cell}/12,{lease_fee_range}),0)"
        add_formula(ws, lease_receivable_cell, formula1)
        
        # 2. 두번째행부터 끝까지: =ROUND(바로 윗행 리스채권*(1+할인율 동일행/12)-동일행 리스료,)
        for row_idx in range(start_row + 1, start_row + num_rows):
            prev_row = row_idx - 1
            prev_lease_receivable_cell = f"{lease_receivable_col_letter}{prev_row}"
            discount_rate_cell = f"{discount_rate_col_letter}{row_idx}"
            lease_fee_cell = f"{lease_fee_col_letter}{row_idx}"
            lease_receivable_cell = f"{lease_receivable_col_letter}{row_idx}"
            formula = f"=ROUND({prev_lease_receivable_cell}*(1+{discount_rate_cell}/12)-{lease_fee_cell},)"
            add_formula(ws, lease_receivable_cell, formula)
        
        # 3. 수령회차가 기간(월)과 동일한 행: 0 값 입력
        target_receipt = period_months
        for row_idx in range(start_row, start_row + num_rows):
            receipt_cell = f"{receipt_count_col_letter}{row_idx}"
            receipt_value = ws[receipt_cell].value
            if receipt_value == target_receipt:
                lease_receivable_cell = f"{lease_receivable_col_letter}{row_idx}"
                set_cell_value(ws, lease_receivable_cell, 0)
                break
        
        print(f"  ✓ 리스채권 입력 완료: 선급/월말 유형, {num_rows}행")
    
    elif payment_type == '후급' and start_basis == '월초':
        # 후급/월초 유형
        # 1. 첫행: =ROUND(NPV(할인율 동일행/12, 리스료 범위),0)
        #    리스료 범위는 수령회차가 1인 행부터 끝까지
        first_row = start_row
        
        # 수령회차가 1인 행 찾기
        receipt_1_row = None
        for row_idx in range(start_row, start_row + num_rows):
            receipt_cell = f"{receipt_count_col_letter}{row_idx}"
            receipt_value = ws[receipt_cell].value
            if receipt_value == 1:
                receipt_1_row = row_idx
                break
        
        if receipt_1_row is None:
            print("    경고: 수령회차가 1인 행을 찾을 수 없습니다.")
            return
        
        # 리스료 범위: 수령회차가 1인 행부터 마지막 행까지
        lease_fee_range_start = f"{lease_fee_col_letter}{receipt_1_row}"
        lease_fee_range_end = f"{lease_fee_col_letter}{start_row + num_rows - 1}"
        lease_fee_range = f"{lease_fee_range_start}:{lease_fee_range_end}"
        
        discount_rate_cell = f"{discount_rate_col_letter}{first_row}"
        lease_receivable_cell = f"{lease_receivable_col_letter}{first_row}"
        formula1 = f"=ROUND(NPV({discount_rate_cell}/12,{lease_fee_range}),0)"
        add_formula(ws, lease_receivable_cell, formula1)
        
        # 2. 두번째행부터 끝까지: =ROUND(바로 윗행 리스채권*(1+할인율 동일행/12)-동일행 리스료,)
        for row_idx in range(start_row + 1, start_row + num_rows):
            prev_row = row_idx - 1
            prev_lease_receivable_cell = f"{lease_receivable_col_letter}{prev_row}"
            discount_rate_cell = f"{discount_rate_col_letter}{row_idx}"
            lease_fee_cell = f"{lease_fee_col_letter}{row_idx}"
            lease_receivable_cell = f"{lease_receivable_col_letter}{row_idx}"
            formula = f"=ROUND({prev_lease_receivable_cell}*(1+{discount_rate_cell}/12)-{lease_fee_cell},)"
            add_formula(ws, lease_receivable_cell, formula)
        
        # 3. 수령회차가 기간(월)과 동일한 행: 0 값 입력
        target_receipt = period_months
        for row_idx in range(start_row, start_row + num_rows):
            receipt_cell = f"{receipt_count_col_letter}{row_idx}"
            receipt_value = ws[receipt_cell].value
            if receipt_value == target_receipt:
                lease_receivable_cell = f"{lease_receivable_col_letter}{row_idx}"
                set_cell_value(ws, lease_receivable_cell, 0)
                break
        
        print(f"  ✓ 리스채권 입력 완료: 후급/월초 유형, {num_rows}행")
    
    elif payment_type == '후급' and start_basis == '월말':
        # 후급/월말 유형
        # 1. 첫행: =ROUND(NPV(할인율 동일행/12, 리스료 범위),0)
        #    리스료 범위는 수령회차가 1인 행부터 끝까지
        first_row = start_row
        
        # 수령회차가 1인 행 찾기
        receipt_1_row = None
        for row_idx in range(start_row, start_row + num_rows):
            receipt_cell = f"{receipt_count_col_letter}{row_idx}"
            receipt_value = ws[receipt_cell].value
            if receipt_value == 1:
                receipt_1_row = row_idx
                break
        
        if receipt_1_row is None:
            print("    경고: 수령회차가 1인 행을 찾을 수 없습니다.")
            return
        
        # 리스료 범위: 수령회차가 1인 행부터 마지막 행까지
        lease_fee_range_start = f"{lease_fee_col_letter}{receipt_1_row}"
        lease_fee_range_end = f"{lease_fee_col_letter}{start_row + num_rows - 1}"
        lease_fee_range = f"{lease_fee_range_start}:{lease_fee_range_end}"
        
        discount_rate_cell = f"{discount_rate_col_letter}{first_row}"
        lease_receivable_cell = f"{lease_receivable_col_letter}{first_row}"
        formula1 = f"=ROUND(NPV({discount_rate_cell}/12,{lease_fee_range}),0)"
        add_formula(ws, lease_receivable_cell, formula1)
        
        # 2. 두번째행부터 끝까지: =ROUND(바로 윗행 리스채권*(1+할인율 동일행/12)-동일행 리스료,)
        for row_idx in range(start_row + 1, start_row + num_rows):
            prev_row = row_idx - 1
            prev_lease_receivable_cell = f"{lease_receivable_col_letter}{prev_row}"
            discount_rate_cell = f"{discount_rate_col_letter}{row_idx}"
            lease_fee_cell = f"{lease_fee_col_letter}{row_idx}"
            lease_receivable_cell = f"{lease_receivable_col_letter}{row_idx}"
            formula = f"=ROUND({prev_lease_receivable_cell}*(1+{discount_rate_cell}/12)-{lease_fee_cell},)"
            add_formula(ws, lease_receivable_cell, formula)
        
        # 3. 수령회차가 기간(월)과 동일한 행: 0 값 입력
        target_receipt = period_months
        for row_idx in range(start_row, start_row + num_rows):
            receipt_cell = f"{receipt_count_col_letter}{row_idx}"
            receipt_value = ws[receipt_cell].value
            if receipt_value == target_receipt:
                lease_receivable_cell = f"{lease_receivable_col_letter}{row_idx}"
                set_cell_value(ws, lease_receivable_cell, 0)
                break
        
        print(f"  ✓ 리스채권 입력 완료: 후급/월말 유형, {num_rows}행")
    
    else:
        # 다른 유형은 아직 구현하지 않음
        print(f"  ⚠ 리스채권: {payment_type}/{start_basis} 유형은 아직 구현되지 않았습니다.")


def fill_interest_income(ws, payment_type, start_basis, num_rows, start_row=2):
    """
    이자수익 열 채우기 (유형별 수식)
    
    Args:
        ws: Worksheet 객체
        payment_type (str): 선급 또는 후급
        start_basis (str): 월초 또는 월말
        num_rows (int): 채울 행 수
        start_row (int): 시작 행 (기본값: 2)
    """
    from openpyxl.utils import get_column_letter
    from excel_utils import add_formula
    
    # 이자수익 열 찾기
    interest_income_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "이자수익":
            interest_income_col_idx = idx
            break
    
    if interest_income_col_idx is None:
        print("    경고: '이자수익' 헤더를 찾을 수 없습니다.")
        return
    
    # 리스채권 열 찾기
    lease_receivable_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "리스채권":
            lease_receivable_col_idx = idx
            break
    
    if lease_receivable_col_idx is None:
        print("    경고: '리스채권' 헤더를 찾을 수 없습니다.")
        return
    
    # 리스료 열 찾기
    lease_fee_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "리스료":
            lease_fee_col_idx = idx
            break
    
    if lease_fee_col_idx is None:
        print("    경고: '리스료' 헤더를 찾을 수 없습니다.")
        return
    
    interest_income_col_letter = get_column_letter(interest_income_col_idx)
    lease_receivable_col_letter = get_column_letter(lease_receivable_col_idx)
    lease_fee_col_letter = get_column_letter(lease_fee_col_idx)
    
    if payment_type == '선급' and start_basis == '월초':
        # 선급/월초 유형
        # 1. 첫행은 비움 (아무것도 하지 않음)
        
        # 2. 두번째행: =H11-H10 (동일행 리스채권 - 바로 윗행 리스채권)
        if num_rows > 1:
            second_row = start_row + 1
            first_row = start_row
            lease_receivable_curr = f"{lease_receivable_col_letter}{second_row}"
            lease_receivable_prev = f"{lease_receivable_col_letter}{first_row}"
            interest_income_cell = f"{interest_income_col_letter}{second_row}"
            formula = f"={lease_receivable_curr}-{lease_receivable_prev}"
            add_formula(ws, interest_income_cell, formula)
        
        # 3. 세번째행부터 끝까지: =H12+F12-H11 (동일행 리스채권 + 동일행 리스료 - 바로 윗행 리스채권)
        for row_idx in range(start_row + 2, start_row + num_rows):
            lease_receivable_curr = f"{lease_receivable_col_letter}{row_idx}"
            lease_fee_curr = f"{lease_fee_col_letter}{row_idx}"
            lease_receivable_prev = f"{lease_receivable_col_letter}{row_idx - 1}"
            interest_income_cell = f"{interest_income_col_letter}{row_idx}"
            formula = f"={lease_receivable_curr}+{lease_fee_curr}-{lease_receivable_prev}"
            add_formula(ws, interest_income_cell, formula)
        
        print(f"  ✓ 이자수익 입력 완료: 선급/월초 유형, {num_rows}행")
    
    elif (payment_type == '선급' and start_basis == '월말') or \
         (payment_type == '후급' and start_basis == '월초') or \
         (payment_type == '후급' and start_basis == '월말'):
        # 선급/월말, 후급/월초, 후급/월말 유형
        # 1. 첫행은 비움 (아무것도 하지 않음)
        
        # 2. 두번째행부터 끝까지: =Q11+O11-Q10 (동일행 리스채권 + 동일행 리스료 - 바로 윗행 리스채권)
        for row_idx in range(start_row + 1, start_row + num_rows):
            lease_receivable_curr = f"{lease_receivable_col_letter}{row_idx}"
            lease_fee_curr = f"{lease_fee_col_letter}{row_idx}"
            lease_receivable_prev = f"{lease_receivable_col_letter}{row_idx - 1}"
            interest_income_cell = f"{interest_income_col_letter}{row_idx}"
            formula = f"={lease_receivable_curr}+{lease_fee_curr}-{lease_receivable_prev}"
            add_formula(ws, interest_income_cell, formula)
        
        print(f"  ✓ 이자수익 입력 완료: {payment_type}/{start_basis} 유형, {num_rows}행")
    
    else:
        print(f"  ⚠ 이자수익: {payment_type}/{start_basis} 유형은 아직 구현되지 않았습니다.")


def fill_depreciation(ws, num_rows, start_row=2):
    """
    상각액 열 채우기 (모든 유형 동일)
    
    Args:
        ws: Worksheet 객체
        num_rows (int): 채울 행 수
        start_row (int): 시작 행 (기본값: 2)
    """
    from openpyxl.utils import get_column_letter
    from excel_utils import add_formula
    
    # 상각액 열 찾기
    depreciation_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "상각액":
            depreciation_col_idx = idx
            break
    
    if depreciation_col_idx is None:
        print("    경고: '상각액' 헤더를 찾을 수 없습니다.")
        return
    
    # 리스료 열 찾기
    lease_fee_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "리스료":
            lease_fee_col_idx = idx
            break
    
    if lease_fee_col_idx is None:
        print("    경고: '리스료' 헤더를 찾을 수 없습니다.")
        return
    
    # 이자수익 열 찾기
    interest_income_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "이자수익":
            interest_income_col_idx = idx
            break
    
    if interest_income_col_idx is None:
        print("    경고: '이자수익' 헤더를 찾을 수 없습니다.")
        return
    
    depreciation_col_letter = get_column_letter(depreciation_col_idx)
    lease_fee_col_letter = get_column_letter(lease_fee_col_idx)
    interest_income_col_letter = get_column_letter(interest_income_col_idx)
    
    # 두번째행부터 끝까지: 동일행의 리스료 - 동일행의 이자수익
    for row_idx in range(start_row + 1, start_row + num_rows):
        lease_fee_cell = f"{lease_fee_col_letter}{row_idx}"
        interest_income_cell = f"{interest_income_col_letter}{row_idx}"
        depreciation_cell = f"{depreciation_col_letter}{row_idx}"
        formula = f"={lease_fee_cell}-{interest_income_cell}"
        add_formula(ws, depreciation_cell, formula)
    
    print(f"  ✓ 상각액 입력 완료: {num_rows}행")


def fill_lease_receivable_current(ws, payment_type, start_basis, num_rows, start_row=2):
    """
    리스채권_유동 열 채우기 (유형별 수식)
    
    Args:
        ws: Worksheet 객체
        payment_type (str): 선급 또는 후급
        start_basis (str): 월초 또는 월말
        num_rows (int): 채울 행 수
        start_row (int): 시작 행 (기본값: 2)
    """
    from openpyxl.utils import get_column_letter
    from excel_utils import add_formula
    
    # 리스채권_유동 열 찾기
    lease_receivable_current_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "리스채권_유동":
            lease_receivable_current_col_idx = idx
            break
    
    if lease_receivable_current_col_idx is None:
        print("    경고: '리스채권_유동' 헤더를 찾을 수 없습니다.")
        return
    
    # 할인율 열 찾기
    discount_rate_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "할인율(연)":
            discount_rate_col_idx = idx
            break
    
    if discount_rate_col_idx is None:
        print("    경고: '할인율(연)' 헤더를 찾을 수 없습니다.")
        return
    
    # 리스료 열 찾기
    lease_fee_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "리스료":
            lease_fee_col_idx = idx
            break
    
    if lease_fee_col_idx is None:
        print("    경고: '리스료' 헤더를 찾을 수 없습니다.")
        return
    
    lease_receivable_current_col_letter = get_column_letter(lease_receivable_current_col_idx)
    discount_rate_col_letter = get_column_letter(discount_rate_col_idx)
    lease_fee_col_letter = get_column_letter(lease_fee_col_idx)
    
    # 마지막 행 번호 계산
    last_row = start_row + num_rows - 1
    
    if payment_type == '선급' and start_basis == '월초':
        # 선급/월초 유형
        # 1. 첫행: =ROUND(NPV(동일행 할인율/12, 바로 아래아래행부터 아래 12개행 리스료 범위),0)
        first_row = start_row
        range_start_row = first_row + 2  # 바로 아래아래행
        range_end_row = min(range_start_row + 11, last_row)  # 아래 12개행 (범위 초과 방지)
        
        if range_start_row <= last_row:
            discount_rate_cell = f"{discount_rate_col_letter}{first_row}"
            lease_fee_range_start = f"{lease_fee_col_letter}{range_start_row}"
            lease_fee_range_end = f"{lease_fee_col_letter}{range_end_row}"
            lease_fee_range = f"{lease_fee_range_start}:{lease_fee_range_end}"
            lease_receivable_current_cell = f"{lease_receivable_current_col_letter}{first_row}"
            formula = f"=ROUND(NPV({discount_rate_cell}/12,{lease_fee_range}),0)"
            add_formula(ws, lease_receivable_current_cell, formula)
        
        # 2. 두번째행부터 끝까지: =ROUND(NPV(동일행 할인율/12, 바로 아래아래행부터 아래 12개행 리스료 범위),0) + 바로 아래행 리스료
        for row_idx in range(start_row + 1, start_row + num_rows):
            range_start_row = row_idx + 2  # 바로 아래아래행
            range_end_row = min(range_start_row + 11, last_row)  # 아래 12개행 (범위 초과 방지)
            next_row_lease_fee = row_idx + 1  # 바로 아래행
            
            if range_start_row <= last_row:
                discount_rate_cell = f"{discount_rate_col_letter}{row_idx}"
                lease_fee_range_start = f"{lease_fee_col_letter}{range_start_row}"
                lease_fee_range_end = f"{lease_fee_col_letter}{range_end_row}"
                lease_fee_range = f"{lease_fee_range_start}:{lease_fee_range_end}"
                
                if next_row_lease_fee <= last_row:
                    next_lease_fee_cell = f"{lease_fee_col_letter}{next_row_lease_fee}"
                    lease_receivable_current_cell = f"{lease_receivable_current_col_letter}{row_idx}"
                    formula = f"=ROUND(NPV({discount_rate_cell}/12,{lease_fee_range}),0)+{next_lease_fee_cell}"
                    add_formula(ws, lease_receivable_current_cell, formula)
                else:
                    # 바로 아래행이 범위를 벗어나면 NPV만
                    lease_receivable_current_cell = f"{lease_receivable_current_col_letter}{row_idx}"
                    formula = f"=ROUND(NPV({discount_rate_cell}/12,{lease_fee_range}),0)"
                    add_formula(ws, lease_receivable_current_cell, formula)
        
        print(f"  ✓ 리스채권_유동 입력 완료: 선급/월초 유형, {num_rows}행")
    
    else:
        # 나머지 유형 (선급/월말, 후급/월초, 후급/월말)
        # 첫행부터 끝까지: =ROUND(NPV(동일행 할인율/12, 바로 아래행부터 아래 12개행 리스료 범위),0)
        for row_idx in range(start_row, start_row + num_rows):
            range_start_row = row_idx + 1  # 바로 아래행
            range_end_row = min(range_start_row + 11, last_row)  # 아래 12개행 (범위 초과 방지)
            
            if range_start_row <= last_row:
                discount_rate_cell = f"{discount_rate_col_letter}{row_idx}"
                lease_fee_range_start = f"{lease_fee_col_letter}{range_start_row}"
                lease_fee_range_end = f"{lease_fee_col_letter}{range_end_row}"
                lease_fee_range = f"{lease_fee_range_start}:{lease_fee_range_end}"
                lease_receivable_current_cell = f"{lease_receivable_current_col_letter}{row_idx}"
                formula = f"=ROUND(NPV({discount_rate_cell}/12,{lease_fee_range}),0)"
                add_formula(ws, lease_receivable_current_cell, formula)
        
        print(f"  ✓ 리스채권_유동 입력 완료: {payment_type}/{start_basis} 유형, {num_rows}행")


def fill_deductible_right_of_use_asset(wb, ws, contract_data, payment_type, start_basis, num_rows, start_row=2):
    """
    차감_사용권자산 열 채우기
    VLOOKUP을 사용하여 Lease_Data 시트에서 데이터를 가져옴
    - 일반 유형: 첫행부터 수령회차가 기간(월)과 같은 행까지만 채움
    - 선급/월말: 첫행부터 수령회차가 기간(월)+1과 같은 행까지 채움
    
    Args:
        wb: Workbook 객체
        ws: 현재 Worksheet 객체
        contract_data: 계약 데이터 (딕셔너리)
        payment_type (str): 선급 또는 후급
        start_basis (str): 월초 또는 월말
        num_rows (int): 전체 행 수
        start_row (int): 시작 행 (기본값: 2)
    """
    from openpyxl.utils import get_column_letter
    from excel_utils import add_formula, find_table_by_name, get_sheet_by_name, get_cell_value
    
    # 차감_사용권자산 열 찾기
    deductible_rua_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "차감_사용권자산":
            deductible_rua_col_idx = idx
            break
    
    if deductible_rua_col_idx is None:
        print("    경고: '차감_사용권자산' 헤더를 찾을 수 없습니다.")
        return
    
    # 상위리스_계약번호(Ref no.) 열 찾기
    parent_lease_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "상위리스_계약번호(Ref no.)":
            parent_lease_col_idx = idx
            break
    
    if parent_lease_col_idx is None:
        print("    경고: '상위리스_계약번호(Ref no.)' 헤더를 찾을 수 없습니다.")
        return
    
    # 전대비율 열 찾기
    sublease_ratio_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "전대비율":
            sublease_ratio_col_idx = idx
            break
    
    if sublease_ratio_col_idx is None:
        print("    경고: '전대비율' 헤더를 찾을 수 없습니다.")
        return
    
    # 수령회차 열 찾기
    receipt_count_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "수령회차":
            receipt_count_col_idx = idx
            break
    
    if receipt_count_col_idx is None:
        print("    경고: '수령회차' 헤더를 찾을 수 없습니다.")
        return
    
    # 기간(월) 값 가져오기
    period_months = None
    for key in contract_data.keys():
        key_str = str(key).strip()
        import re
        key_normalized = re.sub(r'\s+', ' ', key_str.replace('\n', ' ').replace('\r', ' ').replace('\t', ' '))
        
        if '기간' in key_normalized and '월' in key_normalized:
            period_months = contract_data.get(key)
            break
    
    if period_months is None:
        period_months = contract_data.get('기간(월)', 0)
    
    try:
        if isinstance(period_months, str):
            period_months = int(float(period_months))
        elif period_months is None:
            period_months = 0
        else:
            period_months = int(period_months)
    except (ValueError, TypeError):
        print(f"    경고: 기간(월) 값을 숫자로 변환할 수 없습니다: {period_months}")
        period_months = 0
    
    # 수령회차가 기간(월) 또는 기간(월)+1과 같은 행 찾기
    target_row = None
    receipt_count_col_letter = get_column_letter(receipt_count_col_idx)
    
    # 선급/월말 유형은 기간(월)+1까지, 나머지는 기간(월)까지
    if payment_type == '선급' and start_basis == '월말':
        target_receipt = period_months + 1
    else:
        target_receipt = period_months
    
    for row_idx in range(start_row, start_row + num_rows):
        receipt_cell = f"{receipt_count_col_letter}{row_idx}"
        receipt_value = get_cell_value(ws, receipt_cell)
        if receipt_value == target_receipt:
            target_row = row_idx
            break
    
    if target_row is None:
        print(f"    경고: 수령회차가 {target_receipt}인 행을 찾을 수 없습니다. 모든 행에 입력합니다.")
        target_row = start_row + num_rows - 1
    
    # Lease_Data 시트 확인
    lease_data_sheet = get_sheet_by_name(wb, "Lease_Data")
    if lease_data_sheet is None:
        print("    경고: 'Lease_Data' 시트를 찾을 수 없습니다.")
        return
    
    # Lease_Data 시트의 테이블 범위 찾기
    # 먼저 테이블을 찾아보고, 없으면 B7부터 BR86까지 사용
    table_range = None
    lease_data_ws, table = find_table_by_name(wb, "Lease_Data")
    if table:
        table_range = table.ref
        print(f"    ✓ Lease_Data 테이블 범위 찾음: {table_range}")
    else:
        # 테이블이 없으면 B7:BR86 사용 (사용자 지정)
        table_range = "B7:BR86"
        print(f"    ⚠ Lease_Data 테이블을 찾을 수 없어 기본 범위 사용: {table_range}")
    
    # 사용권자산 열이 몇 번째 열인지 찾기
    # 사용자가 62라고 했으므로, B열부터 62번째 열
    use_asset_col_index = 62  # 사용자 지정값
    
    # table_range를 절대참조로 변환 (B7:BR86 -> $B$7:$BR$86)
    def convert_to_absolute_ref(cell_ref):
        """셀 참조를 절대참조로 변환 (B7 -> $B$7)"""
        import re
        # 열 문자와 행 번호 분리
        match = re.match(r'([A-Z]+)(\d+)', cell_ref)
        if match:
            col = match.group(1)
            row = match.group(2)
            return f"${col}${row}"
        return cell_ref
    
    if ':' in table_range:
        start_cell, end_cell = table_range.split(':')
        table_range_absolute = f"{convert_to_absolute_ref(start_cell)}:{convert_to_absolute_ref(end_cell)}"
    else:
        table_range_absolute = convert_to_absolute_ref(table_range)
    
    deductible_rua_col_letter = get_column_letter(deductible_rua_col_idx)
    parent_lease_col_letter = get_column_letter(parent_lease_col_idx)
    sublease_ratio_col_letter = get_column_letter(sublease_ratio_col_idx)
    
    # 첫행부터 수령회차가 기간(월)과 같은 행까지만 VLOOKUP 수식 적용
    filled_rows = 0
    for row_idx in range(start_row, target_row + 1):
        # 행은 고정, 열은 상대참조 (C$2 형식)
        parent_lease_ref = f"{parent_lease_col_letter}${row_idx}"
        deductible_rua_cell = f"{deductible_rua_col_letter}{row_idx}"
        sublease_ratio_cell = f"{sublease_ratio_col_letter}{row_idx}"
        
        # VLOOKUP 수식에 전대비율 곱하기: =VLOOKUP(상위리스_계약번호, Lease_Data!범위, 62, FALSE)*전대비율
        formula = f"=VLOOKUP({parent_lease_ref},Lease_Data!{table_range_absolute}, {use_asset_col_index}, FALSE)*{sublease_ratio_cell}"
        add_formula(ws, deductible_rua_cell, formula)
        filled_rows += 1
    
    print(f"  ✓ 차감_사용권자산 입력 완료: {filled_rows}행 (수령회차 {target_receipt}까지)")


def fill_deductible_depreciation_accumulated(wb, ws, contract_data, payment_type, start_basis, num_rows, start_row=2):
    """
    차감_감가상각누계액 열 채우기
    - 일반 유형: 수령회차가 기간(월)과 같은 행까지 적용 후 차감_사용권자산 참조
    - 선급/월말: 수령회차가 기간(월)과 같은 행까지 적용 후, 기간(월)+1 행에 차감_사용권자산 참조
    
    Args:
        wb: Workbook 객체
        ws: 현재 Worksheet 객체
        contract_data: 계약 데이터 (딕셔너리)
        payment_type (str): 선급 또는 후급
        start_basis (str): 월초 또는 월말
        num_rows (int): 전체 행 수
        start_row (int): 시작 행 (기본값: 2)
    """
    from openpyxl.utils import get_column_letter
    from excel_utils import add_formula, get_cell_value
    
    # 차감_감가상각누계액 열 찾기
    deductible_dep_acc_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "차감_감가상각누계액":
            deductible_dep_acc_col_idx = idx
            break
    
    if deductible_dep_acc_col_idx is None:
        print("    경고: '차감_감가상각누계액' 헤더를 찾을 수 없습니다.")
        return
    
    # 상위리스_계약번호(Ref no.) 열 찾기
    parent_lease_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "상위리스_계약번호(Ref no.)":
            parent_lease_col_idx = idx
            break
    
    if parent_lease_col_idx is None:
        print("    경고: '상위리스_계약번호(Ref no.)' 헤더를 찾을 수 없습니다.")
        return
    
    # 유효일자 열 찾기
    effective_date_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "유효일자":
            effective_date_col_idx = idx
            break
    
    if effective_date_col_idx is None:
        print("    경고: '유효일자' 헤더를 찾을 수 없습니다.")
        return
    
    # 차감_감가상각비 열 찾기
    deductible_dep_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "차감_감가상각비":
            deductible_dep_col_idx = idx
            break
    
    if deductible_dep_col_idx is None:
        print("    경고: '차감_감가상각비' 헤더를 찾을 수 없습니다.")
        return
    
    # 차감_사용권자산 열 찾기
    deductible_rua_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "차감_사용권자산":
            deductible_rua_col_idx = idx
            break
    
    if deductible_rua_col_idx is None:
        print("    경고: '차감_사용권자산' 헤더를 찾을 수 없습니다.")
        return
    
    # 전대비율 열 찾기
    sublease_ratio_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "전대비율":
            sublease_ratio_col_idx = idx
            break
    
    if sublease_ratio_col_idx is None:
        print("    경고: '전대비율' 헤더를 찾을 수 없습니다.")
        return
    
    # 수령회차 열 찾기
    receipt_count_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "수령회차":
            receipt_count_col_idx = idx
            break
    
    if receipt_count_col_idx is None:
        print("    경고: '수령회차' 헤더를 찾을 수 없습니다.")
        return
    
    # 기간(월) 값 가져오기
    period_months = None
    for key in contract_data.keys():
        key_str = str(key).strip()
        import re
        key_normalized = re.sub(r'\s+', ' ', key_str.replace('\n', ' ').replace('\r', ' ').replace('\t', ' '))
        
        if '기간' in key_normalized and '월' in key_normalized:
            period_months = contract_data.get(key)
            break
    
    if period_months is None:
        period_months = contract_data.get('기간(월)', 0)
    
    try:
        if isinstance(period_months, str):
            period_months = int(float(period_months))
        elif period_months is None:
            period_months = 0
        else:
            period_months = int(period_months)
    except (ValueError, TypeError):
        print(f"    경고: 기간(월) 값을 숫자로 변환할 수 없습니다: {period_months}")
        period_months = 0
    
    # 현재 시트 이름 가져오기
    current_sheet_name = ws.title
    
    deductible_dep_acc_col_letter = get_column_letter(deductible_dep_acc_col_idx)
    parent_lease_col_letter = get_column_letter(parent_lease_col_idx)
    effective_date_col_letter = get_column_letter(effective_date_col_idx)
    deductible_dep_col_letter = get_column_letter(deductible_dep_col_idx)
    deductible_rua_col_letter = get_column_letter(deductible_rua_col_idx)
    receipt_count_col_letter = get_column_letter(receipt_count_col_idx)
    sublease_ratio_col_letter = get_column_letter(sublease_ratio_col_idx)
    
    # 1. 첫번째 행: IF/COUNTIFS/SUMIFS 수식에 전대비율 곱하기
    first_row = start_row
    parent_lease_cell = f"{parent_lease_col_letter}{first_row}"
    effective_date_cell = f"{effective_date_col_letter}{first_row}"
    deductible_dep_acc_cell = f"{deductible_dep_acc_col_letter}{first_row}"
    sublease_ratio_cell = f"{sublease_ratio_col_letter}{first_row}"
    
    # 수식: =IF(COUNTIFS(...), SUMIFS(...), SUMIFS(...))*전대비율
    formula1 = f"=IF(COUNTIFS(Lease_Data!B:B,{current_sheet_name}!{parent_lease_cell},Lease_Data!H:H,{current_sheet_name}!{effective_date_cell})>0,SUMIFS(Lease_Data!BJ:BJ,Lease_Data!B:B,{current_sheet_name}!{parent_lease_cell},Lease_Data!H:H,{current_sheet_name}!{effective_date_cell}),SUMIFS(Lease_Data!BJ:BJ,Lease_Data!B:B,{current_sheet_name}!{parent_lease_cell},Lease_Data!H:H,{current_sheet_name}!{effective_date_cell}-1))*{sublease_ratio_cell}"
    add_formula(ws, deductible_dep_acc_cell, formula1)
    
    # 2. 두번째 행부터 수령회차가 기간(월)과 같은 행까지: =바로 윗행 차감_감가상각누계액 + 동일행 차감_감가상각비
    # 수령회차가 기간(월)과 같은 행 찾기
    period_row = None
    for row_idx in range(start_row, start_row + num_rows):
        receipt_cell = f"{receipt_count_col_letter}{row_idx}"
        receipt_value = get_cell_value(ws, receipt_cell)
        if receipt_value == period_months:
            period_row = row_idx
            break
    
    if period_row is None:
        # 수령회차가 기간(월)과 같은 행을 찾을 수 없으면 마지막 행까지
        period_row = start_row + num_rows - 1
    
    # 선급/월말 유형: period_row까지 누적 수식 적용
    # 나머지 유형: period_row 바로 전 행까지 누적 수식 적용
    if payment_type == '선급' and start_basis == '월말':
        # 선급/월말: period_row까지 포함
        for row_idx in range(start_row + 1, period_row + 1):
            prev_row = row_idx - 1
            prev_deductible_dep_acc_cell = f"{deductible_dep_acc_col_letter}{prev_row}"
            deductible_dep_cell = f"{deductible_dep_col_letter}{row_idx}"
            deductible_dep_acc_cell = f"{deductible_dep_acc_col_letter}{row_idx}"
            
            formula2 = f"={prev_deductible_dep_acc_cell}+{deductible_dep_cell}"
            add_formula(ws, deductible_dep_acc_cell, formula2)
        
        # 수령회차가 기간(월)+1과 같은 행 찾기
        period_plus_one_row = None
        for row_idx in range(start_row, start_row + num_rows):
            receipt_cell = f"{receipt_count_col_letter}{row_idx}"
            receipt_value = get_cell_value(ws, receipt_cell)
            if receipt_value == period_months + 1:
                period_plus_one_row = row_idx
                break
        
        # 3. 수령회차가 기간(월)+1과 같은 행: 동일행의 차감_사용권자산 참조
        if period_plus_one_row is not None:
            deductible_rua_cell = f"{deductible_rua_col_letter}{period_plus_one_row}"
            deductible_dep_acc_cell = f"{deductible_dep_acc_col_letter}{period_plus_one_row}"
            formula3 = f"={deductible_rua_cell}"
            add_formula(ws, deductible_dep_acc_cell, formula3)
            filled_rows = period_plus_one_row - start_row + 1
        else:
            filled_rows = period_row - start_row + 1
    else:
        # 나머지 유형: period_row 바로 전 행까지
        for row_idx in range(start_row + 1, period_row):
            prev_row = row_idx - 1
            prev_deductible_dep_acc_cell = f"{deductible_dep_acc_col_letter}{prev_row}"
            deductible_dep_cell = f"{deductible_dep_col_letter}{row_idx}"
            deductible_dep_acc_cell = f"{deductible_dep_acc_col_letter}{row_idx}"
            
            formula2 = f"={prev_deductible_dep_acc_cell}+{deductible_dep_cell}"
            add_formula(ws, deductible_dep_acc_cell, formula2)
        
        # 3. 수령회차가 기간(월)과 같은 행: 동일행의 차감_사용권자산 참조
        if period_row is not None:
            deductible_rua_cell = f"{deductible_rua_col_letter}{period_row}"
            deductible_dep_acc_cell = f"{deductible_dep_acc_col_letter}{period_row}"
            formula3 = f"={deductible_rua_cell}"
            add_formula(ws, deductible_dep_acc_cell, formula3)
        
        filled_rows = period_row - start_row + 1 if period_row else num_rows
    
    print(f"  ✓ 차감_감가상각누계액 입력 완료: {filled_rows}행")


def fill_deductible_depreciation(wb, ws, contract_data, payment_type, start_basis, num_rows, start_row=2):
    """
    차감_감가상각비 열 채우기
    - 일반 유형: 수령회차가 기간(월)과 같은 행까지 적용 후 차감 수식
    - 선급/월말: 수령회차가 기간(월)과 같은 행까지 적용 후, 기간(월)+1 행에 차감 수식
    
    Args:
        wb: Workbook 객체
        ws: 현재 Worksheet 객체
        contract_data: 계약 데이터 (딕셔너리)
        payment_type (str): 선급 또는 후급
        start_basis (str): 월초 또는 월말
        num_rows (int): 전체 행 수
        start_row (int): 시작 행 (기본값: 2)
    """
    from openpyxl.utils import get_column_letter
    from excel_utils import add_formula, get_cell_value
    
    # 차감_감가상각비 열 찾기
    deductible_dep_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "차감_감가상각비":
            deductible_dep_col_idx = idx
            break
    
    if deductible_dep_col_idx is None:
        print("    경고: '차감_감가상각비' 헤더를 찾을 수 없습니다.")
        return
    
    # 차감_사용권자산 열 찾기
    deductible_rua_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "차감_사용권자산":
            deductible_rua_col_idx = idx
            break
    
    if deductible_rua_col_idx is None:
        print("    경고: '차감_사용권자산' 헤더를 찾을 수 없습니다.")
        return
    
    # 차감_감가상각누계액 열 찾기
    deductible_dep_acc_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "차감_감가상각누계액":
            deductible_dep_acc_col_idx = idx
            break
    
    if deductible_dep_acc_col_idx is None:
        print("    경고: '차감_감가상각누계액' 헤더를 찾을 수 없습니다.")
        return
    
    # 감가상각기간(월) 열 찾기
    depreciation_period_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "감가상각기간(월)":
            depreciation_period_col_idx = idx
            break
    
    if depreciation_period_col_idx is None:
        print("    경고: '감가상각기간(월)' 헤더를 찾을 수 없습니다.")
        return
    
    # 수령회차 열 찾기
    receipt_count_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "수령회차":
            receipt_count_col_idx = idx
            break
    
    if receipt_count_col_idx is None:
        print("    경고: '수령회차' 헤더를 찾을 수 없습니다.")
        return
    
    # 기간(월) 값 가져오기
    period_months = None
    for key in contract_data.keys():
        key_str = str(key).strip()
        import re
        key_normalized = re.sub(r'\s+', ' ', key_str.replace('\n', ' ').replace('\r', ' ').replace('\t', ' '))
        
        if '기간' in key_normalized and '월' in key_normalized:
            period_months = contract_data.get(key)
            break
    
    if period_months is None:
        period_months = contract_data.get('기간(월)', 0)
    
    try:
        if isinstance(period_months, str):
            period_months = int(float(period_months))
        elif period_months is None:
            period_months = 0
        else:
            period_months = int(period_months)
    except (ValueError, TypeError):
        print(f"    경고: 기간(월) 값을 숫자로 변환할 수 없습니다: {period_months}")
        period_months = 0
    
    deductible_dep_col_letter = get_column_letter(deductible_dep_col_idx)
    deductible_rua_col_letter = get_column_letter(deductible_rua_col_idx)
    deductible_dep_acc_col_letter = get_column_letter(deductible_dep_acc_col_idx)
    depreciation_period_col_letter = get_column_letter(depreciation_period_col_idx)
    receipt_count_col_letter = get_column_letter(receipt_count_col_idx)
    
    # 첫행의 절대참조 셀
    first_row = start_row
    deductible_rua_first_cell = f"${deductible_rua_col_letter}${first_row}"
    deductible_dep_acc_first_cell = f"${deductible_dep_acc_col_letter}${first_row}"
    
    # 수령회차가 기간(월)과 같은 행 찾기
    period_row = None
    for row_idx in range(start_row, start_row + num_rows):
        receipt_cell = f"{receipt_count_col_letter}{row_idx}"
        receipt_value = get_cell_value(ws, receipt_cell)
        if receipt_value == period_months:
            period_row = row_idx
            break
    
    if period_row is None:
        # 수령회차가 기간(월)과 같은 행을 찾을 수 없으면 마지막 행까지
        period_row = start_row + num_rows - 1
    
    # 1. 첫행은 비움 (아무것도 하지 않음)
    
    # 선급/월말 유형 처리
    if payment_type == '선급' and start_basis == '월말':
        # 2. 두번째행부터 수령회차가 기간(월)과 같은 행까지: =ROUND(($V$2-$X$2)/L3,0)
        for row_idx in range(start_row + 1, period_row + 1):
            depreciation_period_cell = f"{depreciation_period_col_letter}{row_idx}"
            deductible_dep_cell = f"{deductible_dep_col_letter}{row_idx}"
            
            formula = f"=ROUND(({deductible_rua_first_cell}-{deductible_dep_acc_first_cell})/{depreciation_period_cell},0)"
            add_formula(ws, deductible_dep_cell, formula)
        
        # 수령회차가 기간(월)+1과 같은 행 찾기
        period_plus_one_row = None
        for row_idx in range(start_row, start_row + num_rows):
            receipt_cell = f"{receipt_count_col_letter}{row_idx}"
            receipt_value = get_cell_value(ws, receipt_cell)
            if receipt_value == period_months + 1:
                period_plus_one_row = row_idx
                break
        
        # 3. 수령회차가 기간(월)+1과 같은 행: 동일행의 차감_감가상각누계액 - 바로 윗행의 차감_감가상각누계액
        if period_plus_one_row is not None and period_plus_one_row > start_row:
            prev_row = period_plus_one_row - 1
            deductible_dep_acc_curr_cell = f"{deductible_dep_acc_col_letter}{period_plus_one_row}"
            deductible_dep_acc_prev_cell = f"{deductible_dep_acc_col_letter}{prev_row}"
            deductible_dep_cell = f"{deductible_dep_col_letter}{period_plus_one_row}"
            
            formula = f"={deductible_dep_acc_curr_cell}-{deductible_dep_acc_prev_cell}"
            add_formula(ws, deductible_dep_cell, formula)
            filled_rows = period_plus_one_row - start_row
        else:
            filled_rows = period_row - start_row
    else:
        # 나머지 유형: 기존 로직
        # 2. 두번째행부터 수령회차가 기간(월)과 같은 행의 바로 전행까지: =ROUND(($V$2-$X$2)/L3,0)
        for row_idx in range(start_row + 1, period_row):
            depreciation_period_cell = f"{depreciation_period_col_letter}{row_idx}"
            deductible_dep_cell = f"{deductible_dep_col_letter}{row_idx}"
            
            formula = f"=ROUND(({deductible_rua_first_cell}-{deductible_dep_acc_first_cell})/{depreciation_period_cell},0)"
            add_formula(ws, deductible_dep_cell, formula)
        
        # 3. 수령회차가 기간(월)과 같은 행: 동일행의 차감_감가상각누계액 - 바로 윗행의 차감_감가상각누계액
        if period_row is not None and period_row > start_row:
            prev_row = period_row - 1
            deductible_dep_acc_curr_cell = f"{deductible_dep_acc_col_letter}{period_row}"
            deductible_dep_acc_prev_cell = f"{deductible_dep_acc_col_letter}{prev_row}"
            deductible_dep_cell = f"{deductible_dep_col_letter}{period_row}"
            
            formula = f"={deductible_dep_acc_curr_cell}-{deductible_dep_acc_prev_cell}"
            add_formula(ws, deductible_dep_cell, formula)
        
        filled_rows = period_row - start_row if period_row else num_rows - 1
    
    print(f"  ✓ 차감_감가상각비 입력 완료: {filled_rows}행 (수령회차 {period_months}까지)")


def fill_deductible_right_of_use_asset_net(ws, num_rows, start_row=2):
    """
    차감_사용권자산(순) 열 채우기 (모든 유형 동일)
    동일행의 차감_사용권자산 - 동일행의 차감_감가상각누계액
    
    Args:
        ws: 현재 Worksheet 객체
        num_rows (int): 전체 행 수
        start_row (int): 시작 행 (기본값: 2)
    """
    from openpyxl.utils import get_column_letter
    from excel_utils import add_formula
    
    # 차감_사용권자산(순) 열 찾기
    deductible_rua_net_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "차감_사용권자산(순)":
            deductible_rua_net_col_idx = idx
            break
    
    if deductible_rua_net_col_idx is None:
        print("    경고: '차감_사용권자산(순)' 헤더를 찾을 수 없습니다.")
        return
    
    # 차감_사용권자산 열 찾기
    deductible_rua_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "차감_사용권자산":
            deductible_rua_col_idx = idx
            break
    
    if deductible_rua_col_idx is None:
        print("    경고: '차감_사용권자산' 헤더를 찾을 수 없습니다.")
        return
    
    # 차감_감가상각누계액 열 찾기
    deductible_dep_acc_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "차감_감가상각누계액":
            deductible_dep_acc_col_idx = idx
            break
    
    if deductible_dep_acc_col_idx is None:
        print("    경고: '차감_감가상각누계액' 헤더를 찾을 수 없습니다.")
        return
    
    deductible_rua_net_col_letter = get_column_letter(deductible_rua_net_col_idx)
    deductible_rua_col_letter = get_column_letter(deductible_rua_col_idx)
    deductible_dep_acc_col_letter = get_column_letter(deductible_dep_acc_col_idx)
    
    # 전체 대상 행에 수식 적용: 동일행의 차감_사용권자산 - 동일행의 차감_감가상각누계액
    for row_idx in range(start_row, start_row + num_rows):
        deductible_rua_cell = f"{deductible_rua_col_letter}{row_idx}"
        deductible_dep_acc_cell = f"{deductible_dep_acc_col_letter}{row_idx}"
        deductible_rua_net_cell = f"{deductible_rua_net_col_letter}{row_idx}"
        
        formula = f"={deductible_rua_cell}-{deductible_dep_acc_cell}"
        add_formula(ws, deductible_rua_net_cell, formula)
    
    print(f"  ✓ 차감_사용권자산(순) 입력 완료: {num_rows}행")


def fill_lease_deposit(wb, ws, contract_data, contract_no, payment_type, start_basis, num_rows, start_row=2):
    """
    임대보증금 열 채우기
    - 선급/월말: 수령회차가 기간(월)+1인 행까지
    - 나머지 유형: 수령회차가 기간(월)과 같은 행까지
    
    Args:
        wb: Workbook 객체
        ws: 현재 Worksheet 객체
        contract_data: 계약 데이터 (딕셔너리)
        contract_no: 계약번호
        payment_type (str): 선급 또는 후급
        start_basis (str): 월초 또는 월말
        num_rows (int): 전체 행 수
        start_row (int): 시작 행 (기본값: 2)
    """
    from openpyxl.utils import get_column_letter
    from excel_utils import add_formula, get_cell_value, get_sheet_by_name, get_table_data
    
    # 임대보증금 열 찾기
    lease_deposit_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "임대보증금":
            lease_deposit_col_idx = idx
            break
    
    if lease_deposit_col_idx is None:
        print("    경고: '임대보증금' 헤더를 찾을 수 없습니다.")
        return
    
    # 임대보증금할인율(연) 열 찾기
    deposit_discount_rate_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "임대보증금할인율(연)":
            deposit_discount_rate_col_idx = idx
            break
    
    if deposit_discount_rate_col_idx is None:
        print("    경고: '임대보증금할인율(연)' 헤더를 찾을 수 없습니다.")
        return
    
    # 기간(월) 열 찾기
    period_months_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "기간(월)":
            period_months_col_idx = idx
            break
    
    if period_months_col_idx is None:
        print("    경고: '기간(월)' 헤더를 찾을 수 없습니다.")
        return
    
    # 수령회차 열 찾기
    receipt_count_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "수령회차":
            receipt_count_col_idx = idx
            break
    
    if receipt_count_col_idx is None:
        print("    경고: '수령회차' 헤더를 찾을 수 없습니다.")
        return
    
    # 기간(월) 값 가져오기
    period_months = None
    for key in contract_data.keys():
        key_str = str(key).strip()
        import re
        key_normalized = re.sub(r'\s+', ' ', key_str.replace('\n', ' ').replace('\r', ' ').replace('\t', ' '))
        
        if '기간' in key_normalized and '월' in key_normalized:
            period_months = contract_data.get(key)
            break
    
    if period_months is None:
        period_months = contract_data.get('기간(월)', 0)
    
    try:
        if isinstance(period_months, str):
            period_months = int(float(period_months))
        elif period_months is None:
            period_months = 0
        else:
            period_months = int(period_months)
    except (ValueError, TypeError):
        print(f"    경고: 기간(월) 값을 숫자로 변환할 수 없습니다: {period_months}")
        period_months = 0
    
    # Input_data 시트에서 해당 계약의 임차보증금 셀 찾기
    input_data_sheet = get_sheet_by_name(wb, "Input_data")
    if input_data_sheet is None:
        print("    경고: 'Input_data' 시트를 찾을 수 없습니다.")
        return
    
    # input_table에서 계약번호로 해당 행 찾기
    try:
        table_data = get_table_data(input_data_sheet, "input_table")
    except Exception as e:
        print(f"    경고: input_table 데이터 읽기 실패: {e}")
        return
    
    # 계약번호 컬럼 찾기
    ref_no_col = None
    for col_name in table_data[0].keys():
        col_lower = str(col_name).lower().replace(' ', '').replace('_', '').replace('.', '')
        if 'refno' in col_lower or '계약번호' in str(col_name):
            ref_no_col = col_name
            break
    
    if ref_no_col is None:
        print("    경고: Input_data에서 계약번호 컬럼을 찾을 수 없습니다.")
        return
    
    # 임차보증금 컬럼 찾기
    deposit_col = None
    for col_name in table_data[0].keys():
        col_str = str(col_name).strip()
        if '임차보증금' in col_str or '보증금' in col_str:
            deposit_col = col_name
            break
    
    if deposit_col is None:
        print("    경고: Input_data에서 임차보증금 컬럼을 찾을 수 없습니다.")
        return
    
    # 계약번호로 해당 행 찾기
    deposit_cell_ref = None
    for row_idx, row_data in enumerate(table_data, start=1):
        if str(row_data.get(ref_no_col)) == str(contract_no):
            # 테이블의 헤더 행 다음부터 시작하므로 +1
            # input_table의 시작 행을 찾아야 함
            try:
                from excel_utils import find_table_by_name
                input_ws, table = find_table_by_name(wb, "input_table")
                if table:
                    # 테이블 범위 파싱
                    ref = table.ref
                    start_cell, end_cell = ref.split(':')
                    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
                    start_col_str, start_row_num = coordinate_from_string(start_cell)
                    # 헤더 행 다음이 데이터 행이므로 +1, enumerate는 1부터 시작하므로 row_idx 그대로 사용
                    data_row = start_row_num + row_idx
                    
                    # 임차보증금 열 찾기
                    deposit_col_idx_in_table = None
                    headers = list(table_data[0].keys())
                    if deposit_col in headers:
                        deposit_col_idx_in_table = headers.index(deposit_col)
                        deposit_col_letter = get_column_letter(column_index_from_string(start_col_str) + deposit_col_idx_in_table)
                        deposit_cell_ref = f"Input_data!${deposit_col_letter}${data_row}"
                        break
            except Exception as e:
                print(f"    경고: 임차보증금 셀 참조 생성 실패: {e}")
                # 대체 방법: VLOOKUP 사용
                break
    
    if deposit_cell_ref is None:
        print("    경고: Input_data에서 해당 계약의 임차보증금 셀을 찾을 수 없습니다.")
        return
    
    lease_deposit_col_letter = get_column_letter(lease_deposit_col_idx)
    deposit_discount_rate_col_letter = get_column_letter(deposit_discount_rate_col_idx)
    period_months_col_letter = get_column_letter(period_months_col_idx)
    receipt_count_col_letter = get_column_letter(receipt_count_col_idx)
    
    # 선급/월말 유형 처리
    if payment_type == '선급' and start_basis == '월말':
        # 수령회차가 기간(월)+1인 행 찾기
        target_row = None
        target_receipt = period_months + 1
        for row_idx in range(start_row, start_row + num_rows):
            receipt_cell = f"{receipt_count_col_letter}{row_idx}"
            receipt_value = get_cell_value(ws, receipt_cell)
            if receipt_value == target_receipt:
                target_row = row_idx
                break
        
        if target_row is None:
            print(f"    경고: 수령회차가 {target_receipt}인 행을 찾을 수 없습니다.")
            target_row = start_row + num_rows - 1
        
        # 수령회차가 기간(월)+1인 행까지 수식 적용
        for row_idx in range(start_row, target_row + 1):
            deposit_discount_rate_cell = f"{deposit_discount_rate_col_letter}{row_idx}"
            period_months_cell = f"{period_months_col_letter}{row_idx}"
            receipt_count_cell = f"{receipt_count_col_letter}{row_idx}"
            lease_deposit_cell = f"{lease_deposit_col_letter}{row_idx}"
            
            # =ROUND((Input_data!$W$14/(1+P2/12)^(K2-M2+1)),0)
            formula = f"=ROUND(({deposit_cell_ref}/(1+{deposit_discount_rate_cell}/12)^({period_months_cell}-{receipt_count_cell}+1)),0)"
            add_formula(ws, lease_deposit_cell, formula)
        
        filled_rows = target_row - start_row + 1
    else:
        # 나머지 유형: 수령회차가 기간(월)과 같은 행까지
        target_row = None
        for row_idx in range(start_row, start_row + num_rows):
            receipt_cell = f"{receipt_count_col_letter}{row_idx}"
            receipt_value = get_cell_value(ws, receipt_cell)
            if receipt_value == period_months:
                target_row = row_idx
                break
        
        if target_row is None:
            print(f"    경고: 수령회차가 {period_months}인 행을 찾을 수 없습니다.")
            target_row = start_row + num_rows - 1
        
        # 수령회차가 기간(월)과 같은 행까지 수식 적용
        for row_idx in range(start_row, target_row + 1):
            deposit_discount_rate_cell = f"{deposit_discount_rate_col_letter}{row_idx}"
            period_months_cell = f"{period_months_col_letter}{row_idx}"
            receipt_count_cell = f"{receipt_count_col_letter}{row_idx}"
            lease_deposit_cell = f"{lease_deposit_col_letter}{row_idx}"
            
            # =ROUND((Input_data!$W$14/(1+P2/12)^(K2-M2)),0)
            formula = f"=ROUND(({deposit_cell_ref}/(1+{deposit_discount_rate_cell}/12)^({period_months_cell}-{receipt_count_cell})),0)"
            add_formula(ws, lease_deposit_cell, formula)
        
        filled_rows = target_row - start_row + 1
    
    print(f"  ✓ 임대보증금 입력 완료: {filled_rows}행")


def fill_lease_deposit_present_value(wb, ws, contract_data, contract_no, payment_type, start_basis, num_rows, start_row=2):
    """
    임대보증금_현할차 열 채우기
    - 선급/월말: 수령회차가 기간(월)+1인 행까지
    - 나머지 유형: 수령회차가 기간(월)과 같은 행까지
    
    Args:
        wb: Workbook 객체
        ws: 현재 Worksheet 객체
        contract_data: 계약 데이터 (딕셔너리)
        contract_no: 계약번호
        payment_type (str): 선급 또는 후급
        start_basis (str): 월초 또는 월말
        num_rows (int): 전체 행 수
        start_row (int): 시작 행 (기본값: 2)
    """
    from openpyxl.utils import get_column_letter
    from excel_utils import add_formula, get_cell_value, get_sheet_by_name, get_table_data
    
    # 임대보증금_현할차 열 찾기
    deposit_pv_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "임대보증금_현할차":
            deposit_pv_col_idx = idx
            break
    
    if deposit_pv_col_idx is None:
        print("    경고: '임대보증금_현할차' 헤더를 찾을 수 없습니다.")
        return
    
    # 임대보증금 열 찾기
    lease_deposit_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "임대보증금":
            lease_deposit_col_idx = idx
            break
    
    if lease_deposit_col_idx is None:
        print("    경고: '임대보증금' 헤더를 찾을 수 없습니다.")
        return
    
    # 수령회차 열 찾기
    receipt_count_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "수령회차":
            receipt_count_col_idx = idx
            break
    
    if receipt_count_col_idx is None:
        print("    경고: '수령회차' 헤더를 찾을 수 없습니다.")
        return
    
    # 기간(월) 값 가져오기
    period_months = None
    for key in contract_data.keys():
        key_str = str(key).strip()
        import re
        key_normalized = re.sub(r'\s+', ' ', key_str.replace('\n', ' ').replace('\r', ' ').replace('\t', ' '))
        
        if '기간' in key_normalized and '월' in key_normalized:
            period_months = contract_data.get(key)
            break
    
    if period_months is None:
        period_months = contract_data.get('기간(월)', 0)
    
    try:
        if isinstance(period_months, str):
            period_months = int(float(period_months))
        elif period_months is None:
            period_months = 0
        else:
            period_months = int(period_months)
    except (ValueError, TypeError):
        print(f"    경고: 기간(월) 값을 숫자로 변환할 수 없습니다: {period_months}")
        period_months = 0
    
    # Input_data 시트에서 해당 계약의 임차보증금 셀 찾기 (임대보증금 함수와 동일한 로직)
    input_data_sheet = get_sheet_by_name(wb, "Input_data")
    if input_data_sheet is None:
        print("    경고: 'Input_data' 시트를 찾을 수 없습니다.")
        return
    
    try:
        table_data = get_table_data(input_data_sheet, "input_table")
    except Exception as e:
        print(f"    경고: input_table 데이터 읽기 실패: {e}")
        return
    
    # 계약번호 컬럼 찾기
    ref_no_col = None
    for col_name in table_data[0].keys():
        col_lower = str(col_name).lower().replace(' ', '').replace('_', '').replace('.', '')
        if 'refno' in col_lower or '계약번호' in str(col_name):
            ref_no_col = col_name
            break
    
    if ref_no_col is None:
        print("    경고: Input_data에서 계약번호 컬럼을 찾을 수 없습니다.")
        return
    
    # 임차보증금 컬럼 찾기
    deposit_col = None
    for col_name in table_data[0].keys():
        col_str = str(col_name).strip()
        if '임차보증금' in col_str or '보증금' in col_str:
            deposit_col = col_name
            break
    
    if deposit_col is None:
        print("    경고: Input_data에서 임차보증금 컬럼을 찾을 수 없습니다.")
        return
    
    # 계약번호로 해당 행 찾기
    deposit_cell_ref = None
    for row_idx, row_data in enumerate(table_data, start=1):
        if str(row_data.get(ref_no_col)) == str(contract_no):
            try:
                from excel_utils import find_table_by_name
                input_ws, table = find_table_by_name(wb, "input_table")
                if table:
                    ref = table.ref
                    start_cell, end_cell = ref.split(':')
                    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
                    start_col_str, start_row_num = coordinate_from_string(start_cell)
                    data_row = start_row_num + row_idx
                    
                    headers = list(table_data[0].keys())
                    if deposit_col in headers:
                        deposit_col_idx_in_table = headers.index(deposit_col)
                        deposit_col_letter = get_column_letter(column_index_from_string(start_col_str) + deposit_col_idx_in_table)
                        deposit_cell_ref = f"Input_data!${deposit_col_letter}${data_row}"
                        break
            except Exception as e:
                print(f"    경고: 임차보증금 셀 참조 생성 실패: {e}")
                break
    
    if deposit_cell_ref is None:
        print("    경고: Input_data에서 해당 계약의 임차보증금 셀을 찾을 수 없습니다.")
        return
    
    deposit_pv_col_letter = get_column_letter(deposit_pv_col_idx)
    lease_deposit_col_letter = get_column_letter(lease_deposit_col_idx)
    receipt_count_col_letter = get_column_letter(receipt_count_col_idx)
    
    # 선급/월말 유형 처리
    if payment_type == '선급' and start_basis == '월말':
        # 수령회차가 기간(월)+1인 행 찾기
        target_row = None
        target_receipt = period_months + 1
        for row_idx in range(start_row, start_row + num_rows):
            receipt_cell = f"{receipt_count_col_letter}{row_idx}"
            receipt_value = get_cell_value(ws, receipt_cell)
            if receipt_value == target_receipt:
                target_row = row_idx
                break
        
        if target_row is None:
            print(f"    경고: 수령회차가 {target_receipt}인 행을 찾을 수 없습니다.")
            target_row = start_row + num_rows - 1
        
        # 수령회차가 기간(월)+1인 행까지 수식 적용
        for row_idx in range(start_row, target_row + 1):
            lease_deposit_cell = f"{lease_deposit_col_letter}{row_idx}"
            deposit_pv_cell = f"{deposit_pv_col_letter}{row_idx}"
            
            # =Input_data!$W$14-Z2
            formula = f"={deposit_cell_ref}-{lease_deposit_cell}"
            add_formula(ws, deposit_pv_cell, formula)
        
        filled_rows = target_row - start_row + 1
    else:
        # 나머지 유형: 수령회차가 기간(월)과 같은 행까지
        target_row = None
        for row_idx in range(start_row, start_row + num_rows):
            receipt_cell = f"{receipt_count_col_letter}{row_idx}"
            receipt_value = get_cell_value(ws, receipt_cell)
            if receipt_value == period_months:
                target_row = row_idx
                break
        
        if target_row is None:
            print(f"    경고: 수령회차가 {period_months}인 행을 찾을 수 없습니다.")
            target_row = start_row + num_rows - 1
        
        # 수령회차가 기간(월)과 같은 행까지 수식 적용
        for row_idx in range(start_row, target_row + 1):
            lease_deposit_cell = f"{lease_deposit_col_letter}{row_idx}"
            deposit_pv_cell = f"{deposit_pv_col_letter}{row_idx}"
            
            # =Input_data!$W$14-Z2
            formula = f"={deposit_cell_ref}-{lease_deposit_cell}"
            add_formula(ws, deposit_pv_cell, formula)
        
        filled_rows = target_row - start_row + 1
    
    print(f"  ✓ 임대보증금_현할차 입력 완료: {filled_rows}행")


def fill_lease_deposit_interest(ws, contract_data, payment_type, start_basis, num_rows, start_row=2):
    """
    임대보증금이자 열 채우기
    - 선급/월말: 수령회차가 기간(월)+1인 행까지
    - 나머지 유형: 수령회차가 기간(월)과 같은 행까지
    
    Args:
        ws: 현재 Worksheet 객체
        contract_data: 계약 데이터 (딕셔너리)
        payment_type (str): 선급 또는 후급
        start_basis (str): 월초 또는 월말
        num_rows (int): 전체 행 수
        start_row (int): 시작 행 (기본값: 2)
    """
    from openpyxl.utils import get_column_letter
    from excel_utils import add_formula, get_cell_value
    
    # 임대보증금이자 열 찾기
    deposit_interest_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "임대보증금이자":
            deposit_interest_col_idx = idx
            break
    
    if deposit_interest_col_idx is None:
        print("    경고: '임대보증금이자' 헤더를 찾을 수 없습니다.")
        return
    
    # 임대보증금_현할차 열 찾기
    deposit_pv_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "임대보증금_현할차":
            deposit_pv_col_idx = idx
            break
    
    if deposit_pv_col_idx is None:
        print("    경고: '임대보증금_현할차' 헤더를 찾을 수 없습니다.")
        return
    
    # 수령회차 열 찾기
    receipt_count_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "수령회차":
            receipt_count_col_idx = idx
            break
    
    if receipt_count_col_idx is None:
        print("    경고: '수령회차' 헤더를 찾을 수 없습니다.")
        return
    
    # 기간(월) 값 가져오기
    period_months = None
    for key in contract_data.keys():
        key_str = str(key).strip()
        import re
        key_normalized = re.sub(r'\s+', ' ', key_str.replace('\n', ' ').replace('\r', ' ').replace('\t', ' '))
        
        if '기간' in key_normalized and '월' in key_normalized:
            period_months = contract_data.get(key)
            break
    
    if period_months is None:
        period_months = contract_data.get('기간(월)', 0)
    
    try:
        if isinstance(period_months, str):
            period_months = int(float(period_months))
        elif period_months is None:
            period_months = 0
        else:
            period_months = int(period_months)
    except (ValueError, TypeError):
        print(f"    경고: 기간(월) 값을 숫자로 변환할 수 없습니다: {period_months}")
        period_months = 0
    
    deposit_interest_col_letter = get_column_letter(deposit_interest_col_idx)
    deposit_pv_col_letter = get_column_letter(deposit_pv_col_idx)
    receipt_count_col_letter = get_column_letter(receipt_count_col_idx)
    
    # 선급/월말 유형 처리
    if payment_type == '선급' and start_basis == '월말':
        # 수령회차가 기간(월)+1인 행 찾기
        target_row = None
        target_receipt = period_months + 1
        for row_idx in range(start_row, start_row + num_rows):
            receipt_cell = f"{receipt_count_col_letter}{row_idx}"
            receipt_value = get_cell_value(ws, receipt_cell)
            if receipt_value == target_receipt:
                target_row = row_idx
                break
        
        if target_row is None:
            print(f"    경고: 수령회차가 {target_receipt}인 행을 찾을 수 없습니다.")
            target_row = start_row + num_rows - 1
        
        # 두번째 행부터 수령회차가 기간(월)+1인 행까지 수식 적용
        for row_idx in range(start_row + 1, target_row + 1):
            prev_row = row_idx - 1
            prev_deposit_pv_cell = f"{deposit_pv_col_letter}{prev_row}"
            deposit_pv_cell = f"{deposit_pv_col_letter}{row_idx}"
            deposit_interest_cell = f"{deposit_interest_col_letter}{row_idx}"
            
            # =AA2-AA3 (바로 윗행 임대보증금_현할차 - 동일행 임대보증금_현할차)
            formula = f"={prev_deposit_pv_cell}-{deposit_pv_cell}"
            add_formula(ws, deposit_interest_cell, formula)
        
        filled_rows = target_row - start_row
    else:
        # 나머지 유형: 수령회차가 기간(월)과 같은 행까지
        target_row = None
        for row_idx in range(start_row, start_row + num_rows):
            receipt_cell = f"{receipt_count_col_letter}{row_idx}"
            receipt_value = get_cell_value(ws, receipt_cell)
            if receipt_value == period_months:
                target_row = row_idx
                break
        
        if target_row is None:
            print(f"    경고: 수령회차가 {period_months}인 행을 찾을 수 없습니다.")
            target_row = start_row + num_rows - 1
        
        # 두번째 행부터 수령회차가 기간(월)과 같은 행까지 수식 적용
        for row_idx in range(start_row + 1, target_row + 1):
            prev_row = row_idx - 1
            prev_deposit_pv_cell = f"{deposit_pv_col_letter}{prev_row}"
            deposit_pv_cell = f"{deposit_pv_col_letter}{row_idx}"
            deposit_interest_cell = f"{deposit_interest_col_letter}{row_idx}"
            
            # =AA2-AA3 (바로 윗행 임대보증금_현할차 - 동일행 임대보증금_현할차)
            formula = f"={prev_deposit_pv_cell}-{deposit_pv_cell}"
            add_formula(ws, deposit_interest_cell, formula)
        
        filled_rows = target_row - start_row
    
    print(f"  ✓ 임대보증금이자 입력 완료: {filled_rows}행")


def apply_cell_formats(ws, num_rows, start_row=2):
    """
    날짜 열과 금액 열에 엑셀 형식 적용
    
    Args:
        ws: Worksheet 객체
        num_rows (int): 데이터 행 수
        start_row (int): 시작 행 (기본값: 2)
    """
    from openpyxl.utils import get_column_letter
    
    # 날짜 열 목록
    date_headers = [
        "리스개시일",
        "리스종료일",
        "지급일자",
        "유효일자",
    ]
    
    # 금액 열 목록
    amount_headers = [
        "리스료",
        "리스채권",
        "이자수익",
        "상각액",
        "리스채권_유동",
        "차감_사용권자산",
        "차감_감가상각비",
        "차감_감가상각누계액",
        "차감_사용권자산(순)",
        "임대보증금",
        "임대보증금_현할차",
        "임대보증금이자",
    ]
    
    # 날짜 열 인덱스 찾기
    date_col_indices = []
    for header in date_headers:
        for idx, h in enumerate(HEADERS, start=1):
            if h == header:
                date_col_indices.append(idx)
                break
    
    # 금액 열 인덱스 찾기
    amount_col_indices = []
    for header in amount_headers:
        for idx, h in enumerate(HEADERS, start=1):
            if h == header:
                amount_col_indices.append(idx)
                break
    
    # 날짜 형식 적용 (yyyy-mm-dd 형식)
    date_format = "yyyy-mm-dd"
    for col_idx in date_col_indices:
        col_letter = get_column_letter(col_idx)
        for row_idx in range(start_row, start_row + num_rows):
            cell = ws[f"{col_letter}{row_idx}"]
            cell.number_format = date_format
    
    # 금액 형식 적용 (회계 형식: 천단위 구분, 소수점 없음)
    amount_format = "#,##0"
    for col_idx in amount_col_indices:
        col_letter = get_column_letter(col_idx)
        for row_idx in range(start_row, start_row + num_rows):
            cell = ws[f"{col_letter}{row_idx}"]
            cell.number_format = amount_format
    
    if date_col_indices:
        print(f"  ✓ 날짜 형식 적용: {len(date_col_indices)}개 열")
    if amount_col_indices:
        print(f"  ✓ 금액 형식 적용: {len(amount_col_indices)}개 열")


def apply_border_to_table(ws, header_row, num_data_rows, start_col, num_cols):
    """
    테이블 범위에 모든 테두리 적용
    
    Args:
        ws: Worksheet 객체
        header_row (int): 헤더 행 번호
        num_data_rows (int): 헤더를 제외한 데이터 행 수
        start_col (int): 시작 열 번호 (1부터 시작, A열=1)
        num_cols (int): 열 개수
    """
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter
    
    # 테두리 스타일 정의
    thin_border = Side(style='thin')
    border = Border(
        left=thin_border,
        right=thin_border,
        top=thin_border,
        bottom=thin_border
    )
    
    # 헤더 행부터 마지막 데이터 행까지 모든 셀에 테두리 적용
    # num_data_rows는 헤더를 제외한 데이터 행 수이므로 +1을 해서 헤더 포함
    for row_idx in range(header_row, header_row + num_data_rows + 1):
        for col_idx in range(start_col, start_col + num_cols):
            col_letter = get_column_letter(col_idx)
            cell = ws[f"{col_letter}{row_idx}"]
            cell.border = border
    
    print(f"  ✓ 테두리 적용 완료: {num_cols}개 열, {num_data_rows + 1}행 (헤더 포함)")


def create_sheets_for_all_contracts(wb, data_manager):
    """
    모든 전대 계약에 대해 시트 생성 및 기본 설정
    
    Args:
        wb: Workbook 객체
        data_manager: DataManager 객체
        
    Returns:
        dict: 계약번호를 키로 하는 시트 딕셔너리
    """
    print("\n[시트 생성 시작]")
    
    sheets = {}
    
    for contract_no, contract_data in data_manager.contract_data.items():
        print(f"\n계약번호: {contract_no}")
        
        # 1. 시트 생성
        ws = create_contract_sheet(wb, contract_no, contract_data)
        sheets[contract_no] = ws
        
        # 2. 헤더 설정 (1행에 헤더 이름들)
        last_col = set_headers(ws, HEADERS, start_row=1, start_col=1)
        
        # 3. 유형별 행수 계산
        payment_type = str(contract_data.get('선급/후급', '')).strip()
        start_basis = str(contract_data.get('리스개시기준', '')).strip()
        
        # 감가상각기간 컬럼명 찾기 (다양한 이름 대응)
        depreciation_period = None
        for key in contract_data.keys():
            key_str = str(key)
            if '감가상각' in key_str and ('기간' in key_str or '기' in key_str or 'period' in key_str.lower()):
                depreciation_period = contract_data.get(key)
                break
        
        if depreciation_period is None:
            # 여러 가능한 컬럼명 시도
            depreciation_period = contract_data.get('감가상각기간(월)', 
                          contract_data.get('감가상각기간',
                          contract_data.get('감가상각기', 0)))
        
        # 감가상각기 값을 숫자로 변환
        try:
            if isinstance(depreciation_period, str):
                depreciation_period = int(float(depreciation_period))
            elif depreciation_period is None:
                depreciation_period = 0
            else:
                depreciation_period = int(depreciation_period)
        except (ValueError, TypeError) as e:
            print(f"    경고: 감가상각기 값을 숫자로 변환할 수 없습니다: {depreciation_period} (에러: {e})")
            depreciation_period = 0
        
        data_rows = calculate_data_rows(payment_type, start_basis, depreciation_period)
        print(f"  ✓ 데이터 행수 계산: {payment_type}/{start_basis} -> {data_rows}행 (감가상각기: {depreciation_period})")
        
        # 4. Databody 입력 (2행부터 유형별 행수만큼)
        fill_databody_from_contract(ws, contract_data, HEADERS_FROM_DATA, start_row=2, num_rows=data_rows)
        
        # 5. 수령회차 열 채우기
        fill_receipt_count(ws, payment_type, start_basis, data_rows, start_row=2)
        
        # 6. 리스료 열 채우기
        fill_lease_fee(ws, contract_data, data_rows, start_row=2)
        
        # 7. 유효일자 열 채우기 (수식)
        fill_effective_date(ws, start_basis, data_rows, start_row=2)
        
        # 8. 지급일자 열 채우기 (수식)
        fill_payment_date(ws, payment_type, start_basis, data_rows, start_row=2)
        
        # 9. 지급연도 열 채우기 (수식)
        fill_payment_year(ws, payment_type, start_basis, data_rows, start_row=2)
        
        # 10. 회계연도 및 결산월 열 채우기 (수식)
        fill_fiscal_year_and_month(ws, data_rows, start_row=2)
        
        # 11. 리스채권 열 채우기 (수식)
        fill_lease_receivable(ws, payment_type, start_basis, contract_data, data_rows, start_row=2)
        
        # 12. 이자수익 열 채우기 (수식)
        fill_interest_income(ws, payment_type, start_basis, data_rows, start_row=2)
        
        # 13. 상각액 열 채우기 (수식)
        fill_depreciation(ws, data_rows, start_row=2)
        
        # 14. 리스채권_유동 열 채우기 (수식)
        fill_lease_receivable_current(ws, payment_type, start_basis, data_rows, start_row=2)
        
        # 15. 차감_사용권자산 열 채우기 (수식)
        fill_deductible_right_of_use_asset(wb, ws, contract_data, payment_type, start_basis, data_rows, start_row=2)
        
        # 16. 차감_감가상각누계액 열 채우기 (수식)
        fill_deductible_depreciation_accumulated(wb, ws, contract_data, payment_type, start_basis, data_rows, start_row=2)
        
        # 17. 차감_감가상각비 열 채우기 (수식)
        fill_deductible_depreciation(wb, ws, contract_data, payment_type, start_basis, data_rows, start_row=2)
        
        # 18. 차감_사용권자산(순) 열 채우기 (수식)
        fill_deductible_right_of_use_asset_net(ws, data_rows, start_row=2)
        
        # 19. 임대보증금 열 채우기 (수식)
        fill_lease_deposit(wb, ws, contract_data, contract_no, payment_type, start_basis, data_rows, start_row=2)
        
        # 20. 임대보증금_현할차 열 채우기 (수식)
        fill_lease_deposit_present_value(wb, ws, contract_data, contract_no, payment_type, start_basis, data_rows, start_row=2)
        
        # 21. 임대보증금이자 열 채우기 (수식)
        fill_lease_deposit_interest(ws, contract_data, payment_type, start_basis, data_rows, start_row=2)
        
        # 22. 날짜 열 및 금액 열 형식 적용
        apply_cell_formats(ws, data_rows, start_row=2)
        
        # 23. 테이블 범위에 테두리 적용
        # 헤더 행(1행)부터 마지막 데이터 행(1 + data_rows)까지
        apply_border_to_table(ws, 1, data_rows, 1, len(HEADERS))  # 헤더 포함
    
    print(f"\n[시트 생성 완료] 총 {len(sheets)}개 시트 생성")
    
    return sheets

