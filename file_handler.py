"""
파일 처리 모듈
- 엑셀 파일 선택
- 파일 경로 생성 및 관리
- 파일 로드 및 저장
"""

from pathlib import Path
from tkinter import filedialog
import openpyxl


def select_excel_file():
    """
    탐색기를 통해 엑셀 파일 선택
    
    Returns:
        str: 선택된 파일 경로, 취소 시 None
    """
    root = filedialog.Tk()
    root.withdraw()  # 메인 윈도우 숨기기
    
    file_path = filedialog.askopenfilename(
        title="엑셀 파일 선택",
        filetypes=[
            ("Excel files", "*.xlsx *.xlsm *.xls"),
            ("All files", "*.*")
        ]
    )
    
    root.destroy()
    return file_path if file_path else None


def generate_output_path(original_path, suffix="_전대반영후", avoid_overwrite=True):
    """
    원본 파일 경로에서 지정된 접미사를 추가한 새 경로 생성
    기존 파일이 있으면 번호를 추가하여 중복 방지
    
    Args:
        original_path (str): 원본 파일 경로
        suffix (str): 추가할 접미사 (기본값: "_전대반영후")
        avoid_overwrite (bool): 기존 파일 덮어쓰기 방지 여부 (기본값: True)
        
    Returns:
        str: 새 파일 경로 (기존 파일이 있으면 번호 추가)
    """
    path_obj = Path(original_path)
    directory = path_obj.parent
    stem = path_obj.stem
    file_suffix = path_obj.suffix
    
    if avoid_overwrite:
        # 기존 파일이 없을 때까지 번호를 증가시키며 파일명 생성
        counter = 0
        while True:
            if counter == 0:
                new_filename = f"{stem}{suffix}{file_suffix}"
            else:
                new_filename = f"{stem}{suffix}_{counter}{file_suffix}"
            
            output_path = directory / new_filename
            if not output_path.exists():
                return str(output_path)
            counter += 1
    else:
        # 기존 파일 덮어쓰기 허용
        new_filename = f"{stem}{suffix}{file_suffix}"
        return str(directory / new_filename)


def load_workbook(file_path):
    """
    엑셀 파일 로드
    
    Args:
        file_path (str): 엑셀 파일 경로
        
    Returns:
        Workbook: openpyxl Workbook 객체
        
    Raises:
        Exception: 파일 로드 실패 시
    """
    try:
        # .xlsm 파일인 경우 매크로(VBA) 유지를 위해 keep_vba=True 옵션 사용
        file_ext = Path(file_path).suffix.lower()
        if file_ext == '.xlsm':
            wb = openpyxl.load_workbook(file_path, keep_vba=True)
        else:
            wb = openpyxl.load_workbook(file_path)
        print(f"파일 로드 완료: {file_path}")
        return wb
    except Exception as e:
        print(f"파일 로드 오류: {e}")
        raise


def save_workbook(wb, output_path):
    """
    워크북을 파일로 저장
    (keep_vba=True로 로드된 경우 매크로가 자동으로 유지됨)
    
    Args:
        wb: openpyxl Workbook 객체
        output_path (str): 저장할 파일 경로
        
    Raises:
        Exception: 파일 저장 실패 시
    """
    try:
        wb.save(output_path)
        print(f"파일 저장 완료: {output_path}")
    except Exception as e:
        print(f"파일 저장 오류: {e}")
        raise
    finally:
        wb.close()

