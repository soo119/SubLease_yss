"""
임대보증금 시트 생성 모듈
- 임대보증금 시트 생성 및 헤더 설정
"""

from excel_utils import add_sheet, set_cell_value, add_formula
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


# ============================================================================
# 헤더 설정 (이미지 순서대로, 총 43개)
# ============================================================================

HEADERS = [
    "계약번호(Ref no.)",
    "리스명",
    "거래처",
    "자산구분",
    "비용구분",
    "내부거래여부",
    "주석구분",
    "전대_리스개시일",
    "전대_리스종료일",
    "전대_중도해지일",
    "통화",
    "기초",
    "취득",
    "원금상환",
    "이자비용",
    "외화환산손익",
    "변경",
    "중도해지",
    "기말",
    "1월",
    "2월",
    "3월",
    "4월",
    "5월",
    "6월",
    "7월",
    "8월",
    "9월",
    "10월",
    "11월",
    "12월",
    "1월",
    "2월",
    "3월",
    "4월",
    "5월",
    "6월",
    "7월",
    "8월",
    "9월",
    "10월",
    "11월",
    "12월",
]


# ============================================================================
# 임대보증금 시트 생성 함수
# ============================================================================

def create_lease_deposit_sheet(wb):
    """
    임대보증금 시트 생성 및 기본 설정
    
    Args:
        wb: Workbook 객체
    
    Returns:
        Worksheet: 생성된 시트 객체
    """
    print("\n[임대보증금 시트 생성]")
    
    # 시트 생성
    sheet_name = "임대보증금"
    ws = add_sheet(wb, sheet_name)
    
    # B2셀: 임대보증금
    set_cell_value(ws, "B2", "임대보증금")
    
    # B4셀: 기준일
    set_cell_value(ws, "B4", "기준일")
    
    # C4셀: =Input_data!$C$7
    add_formula(ws, "C4", "=Input_data!$C$7")
    # C4셀에 날짜 형식 적용
    ws["C4"].number_format = "yyyy-mm-dd"
    
    # 6행: 셀 병합 및 그룹 헤더 설정 (가운데 정렬)
    # B6~L6: 기본정보 (11개 열)
    ws.merge_cells("B6:L6")
    set_cell_value(ws, "B6", "기본정보")
    ws["B6"].alignment = Alignment(horizontal='center', vertical='center')
    
    # M6~T6: 임대보증금 (8개 열)
    ws.merge_cells("M6:T6")
    set_cell_value(ws, "M6", "임대보증금")
    ws["M6"].alignment = Alignment(horizontal='center', vertical='center')
    
    # U6~AF6: 이자비용 (18개 열)
    ws.merge_cells("U6:AF6")
    set_cell_value(ws, "U6", "이자비용")
    ws["U6"].alignment = Alignment(horizontal='center', vertical='center')
    
    # AG6~AR6: 외화환산손익 (12개 열)
    ws.merge_cells("AG6:AR6")
    set_cell_value(ws, "AG6", "외화환산손익")
    ws["AG6"].alignment = Alignment(horizontal='center', vertical='center')
    
    # 7행: 헤더 행 설정 (B7부터 시작)
    for idx, header in enumerate(HEADERS):
        col_letter = get_column_letter(2 + idx)  # B열부터 시작 (2 = B)
        cell = f"{col_letter}7"
        set_cell_value(ws, cell, header)
    
    print(f"  ✓ 임대보증금 시트 생성 완료")
    print(f"  ✓ 헤더 설정 완료: {len(HEADERS)}개 열")
    
    return ws


def fill_lease_deposit_sheet(wb, data_manager):
    """
    임대보증금 시트에 데이터 채우기
    Input_data 시트의 전대 리스 데이터를 임대보증금 시트에 채움
    
    Args:
        wb: Workbook 객체
        data_manager: DataManager 객체
    """
    from excel_utils import get_sheet_by_name, get_table_data, set_cell_value
    from openpyxl.utils import get_column_letter
    
    print("\n[임대보증금 시트 데이터 채우기]")
    
    # 임대보증금 시트 확인
    lease_deposit_sheet = get_sheet_by_name(wb, "임대보증금")
    if lease_deposit_sheet is None:
        print("  경고: '임대보증금' 시트를 찾을 수 없습니다.")
        return
    
    # Input_data 시트 확인
    input_data_sheet = get_sheet_by_name(wb, "Input_data")
    if input_data_sheet is None:
        print("  경고: 'Input_data' 시트를 찾을 수 없습니다.")
        return
    
    # input_table 테이블에서 데이터 읽기
    try:
        table_data = get_table_data(input_data_sheet, "input_table")
    except Exception as e:
        print(f"  경고: input_table 데이터 읽기 실패: {e}")
        return
    
    # 전대 리스 데이터 필터링 (data_manager의 계약번호 목록 사용)
    filtered_data = []
    contract_nos = set(data_manager.contract_data.keys())
    
    # 계약번호 컬럼 찾기
    ref_no_col = None
    if table_data:
        for col_name in table_data[0].keys():
            col_lower = str(col_name).lower().replace(' ', '').replace('_', '').replace('.', '')
            if 'refno' in col_lower or '계약번호' in str(col_name):
                ref_no_col = col_name
                break
    
    if ref_no_col is None:
        print("  경고: Input_data에서 계약번호 컬럼을 찾을 수 없습니다.")
        return
    
    # 전대 리스 데이터 필터링
    for row in table_data:
        contract_no = row.get(ref_no_col)
        if contract_no in contract_nos:
            filtered_data.append(row)
    
    if len(filtered_data) == 0:
        print("  경고: 필터링된 데이터가 없습니다.")
        return
    
    print(f"  ✓ 전대 리스 데이터 필터링 완료: {len(filtered_data)}행")
    
    # 임대보증금 시트의 헤더와 Input_data의 컬럼 매핑
    # 이미지에서 보이는 열들:
    # 1. 계약번호 (Ref no.) -> 계약번호(Ref no.)
    # 2. 리스명 -> 리스명
    # 3. 거래처 -> 거래처
    # 4. 자산구분 -> 자산구분
    # 5. 비용구분 -> 비용구분
    # 6. 내부거래여부 -> 내부거래여부
    # 7. 주석구분 -> 주석구분
    # 8. 전대_리스개시일 -> 전대_리스개시일
    # 9. 전대_리스종료일 -> 전대_리스종료일
    # 10. 전대_중도해지일 -> 전대_중도해지일
    # 11. 통화 -> 통화
    
    # Input_data의 컬럼명 찾기 (유연한 매칭)
    def find_column(table_data, search_terms):
        """컬럼명을 유연하게 찾기"""
        if not table_data:
            return None
        for col_name in table_data[0].keys():
            col_str = str(col_name).strip().lower().replace(' ', '').replace('_', '').replace('.', '')
            for term in search_terms:
                if term.lower().replace(' ', '').replace('_', '').replace('.', '') in col_str:
                    return col_name
        return None
    
    # 매핑 정의
    column_mapping = [
        ("계약번호(Ref no.)", ["계약번호", "ref no", "refno"]),
        ("리스명", ["리스명"]),
        ("거래처", ["거래처"]),
        ("자산구분", ["자산구분"]),
        ("비용구분", ["비용구분"]),
        ("내부거래여부", ["내부거래", "내부거래여부"]),
        ("주석구분", ["주석구분"]),
        ("전대_리스개시일", ["리스개시일"]),
        ("전대_리스종료일", ["리스종료일"]),
        ("전대_중도해지일", ["리스변경일", "중도해지일"]),
        ("통화", ["통화"]),
    ]
    
    # Input_data의 컬럼 찾기
    input_columns = {}
    for lease_deposit_header, search_terms in column_mapping:
        col = find_column(table_data, search_terms)
        if col:
            input_columns[lease_deposit_header] = col
        else:
            print(f"  경고: '{lease_deposit_header}'에 해당하는 Input_data 컬럼을 찾을 수 없습니다.")
    
    # 데이터 채우기 (8행부터 시작)
    start_row = 8
    date_headers = ["전대_리스개시일", "전대_리스종료일", "전대_중도해지일"]
    
    for row_idx, row_data in enumerate(filtered_data):
        current_row = start_row + row_idx
        
        for lease_deposit_header, input_col in input_columns.items():
            # 임대보증금 시트의 헤더 인덱스 찾기
            header_idx = None
            for idx, header in enumerate(HEADERS):
                if header == lease_deposit_header:
                    header_idx = idx
                    break
            
            if header_idx is None:
                continue
            
            # 셀 주소 계산 (B열부터 시작)
            col_letter = get_column_letter(2 + header_idx)  # B열 = 2
            cell = f"{col_letter}{current_row}"
            
            # 값 가져오기
            value = row_data.get(input_col)
            
            # 날짜 열인 경우 날짜 형식 적용
            if lease_deposit_header in date_headers:
                set_cell_value(lease_deposit_sheet, cell, value)
                lease_deposit_sheet[cell].number_format = "yyyy-mm-dd"
            else:
                set_cell_value(lease_deposit_sheet, cell, value)
    
    print(f"  ✓ 임대보증금 시트 데이터 채우기 완료: {len(filtered_data)}행")
    
    # 기초 열 채우기
    fill_beginning_balance_column(wb, lease_deposit_sheet, len(filtered_data), start_row)


def fill_beginning_balance_column(wb, lease_deposit_sheet, num_rows, start_row=8):
    """
    임대보증금 시트의 기초 열 채우기
    
    수식: =IF(I8>=Input_data!$C$6,0,SUMIFS(전대_Lease_Data!$AC:$AC,전대_Lease_Data!$B:$B,B8,전대_Lease_Data!$I:$I,EOMONTH(Input_data!$C$6,-1)))
    
    Args:
        wb: Workbook 객체
        lease_deposit_sheet: 임대보증금 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8)
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[기초 열 채우기]")
    
    # 임대보증금 시트의 헤더 인덱스 찾기
    contract_no_col_idx = None
    lease_start_date_col_idx = None
    beginning_balance_col_idx = None
    
    for idx, header in enumerate(HEADERS):
        if header == "계약번호(Ref no.)":
            contract_no_col_idx = idx
        elif header == "전대_리스개시일":
            lease_start_date_col_idx = idx
        elif header == "기초":
            beginning_balance_col_idx = idx
    
    if contract_no_col_idx is None or lease_start_date_col_idx is None or beginning_balance_col_idx is None:
        print("  경고: 필요한 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산 (B열부터 시작)
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열 = 2
    lease_start_date_col_letter = get_column_letter(2 + lease_start_date_col_idx)  # I열
    beginning_balance_col_letter = get_column_letter(2 + beginning_balance_col_idx)  # L열
    
    # 전대_Lease_Data 시트의 열 문자
    # 계약번호(Ref no.): B열
    # 유효일자: I열
    # 임대보증금: AC열 (임대보증금_현할차 열)
    lease_data_contract_no_col = "B"  # 전대_Lease_Data 시트의 계약번호 열
    lease_data_effective_date_col = "I"  # 전대_Lease_Data 시트의 유효일자 열
    lease_data_deposit_col = "AC"  # 전대_Lease_Data 시트의 임대보증금 열 (AC열)
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =IF(I8>=Input_data!$C$6,0,SUMIFS(전대_Lease_Data!$AC:$AC,전대_Lease_Data!$B:$B,B8,전대_Lease_Data!$I:$I,EOMONTH(Input_data!$C$6,-1)))
        
        lease_start_date_cell = f"{lease_start_date_col_letter}{row_idx}"
        contract_no_cell = f"{contract_no_col_letter}{row_idx}"
        beginning_balance_cell = f"{beginning_balance_col_letter}{row_idx}"
        
        formula = f"=IF({lease_start_date_cell}>=Input_data!$C$6,0,SUMIFS(전대_Lease_Data!${lease_data_deposit_col}:${lease_data_deposit_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_effective_date_col}:${lease_data_effective_date_col},EOMONTH(Input_data!$C$6,-1)))"
        
        add_formula(lease_deposit_sheet, beginning_balance_cell, formula)
    
    print(f"  ✓ 기초 열 입력 완료: {num_rows}행")
    
    # 취득 열 채우기
    fill_acquisition_column(wb, lease_deposit_sheet, num_rows, start_row)


def fill_acquisition_column(wb, lease_deposit_sheet, num_rows, start_row=8):
    """
    임대보증금 시트의 취득 열 채우기
    
    수식: =IF(I8>=Input_data!$C$6,SUMIFS(전대_Lease_Data!$AC:$AC,전대_Lease_Data!$B:$B,B8,전대_Lease_Data!$N:$N, MINIFS(전대_Lease_Data!$N:$N, 전대_Lease_Data!$B:$B, B8)),0)
    
    Args:
        wb: Workbook 객체
        lease_deposit_sheet: 임대보증금 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8)
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[취득 열 채우기]")
    
    # 임대보증금 시트의 헤더 인덱스 찾기
    contract_no_col_idx = None
    lease_start_date_col_idx = None
    acquisition_col_idx = None
    
    for idx, header in enumerate(HEADERS):
        if header == "계약번호(Ref no.)":
            contract_no_col_idx = idx
        elif header == "전대_리스개시일":
            lease_start_date_col_idx = idx
        elif header == "취득":
            acquisition_col_idx = idx
    
    if contract_no_col_idx is None or lease_start_date_col_idx is None or acquisition_col_idx is None:
        print("  경고: 필요한 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산 (B열부터 시작)
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열 = 2
    lease_start_date_col_letter = get_column_letter(2 + lease_start_date_col_idx)  # I열
    acquisition_col_letter = get_column_letter(2 + acquisition_col_idx)  # M열
    
    # 전대_Lease_Data 시트의 열 문자
    # 계약번호(Ref no.): B열
    # 수령회차: N열
    # 임대보증금: AC열
    lease_data_contract_no_col = "B"  # 전대_Lease_Data 시트의 계약번호 열
    lease_data_receipt_count_col = "N"  # 전대_Lease_Data 시트의 수령회차 열
    lease_data_deposit_col = "AC"  # 전대_Lease_Data 시트의 임대보증금 열
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =IF(I8>=Input_data!$C$6,SUMIFS(전대_Lease_Data!$AC:$AC,전대_Lease_Data!$B:$B,B8,전대_Lease_Data!$N:$N, MINIFS(전대_Lease_Data!$N:$N, 전대_Lease_Data!$B:$B, B8)),0)
        
        lease_start_date_cell = f"{lease_start_date_col_letter}{row_idx}"
        contract_no_cell = f"{contract_no_col_letter}{row_idx}"
        acquisition_cell = f"{acquisition_col_letter}{row_idx}"
        
        formula = f"=IF({lease_start_date_cell}>=Input_data!$C$6,SUMIFS(전대_Lease_Data!${lease_data_deposit_col}:${lease_data_deposit_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_receipt_count_col}:${lease_data_receipt_count_col},MINIFS(전대_Lease_Data!${lease_data_receipt_count_col}:${lease_data_receipt_count_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell})),0)"
        
        add_formula(lease_deposit_sheet, acquisition_cell, formula)
    
    print(f"  ✓ 취득 열 입력 완료: {num_rows}행")
    
    # 원금상환 열 채우기
    fill_principal_repayment_column(wb, lease_deposit_sheet, num_rows, start_row)


def fill_principal_repayment_column(wb, lease_deposit_sheet, num_rows, start_row=8):
    """
    임대보증금 시트의 원금상환 열 채우기
    
    수식: =-IF(OR(J8<=C$4$,AND(K8<>0,K8<=C$4$)),Input_data!W15,0)
    
    Args:
        wb: Workbook 객체
        lease_deposit_sheet: 임대보증금 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8)
    """
    from excel_utils import add_formula, get_sheet_by_name, get_table_data, find_table_by_name
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    
    print("\n[원금상환 열 채우기]")
    
    # 임대보증금 시트의 헤더 인덱스 찾기
    contract_no_col_idx = None
    lease_end_date_col_idx = None
    early_termination_date_col_idx = None
    principal_repayment_col_idx = None
    
    for idx, header in enumerate(HEADERS):
        if header == "계약번호(Ref no.)":
            contract_no_col_idx = idx
        elif header == "전대_리스종료일":
            lease_end_date_col_idx = idx
        elif header == "전대_중도해지일":
            early_termination_date_col_idx = idx
        elif header == "원금상환":
            principal_repayment_col_idx = idx
    
    if contract_no_col_idx is None or lease_end_date_col_idx is None or early_termination_date_col_idx is None or principal_repayment_col_idx is None:
        print("  경고: 필요한 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산 (B열부터 시작)
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열 = 2
    lease_end_date_col_letter = get_column_letter(2 + lease_end_date_col_idx)  # J열
    early_termination_date_col_letter = get_column_letter(2 + early_termination_date_col_idx)  # K열
    principal_repayment_col_letter = get_column_letter(2 + principal_repayment_col_idx)  # N열
    
    # Input_data 시트에서 임차보증금 열 찾기
    input_data_sheet = get_sheet_by_name(wb, "Input_data")
    if input_data_sheet is None:
        print("  경고: 'Input_data' 시트를 찾을 수 없습니다.")
        return
    
    # input_table 테이블 찾기
    input_ws, table = find_table_by_name(wb, "input_table")
    if table is None:
        print("  경고: 'input_table' 테이블을 찾을 수 없습니다.")
        return
    
    # 테이블 데이터 읽기
    try:
        table_data = get_table_data(input_data_sheet, "input_table")
    except Exception as e:
        print(f"  경고: input_table 데이터 읽기 실패: {e}")
        return
    
    # 계약번호(Ref no.) 열 찾기
    ref_no_col = None
    headers = list(table_data[0].keys())
    for col_name in headers:
        col_str = str(col_name).strip()
        if '계약번호' in col_str and 'Ref no' in col_str and ('상위' not in col_str and 'parent' not in col_str.lower()):
            ref_no_col = col_name
            break
    
    if ref_no_col is None:
        print("  경고: Input_data에서 계약번호(Ref no.) 컬럼을 찾을 수 없습니다.")
        return
    
    # 임차보증금 열 찾기
    deposit_col = None
    for col_name in headers:
        col_str = str(col_name).strip()
        if '임차보증금' in col_str or '보증금' in col_str:
            deposit_col = col_name
            break
    
    if deposit_col is None:
        print("  경고: Input_data에서 임차보증금 컬럼을 찾을 수 없습니다.")
        return
    
    # 테이블 범위 파싱
    ref = table.ref
    start_cell, end_cell = ref.split(':')
    start_col_str, start_row_num = coordinate_from_string(start_cell)
    start_col_idx = column_index_from_string(start_col_str)
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 계약번호 가져오기
        contract_no_cell = f"{contract_no_col_letter}{row_idx}"
        contract_no_value = lease_deposit_sheet[contract_no_cell].value
        
        if contract_no_value is None:
            continue
        
        # Input_data에서 해당 계약번호의 임차보증금 셀 참조 찾기
        deposit_cell_ref = None
        for data_row_idx, row_data in enumerate(table_data, start=1):
            if str(row_data.get(ref_no_col)) == str(contract_no_value):
                # 테이블의 헤더 행 다음부터 시작하므로 +1
                data_row = start_row_num + data_row_idx
                
                # 임차보증금 열 인덱스 찾기
                deposit_col_idx_in_table = headers.index(deposit_col)
                deposit_col_letter = get_column_letter(start_col_idx + deposit_col_idx_in_table)
                deposit_cell_ref = f"Input_data!${deposit_col_letter}${data_row}"
                break
        
        if deposit_cell_ref is None:
            print(f"  경고: 계약번호 {contract_no_value}에 대한 임차보증금 셀을 찾을 수 없습니다.")
            continue
        
        # 수식 구성
        # =-IF(OR(J8<=C$4$,AND(K8<>0,K8<=C$4$)),Input_data!W15,0)
        lease_end_date_cell = f"{lease_end_date_col_letter}{row_idx}"
        early_termination_date_cell = f"{early_termination_date_col_letter}{row_idx}"
        principal_repayment_cell = f"{principal_repayment_col_letter}{row_idx}"
        
        # C$4는 임대보증금 시트의 C4셀 (기준일)
        formula = f"=-IF(OR({lease_end_date_cell}<=$C$4,AND({early_termination_date_cell}<>0,{early_termination_date_cell}<=$C$4)),{deposit_cell_ref},0)"
        
        add_formula(lease_deposit_sheet, principal_repayment_cell, formula)
    
    print(f"  ✓ 원금상환 열 입력 완료: {num_rows}행")
    
    # 이자비용 열 채우기
    fill_interest_expense_column(lease_deposit_sheet, num_rows, start_row)


def fill_interest_expense_column(lease_deposit_sheet, num_rows, start_row=8):
    """
    임대보증금 시트의 이자비용 열 채우기
    
    수식: =SUMIFS(전대_Lease_Data!$AE:$AE,전대_Lease_Data!$B:$B,B8,전대_Lease_Data!$I:$I,">="&Input_data!$C$6,전대_Lease_Data!$I:$I,"<="&IF(K8=0,$C$4,MIN(K8,$C$4)))
    
    Args:
        lease_deposit_sheet: 임대보증금 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8)
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[이자비용 열 채우기]")
    
    # 임대보증금 시트의 헤더 인덱스 찾기
    contract_no_col_idx = None
    early_termination_date_col_idx = None
    interest_expense_col_idx = None
    
    for idx, header in enumerate(HEADERS):
        if header == "계약번호(Ref no.)":
            contract_no_col_idx = idx
        elif header == "전대_중도해지일":
            early_termination_date_col_idx = idx
        elif header == "이자비용":
            interest_expense_col_idx = idx
    
    if contract_no_col_idx is None or early_termination_date_col_idx is None or interest_expense_col_idx is None:
        print("  경고: 필요한 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산 (B열부터 시작)
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열 = 2
    early_termination_date_col_letter = get_column_letter(2 + early_termination_date_col_idx)  # K열
    interest_expense_col_letter = get_column_letter(2 + interest_expense_col_idx)  # O열
    
    # 전대_Lease_Data 시트의 열 문자
    # 계약번호(Ref no.): B열
    # 유효일자: I열
    # 임대보증금이자: AE열
    lease_data_contract_no_col = "B"  # 전대_Lease_Data 시트의 계약번호 열
    lease_data_effective_date_col = "I"  # 전대_Lease_Data 시트의 유효일자 열
    lease_data_deposit_interest_col = "AE"  # 전대_Lease_Data 시트의 임대보증금이자 열
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =SUMIFS(전대_Lease_Data!$AE:$AE,전대_Lease_Data!$B:$B,B8,전대_Lease_Data!$I:$I,">="&Input_data!$C$6,전대_Lease_Data!$I:$I,"<="&IF(K8=0,$C$4,MIN(K8,$C$4)))
        
        contract_no_cell = f"{contract_no_col_letter}{row_idx}"
        early_termination_date_cell = f"{early_termination_date_col_letter}{row_idx}"
        interest_expense_cell = f"{interest_expense_col_letter}{row_idx}"
        
        formula = f"=SUMIFS(전대_Lease_Data!${lease_data_deposit_interest_col}:${lease_data_deposit_interest_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_effective_date_col}:${lease_data_effective_date_col},\">=\"&Input_data!$C$6,전대_Lease_Data!${lease_data_effective_date_col}:${lease_data_effective_date_col},\"<=\"&IF({early_termination_date_cell}=0,$C$4,MIN({early_termination_date_cell},$C$4)))"
        
        add_formula(lease_deposit_sheet, interest_expense_cell, formula)
    
    print(f"  ✓ 이자비용 열 입력 완료: {num_rows}행")
    
    # 중도해지 열 채우기
    fill_early_termination_column(lease_deposit_sheet, num_rows, start_row)


def fill_early_termination_column(lease_deposit_sheet, num_rows, start_row=8):
    """
    임대보증금 시트의 중도해지 열 채우기
    
    수식: =SUMIFS(전대_Lease_Data!AD:AD,전대_Lease_Data!B:B,임대보증금!B8,전대_Lease_Data!AP:AP,"O")
    
    Args:
        lease_deposit_sheet: 임대보증금 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8)
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[중도해지 열 채우기]")
    
    # 임대보증금 시트의 헤더 인덱스 찾기
    contract_no_col_idx = None
    early_termination_col_idx = None
    
    for idx, header in enumerate(HEADERS):
        if header == "계약번호(Ref no.)":
            contract_no_col_idx = idx
        elif header == "중도해지":
            early_termination_col_idx = idx
    
    if contract_no_col_idx is None or early_termination_col_idx is None:
        print("  경고: 필요한 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산 (B열부터 시작)
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열 = 2
    early_termination_col_letter = get_column_letter(2 + early_termination_col_idx)  # Q열
    
    # 전대_Lease_Data 시트의 열 문자
    # 계약번호(Ref no.): B열
    # 임대보증금_현할차: AD열
    # 중도해지: AP열
    lease_data_contract_no_col = "B"  # 전대_Lease_Data 시트의 계약번호 열
    lease_data_deposit_pv_col = "AD"  # 전대_Lease_Data 시트의 임대보증금_현할차 열
    lease_data_early_termination_col = "AP"  # 전대_Lease_Data 시트의 중도해지 열
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =SUMIFS(전대_Lease_Data!AD:AD,전대_Lease_Data!B:B,임대보증금!B8,전대_Lease_Data!AP:AP,"O")
        
        contract_no_cell = f"{contract_no_col_letter}{row_idx}"
        early_termination_cell = f"{early_termination_col_letter}{row_idx}"
        
        formula = f"=SUMIFS(전대_Lease_Data!${lease_data_deposit_pv_col}:${lease_data_deposit_pv_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_early_termination_col}:${lease_data_early_termination_col},\"O\")"
        
        add_formula(lease_deposit_sheet, early_termination_cell, formula)
    
    print(f"  ✓ 중도해지 열 입력 완료: {num_rows}행")
    
    # 기말 열 채우기
    fill_ending_balance_column(lease_deposit_sheet, num_rows, start_row)


def fill_ending_balance_column(lease_deposit_sheet, num_rows, start_row=8):
    """
    임대보증금 시트의 기말 열 채우기
    
    수식: =SUM(M8:S8)
    M8: 동일행의 기초 열
    S8: 동일행의 중도해지 열
    
    Args:
        lease_deposit_sheet: 임대보증금 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8)
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[기말 열 채우기]")
    
    # 임대보증금 시트의 헤더 인덱스 찾기
    beginning_balance_col_idx = None
    early_termination_col_idx = None
    ending_balance_col_idx = None
    
    for idx, header in enumerate(HEADERS):
        if header == "기초":
            beginning_balance_col_idx = idx
        elif header == "중도해지":
            early_termination_col_idx = idx
        elif header == "기말":
            ending_balance_col_idx = idx
    
    if beginning_balance_col_idx is None or early_termination_col_idx is None or ending_balance_col_idx is None:
        print("  경고: 필요한 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산 (B열부터 시작)
    beginning_balance_col_letter = get_column_letter(2 + beginning_balance_col_idx)  # M열
    early_termination_col_letter = get_column_letter(2 + early_termination_col_idx)  # S열
    ending_balance_col_letter = get_column_letter(2 + ending_balance_col_idx)  # T열
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =SUM(M8:S8)
        
        beginning_balance_cell = f"{beginning_balance_col_letter}{row_idx}"
        early_termination_cell = f"{early_termination_col_letter}{row_idx}"
        ending_balance_cell = f"{ending_balance_col_letter}{row_idx}"
        
        formula = f"=SUM({beginning_balance_cell}:{early_termination_cell})"
        
        add_formula(lease_deposit_sheet, ending_balance_cell, formula)
    
    print(f"  ✓ 기말 열 입력 완료: {num_rows}행")
    
    # 1월~12월 열 채우기 (이자비용 그룹)
    fill_monthly_interest_expense_columns(lease_deposit_sheet, num_rows, start_row)


def fill_monthly_interest_expense_columns(lease_deposit_sheet, num_rows, start_row=8):
    """
    임대보증금 시트의 1월~12월 열 채우기 (이자비용 그룹)
    
    수식: =SUMIFS(전대_Lease_Data!$AE:$AE,전대_Lease_Data!$B:$B,$B8,전대_Lease_Data!$I:$I,"<="&EOMONTH(MIN($J8,$K8),0),전대_Lease_Data!$J:$J,YEAR($C$4),전대_Lease_Data!$K:$K,SUBSTITUTE(U7,"월",""))
    
    Args:
        lease_deposit_sheet: 임대보증금 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8)
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[1월~12월 열 채우기 (이자비용 그룹)]")
    
    # 임대보증금 시트의 헤더 인덱스 찾기
    contract_no_col_idx = None
    lease_end_date_col_idx = None
    early_termination_date_col_idx = None
    
    for idx, header in enumerate(HEADERS):
        if header == "계약번호(Ref no.)":
            contract_no_col_idx = idx
        elif header == "전대_리스종료일":
            lease_end_date_col_idx = idx
        elif header == "전대_중도해지일":
            early_termination_date_col_idx = idx
    
    if contract_no_col_idx is None or lease_end_date_col_idx is None or early_termination_date_col_idx is None:
        print("  경고: 필요한 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산 (B열부터 시작)
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열 = 2
    lease_end_date_col_letter = get_column_letter(2 + lease_end_date_col_idx)  # J열
    early_termination_date_col_letter = get_column_letter(2 + early_termination_date_col_idx)  # K열
    
    # 전대_Lease_Data 시트의 열 문자
    # 계약번호(Ref no.): B열
    # 유효일자: I열
    # 회계연도: J열
    # 결산월: K열
    # 임대보증금이자: AE열
    lease_data_contract_no_col = "B"  # 전대_Lease_Data 시트의 계약번호 열
    lease_data_effective_date_col = "I"  # 전대_Lease_Data 시트의 유효일자 열
    lease_data_fiscal_year_col = "J"  # 전대_Lease_Data 시트의 회계연도 열
    lease_data_fiscal_month_col = "K"  # 전대_Lease_Data 시트의 결산월 열
    lease_data_deposit_interest_col = "AE"  # 전대_Lease_Data 시트의 임대보증금이자 열
    
    # 임대보증금 시트의 기준일 셀 (C4)
    base_date_cell = "C4"
    
    # 헤더 행 (7행)
    header_row = 7
    
    # 1월~12월 열 찾기 (첫 번째 세트, 이자비용 그룹)
    monthly_headers = ["1월", "2월", "3월", "4월", "5월", "6월", "7월", "8월", "9월", "10월", "11월", "12월"]
    monthly_col_indices = []
    
    # 첫 번째 1월~12월 세트 찾기 (이자비용 그룹)
    # HEADERS 리스트에서 첫 번째로 나타나는 1월~12월을 찾음
    for month_header in monthly_headers:
        for idx, header in enumerate(HEADERS):
            if header == month_header and idx not in monthly_col_indices:
                monthly_col_indices.append(idx)
                break
    
    if len(monthly_col_indices) != 12:
        print("  경고: 1월~12월 헤더를 모두 찾을 수 없습니다.")
        return
    
    # 각 월별 열에 수식 입력
    for month_idx, col_idx in enumerate(monthly_col_indices):
        month_col_letter = get_column_letter(2 + col_idx)  # U열부터 시작
        header_cell = f"{month_col_letter}{header_row}"  # 헤더 셀 (예: U7)
        
        for row_idx in range(start_row, start_row + num_rows):
            # 수식 구성
            # =SUMIFS(전대_Lease_Data!$AE:$AE,전대_Lease_Data!$B:$B,$B8,전대_Lease_Data!$I:$I,"<="&EOMONTH(MIN($J8,$K8),0),전대_Lease_Data!$J:$J,YEAR($C$4),전대_Lease_Data!$K:$K,SUBSTITUTE(U7,"월",""))
            
            contract_no_cell = f"${contract_no_col_letter}${row_idx}"  # $B8
            lease_end_date_cell = f"${lease_end_date_col_letter}${row_idx}"  # $J8
            early_termination_date_cell = f"${early_termination_date_col_letter}${row_idx}"  # $K8
            month_cell = f"{month_col_letter}{row_idx}"
            
            formula = f"=SUMIFS(전대_Lease_Data!${lease_data_deposit_interest_col}:${lease_data_deposit_interest_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_effective_date_col}:${lease_data_effective_date_col},\"<=\"&EOMONTH(MIN({lease_end_date_cell},{early_termination_date_cell}),0),전대_Lease_Data!${lease_data_fiscal_year_col}:${lease_data_fiscal_year_col},YEAR(${base_date_cell}),전대_Lease_Data!${lease_data_fiscal_month_col}:${lease_data_fiscal_month_col},SUBSTITUTE({header_cell},\"월\",\"\"))"
            
            add_formula(lease_deposit_sheet, month_cell, formula)
    
    print(f"  ✓ 1월~12월 열 입력 완료 (이자비용 그룹): {num_rows}행")
    
    # 회계 형식 적용 (기초 열부터 마지막 12월 열까지)
    apply_accounting_format(lease_deposit_sheet, num_rows, start_row)
    
    # 테이블 범위에 테두리 적용
    apply_border_to_lease_deposit_table(lease_deposit_sheet, num_rows, start_row)


def apply_accounting_format(lease_deposit_sheet, num_rows, start_row=8):
    """
    임대보증금 시트의 기초 열부터 마지막 12월 열까지 회계 형식 적용
    
    Args:
        lease_deposit_sheet: 임대보증금 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8)
    """
    from openpyxl.utils import get_column_letter
    
    print("\n[회계 형식 적용]")
    
    # 기초 열의 인덱스 찾기
    beginning_balance_col_idx = None
    for idx, header in enumerate(HEADERS):
        if header == "기초":
            beginning_balance_col_idx = idx
            break
    
    if beginning_balance_col_idx is None:
        print("  경고: '기초' 헤더를 찾을 수 없습니다.")
        return
    
    # 마지막 12월 열의 인덱스 찾기 (두 번째 세트의 12월)
    last_dec_col_idx = None
    first_dec_found = False
    for idx, header in enumerate(HEADERS):
        if header == "12월":
            if not first_dec_found:
                first_dec_found = True
            else:
                last_dec_col_idx = idx
                break
    
    if last_dec_col_idx is None:
        print("  경고: 마지막 '12월' 헤더를 찾을 수 없습니다.")
        return
    
    # 회계 형식 (#,##0)
    accounting_format = "#,##0"
    
    # 기초 열부터 마지막 12월 열까지 회계 형식 적용
    for col_idx in range(beginning_balance_col_idx, last_dec_col_idx + 1):
        col_letter = get_column_letter(2 + col_idx)  # B열부터 시작
        for row_idx in range(start_row, start_row + num_rows):
            cell = lease_deposit_sheet[f"{col_letter}{row_idx}"]
            cell.number_format = accounting_format
    
    num_cols = last_dec_col_idx - beginning_balance_col_idx + 1
    print(f"  ✓ 회계 형식 적용 완료: {num_cols}개 열 ({num_rows}행)")


def apply_border_to_lease_deposit_table(lease_deposit_sheet, num_rows, start_row=8):
    """
    임대보증금 시트의 테이블 범위에 모든 테두리 적용
    
    Args:
        lease_deposit_sheet: 임대보증금 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8, 헤더는 7행)
    """
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter
    
    # 테두리 스타일 정의
    thin_border = Side(style='thin')
    border = Border(
        left=thin_border,
        right=thin_border,
        top=thin_border,
        bottom=thin_border
    )
    
    # 셀병합된 행 (6행)부터 마지막 데이터 행까지
    merged_row = 6  # 셀병합된 행
    header_row = 7  # 헤더 행
    num_cols = len(HEADERS)  # 헤더 개수
    start_col = 2  # B열부터 시작
    
    # 셀병합된 행(6행)에 테두리 적용
    for col_idx in range(start_col, start_col + num_cols):
        col_letter = get_column_letter(col_idx)
        cell = lease_deposit_sheet[f"{col_letter}{merged_row}"]
        cell.border = border
    
    # 헤더 행(7행)부터 마지막 데이터 행까지 테두리 적용
    for row_idx in range(header_row, start_row + num_rows):
        for col_idx in range(start_col, start_col + num_cols):
            col_letter = get_column_letter(col_idx)
            cell = lease_deposit_sheet[f"{col_letter}{row_idx}"]
            cell.border = border
    
    print(f"  ✓ 테두리 적용 완료: {num_cols}개 열, {num_rows + 2}행 (셀병합 행 + 헤더 포함)")

