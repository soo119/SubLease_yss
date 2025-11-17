"""
전대사용권자산 시트 생성 모듈
- 전대사용권자산 시트 생성 및 데이터 추가
"""

from excel_utils import get_sheet_by_name, get_table_data, set_cell_value, add_sheet, add_formula
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


# 기본 정보 헤더 (11개)
BASIC_HEADERS = [
    "계약번호(Ref no.)",
    "리스명",
    "거래처",
    "자산구분",
    "비용구분",
    "내부거래여부",
    "주석구분",
    "리스개시일",
    "리스종료일",
    "리스변경일",
    "통화",
]

# 사용권자산_취득가액 그룹 헤더 (4개)
ACQUISITION_HEADERS = [
    "기초",
    "증가",
    "감소",
    "기말",
]

# 사용권자산_감가상각누계액 그룹 헤더 (5개) - 증가와 변경 사이에 상각 포함
DEPRECIATION_HEADERS = [
    "기초",
    "증가",
    "상각",  # 증가와 변경 사이에 상각 열 추가
    "감소",  # 변경 열
    "기말",
]

# 사용권자산_손상차손누계액 그룹 헤더 (5개)
IMPAIRMENT_HEADERS = [
    "기초",
    "증가",
    "감소",
    "종료",
    "기말",
]

# 사용권자산_장부가액 그룹 헤더 (3개)
BOOK_VALUE_HEADERS = [
    "임차료",
    "보증금",
    "복구충당부채",
]

# 감가상각비 그룹 헤더 (12개) - 1월~12월
DEPRECIATION_EXPENSE_HEADERS = [
    "1월",
    "2월",
    "3월",
    "4월",
    "5월",
    "6월",
    "7월",
    "8월",
    "9월",
    "10월",
    "11월",
    "12월",
]

# 전체 헤더 목록
HEADERS = BASIC_HEADERS + ACQUISITION_HEADERS + DEPRECIATION_HEADERS + IMPAIRMENT_HEADERS + BOOK_VALUE_HEADERS + DEPRECIATION_EXPENSE_HEADERS


def create_right_of_use_asset_sheet(wb):
    """
    전대사용권자산 시트 생성 및 기본 설정
    
    Args:
        wb: Workbook 객체
    
    Returns:
        Worksheet: 생성된 시트 객체
    """
    print("\n[전대사용권자산 시트 생성]")
    
    # 시트 생성
    sheet_name = "전대사용권자산"
    ws = add_sheet(wb, sheet_name)
    
    # B2셀: 사용권자산
    set_cell_value(ws, "B2", "사용권자산")
    
    # B4셀: 기준일
    set_cell_value(ws, "B4", "기준일")
    
    # C4셀: =Input_data!$C$7
    add_formula(ws, "C4", "=Input_data!$C$7")
    # C4셀에 날짜 형식 적용
    ws["C4"].number_format = "yyyy-mm-dd"
    
    # 6행: 셀 병합 및 그룹 헤더 설정 (가운데 정렬)
    col_idx = 2  # B열부터 시작
    
    # B6~L6: 기본정보 (11개 열)
    start_col = get_column_letter(col_idx)
    end_col = get_column_letter(col_idx + len(BASIC_HEADERS) - 1)
    ws.merge_cells(f"{start_col}6:{end_col}6")
    set_cell_value(ws, f"{start_col}6", "기본정보")
    ws[f"{start_col}6"].alignment = Alignment(horizontal='center', vertical='center')
    col_idx += len(BASIC_HEADERS)
    
    # 사용권자산_취득가액 그룹 (4개 열)
    start_col = get_column_letter(col_idx)
    end_col = get_column_letter(col_idx + len(ACQUISITION_HEADERS) - 1)
    ws.merge_cells(f"{start_col}6:{end_col}6")
    set_cell_value(ws, f"{start_col}6", "사용권자산_취득가액")
    ws[f"{start_col}6"].alignment = Alignment(horizontal='center', vertical='center')
    col_idx += len(ACQUISITION_HEADERS)
    
    # 사용권자산_감가상각누계액 그룹 (5개 열)
    start_col = get_column_letter(col_idx)
    end_col = get_column_letter(col_idx + len(DEPRECIATION_HEADERS) - 1)
    ws.merge_cells(f"{start_col}6:{end_col}6")
    set_cell_value(ws, f"{start_col}6", "사용권자산_감가상각누계액")
    ws[f"{start_col}6"].alignment = Alignment(horizontal='center', vertical='center')
    col_idx += len(DEPRECIATION_HEADERS)
    
    # 사용권자산_손상차손누계액 그룹 (5개 열)
    start_col = get_column_letter(col_idx)
    end_col = get_column_letter(col_idx + len(IMPAIRMENT_HEADERS) - 1)
    ws.merge_cells(f"{start_col}6:{end_col}6")
    set_cell_value(ws, f"{start_col}6", "사용권자산_손상차손누계액")
    ws[f"{start_col}6"].alignment = Alignment(horizontal='center', vertical='center')
    col_idx += len(IMPAIRMENT_HEADERS)
    
    # 사용권자산_장부가액 그룹 (3개 열)
    start_col = get_column_letter(col_idx)
    end_col = get_column_letter(col_idx + len(BOOK_VALUE_HEADERS) - 1)
    ws.merge_cells(f"{start_col}6:{end_col}6")
    set_cell_value(ws, f"{start_col}6", "사용권자산_장부가액")
    ws[f"{start_col}6"].alignment = Alignment(horizontal='center', vertical='center')
    col_idx += len(BOOK_VALUE_HEADERS)
    
    # 감가상각비 그룹 (12개 열)
    start_col = get_column_letter(col_idx)
    end_col = get_column_letter(col_idx + len(DEPRECIATION_EXPENSE_HEADERS) - 1)
    ws.merge_cells(f"{start_col}6:{end_col}6")
    set_cell_value(ws, f"{start_col}6", "감가상각비")
    ws[f"{start_col}6"].alignment = Alignment(horizontal='center', vertical='center')
    
    # 7행: 헤더 행 설정 (B7부터 시작)
    col_idx = 2  # B열부터 시작
    for header in HEADERS:
        col_letter = get_column_letter(col_idx)
        cell = f"{col_letter}7"
        set_cell_value(ws, cell, header)
        col_idx += 1
    
    print(f"  ✓ 전대사용권자산 시트 생성 완료")
    print(f"  ✓ 헤더 설정 완료: {len(HEADERS)}개 열")
    
    return ws


def fill_right_of_use_asset_sheet(wb, data_manager):
    """
    전대사용권자산 시트에 전대 리스 데이터 추가
    
    Args:
        wb: Workbook 객체
        data_manager: DataManager 객체
    """
    print("\n[전대사용권자산 시트 데이터 추가]")
    
    # 전대사용권자산 시트 확인
    right_of_use_asset_sheet = get_sheet_by_name(wb, "전대사용권자산")
    if right_of_use_asset_sheet is None:
        print("  경고: '전대사용권자산' 시트를 찾을 수 없습니다.")
        return
    
    # Input_data 시트 확인
    input_data_sheet = get_sheet_by_name(wb, "Input_data")
    if input_data_sheet is None:
        print("  경고: 'Input_data' 시트를 찾을 수 없습니다.")
        return
    
    # input_table 테이블에서 데이터 읽기
    try:
        table_data = get_table_data(input_data_sheet, "input_table")
    except Exception as e:
        print(f"  경고: input_table 데이터 읽기 실패: {e}")
        return
    
    # 전대 리스 데이터 필터링 (data_manager의 계약번호 목록 사용)
    filtered_data = []
    contract_nos = set(data_manager.contract_data.keys())
    
    # 계약번호 컬럼 찾기
    ref_no_col = None
    if table_data:
        for col_name in table_data[0].keys():
            col_lower = str(col_name).lower().replace(' ', '').replace('_', '').replace('.', '')
            if 'refno' in col_lower or '계약번호' in str(col_name):
                ref_no_col = col_name
                break
    
    if ref_no_col is None:
        print("  경고: Input_data에서 계약번호 컬럼을 찾을 수 없습니다.")
        return
    
    # 전대 리스 데이터 필터링
    for row in table_data:
        contract_no = row.get(ref_no_col)
        if contract_no in contract_nos:
            filtered_data.append(row)
    
    if len(filtered_data) == 0:
        print("  경고: 필터링된 데이터가 없습니다.")
        return
    
    print(f"  ✓ 전대 리스 데이터 필터링 완료: {len(filtered_data)}행")
    
    # B7셀부터 헤더행이 존재
    header_row = 7
    
    # 데이터 시작 행 (헤더 행 다음)
    start_row = header_row + 1
    print(f"  ✓ 데이터 시작 행: {start_row}행")
    
    # Input_data의 컬럼명 찾기 (유연한 매칭)
    def find_column(table_data, search_terms):
        """컬럼명을 유연하게 찾기"""
        if not table_data:
            return None
        for col_name in table_data[0].keys():
            col_str = str(col_name).strip().lower().replace(' ', '').replace('_', '').replace('.', '')
            for term in search_terms:
                if term.lower().replace(' ', '').replace('_', '').replace('.', '') in col_str:
                    return col_name
        return None
    
    # 매핑 정의
    column_mapping = [
        ("계약번호(Ref no.)", ["계약번호", "ref no", "refno"]),
        ("리스명", ["리스명"]),
        ("거래처", ["거래처"]),
        ("자산구분", ["자산구분"]),
        ("비용구분", ["비용구분"]),
        ("내부거래여부", ["내부거래", "내부거래여부"]),
        ("주석구분", ["주석구분"]),
        ("리스개시일", ["리스개시일"]),
        ("리스종료일", ["리스종료일"]),
        ("리스변경일", ["리스변경일", "중도해지일"]),
        ("통화", ["통화"]),
    ]
    
    # Input_data의 컬럼 찾기
    input_columns = {}
    for right_of_use_asset_header, search_terms in column_mapping:
        col = find_column(table_data, search_terms)
        if col:
            input_columns[right_of_use_asset_header] = col
        else:
            print(f"  경고: '{right_of_use_asset_header}'에 해당하는 Input_data 컬럼을 찾을 수 없습니다.")
    
    # 데이터 채우기
    date_headers = ["리스개시일", "리스종료일", "리스변경일"]
    
    for row_idx, row_data in enumerate(filtered_data):
        current_row = start_row + row_idx
        
        for right_of_use_asset_header, input_col in input_columns.items():
            # 전대사용권자산 시트의 헤더 인덱스 찾기 (기본 정보 헤더만)
            header_idx = None
            for idx, header in enumerate(BASIC_HEADERS):
                if header == right_of_use_asset_header:
                    header_idx = idx
                    break
            
            if header_idx is None:
                continue
            
            # 셀 주소 계산 (B열부터 시작)
            col_letter = get_column_letter(2 + header_idx)  # B열 = 2
            cell = f"{col_letter}{current_row}"
            
            # 값 가져오기
            value = row_data.get(input_col)
            
            # 날짜 열인 경우 날짜 형식 적용
            if right_of_use_asset_header in date_headers:
                set_cell_value(right_of_use_asset_sheet, cell, value)
                right_of_use_asset_sheet[cell].number_format = "yyyy-mm-dd"
            else:
                set_cell_value(right_of_use_asset_sheet, cell, value)
    
    print(f"  ✓ 사용권자산 시트 데이터 추가 완료: {len(filtered_data)}행")
    
    # 기초 열 채우기 (이후 증가, 감소, 기말, 상각 열이 자동으로 채워짐)
    fill_beginning_balance_column(right_of_use_asset_sheet, len(filtered_data), start_row)
    
    # 사용권자산_감가상각누계액 그룹의 기초 열 채우기
    fill_depreciation_beginning_balance_column(right_of_use_asset_sheet, len(filtered_data), start_row)
    
    # 사용권자산_감가상각누계액 그룹의 증가 열 채우기
    fill_depreciation_increase_column(right_of_use_asset_sheet, len(filtered_data), start_row)
    
    # 사용권자산_감가상각누계액 그룹의 감소 열 채우기
    fill_depreciation_decrease_column(right_of_use_asset_sheet, len(filtered_data), start_row)
    
    # 사용권자산_감가상각누계액 그룹의 기말 열 채우기
    fill_depreciation_ending_balance_column(right_of_use_asset_sheet, len(filtered_data), start_row)
    
    # 상각 열 채우기 (사용권자산_감가상각누계액 그룹)
    fill_depreciation_column(right_of_use_asset_sheet, len(filtered_data), start_row)
    
    # 임차료 열 채우기
    fill_rent_column(right_of_use_asset_sheet, len(filtered_data), start_row)
    
    # 감가상각비 그룹의 1월~12월 열 채우기
    fill_monthly_depreciation_expense_columns(right_of_use_asset_sheet, len(filtered_data), start_row)
    
    # 회계 형식 적용 (기초 열부터 마지막 12월 열까지)
    apply_accounting_format_to_right_of_use_asset(right_of_use_asset_sheet, len(filtered_data), start_row)
    
    # 테이블 범위에 테두리 적용
    apply_border_to_right_of_use_asset_table(right_of_use_asset_sheet, len(filtered_data), start_row)


def fill_beginning_balance_column(right_of_use_asset_sheet, num_rows, start_row):
    """
    사용권자산 시트의 기초 열 채우기
    
    수식: =-SUMIFS(전대_Lease_Data!$AL:$AL,전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$I:$I,EOMONTH(Input_data!$C$6,-1))
    
    Args:
        right_of_use_asset_sheet: 사용권자산 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[기초 열 채우기]")
    
    # 사용권자산 시트의 헤더 행 (7행)에서 기초 열 찾기
    header_row = 7
    beginning_balance_col_idx = None
    
    # 헤더 행에서 "기초" 열 찾기
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "기초" in str(header_value):
            beginning_balance_col_idx = col_idx - 2  # B열 기준 인덱스
            break
    
    if beginning_balance_col_idx is None:
        print("  경고: '기초' 헤더를 찾을 수 없습니다.")
        return
    
    # 계약번호 열 찾기 (B열, 인덱스 0)
    contract_no_col_idx = 0
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열
    
    # 기초 열 문자 계산
    beginning_balance_col_letter = get_column_letter(2 + beginning_balance_col_idx)
    
    # 전대_Lease_Data 시트의 열 문자
    # 계약번호(Ref no.): B열
    # 유효일자: I열
    # 차감_사용권자산: AL열 (전대_Lease_Data 시트 헤더 기준 인덱스 37, B+37=AL)
    lease_data_contract_no_col = "B"  # 전대_Lease_Data 시트의 계약번호 열
    lease_data_effective_date_col = "I"  # 전대_Lease_Data 시트의 유효일자 열
    lease_data_right_of_use_asset_col = "AL"  # 전대_Lease_Data 시트의 차감_사용권자산 열
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =-SUMIFS(전대_Lease_Data!$AL:$AL,전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$I:$I,EOMONTH(Input_data!$C$6,-1))
        
        contract_no_cell = f"{contract_no_col_letter}{row_idx}"
        beginning_balance_cell = f"{beginning_balance_col_letter}{row_idx}"
        
        formula = f"=-SUMIFS(전대_Lease_Data!${lease_data_right_of_use_asset_col}:${lease_data_right_of_use_asset_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_effective_date_col}:${lease_data_effective_date_col},EOMONTH(Input_data!$C$6,-1))"
        
        add_formula(right_of_use_asset_sheet, beginning_balance_cell, formula)
    
    print(f"  ✓ 기초 열 입력 완료: {num_rows}행")
    
    # 증가 열 채우기
    fill_increase_column(right_of_use_asset_sheet, num_rows, start_row)


def fill_increase_column(right_of_use_asset_sheet, num_rows, start_row):
    """
    사용권자산 시트의 증가 열 채우기
    
    수식: =-IF(I9>=Input_data!$C$6,SUMIFS(전대_Lease_Data!$AL:$AL,전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$N:$N, MIN(IF(전대_Lease_Data!$B:$B=B9, 전대_Lease_Data!$N:$N))),0)
    
    Args:
        right_of_use_asset_sheet: 사용권자산 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[증가 열 채우기]")
    
    # 사용권자산 시트의 헤더 행 (7행)에서 필요한 열 찾기
    header_row = 7
    contract_no_col_idx = None
    lease_start_date_col_idx = None
    increase_col_idx = None
    
    # 계약번호 열 찾기 (B열, 인덱스 0)
    contract_no_col_idx = 0
    
    # 리스개시일 열 찾기
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "리스개시일" in str(header_value):
            lease_start_date_col_idx = col_idx - 2  # B열 기준 인덱스
            break
    
    # 증가 열 찾기
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "증가" in str(header_value):
            increase_col_idx = col_idx - 2  # B열 기준 인덱스
            break
    
    if lease_start_date_col_idx is None:
        print("  경고: '리스개시일' 헤더를 찾을 수 없습니다.")
        return
    
    if increase_col_idx is None:
        print("  경고: '증가' 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열
    lease_start_date_col_letter = get_column_letter(2 + lease_start_date_col_idx)  # I열
    increase_col_letter = get_column_letter(2 + increase_col_idx)
    
    # 전대_Lease_Data 시트의 열 문자
    # 계약번호(Ref no.): B열
    # 수령회차: N열
    # 차감_사용권자산: AL열
    lease_data_contract_no_col = "B"  # 전대_Lease_Data 시트의 계약번호 열
    lease_data_receipt_count_col = "N"  # 전대_Lease_Data 시트의 수령회차 열
    lease_data_right_of_use_asset_col = "AL"  # 전대_Lease_Data 시트의 차감_사용권자산 열
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =-IF(I9>=Input_data!$C$6,SUMIFS(전대_Lease_Data!$AL:$AL,전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$N:$N, MIN(IF(전대_Lease_Data!$B:$B=B9, 전대_Lease_Data!$N:$N))),0)
        
        lease_start_date_cell = f"{lease_start_date_col_letter}{row_idx}"
        contract_no_cell = f"{contract_no_col_letter}{row_idx}"
        increase_cell = f"{increase_col_letter}{row_idx}"
        
        # 새로운 수식: IF(COUNTIFS(...) > 0, 0, 1) 사용
        formula = f"=-IF({lease_start_date_cell}>=Input_data!$C$6,SUMIFS(전대_Lease_Data!${lease_data_right_of_use_asset_col}:${lease_data_right_of_use_asset_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_receipt_count_col}:${lease_data_receipt_count_col},IF(COUNTIFS(전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_receipt_count_col}:${lease_data_receipt_count_col},0)>0,0,1)),0)"
        
        add_formula(right_of_use_asset_sheet, increase_cell, formula)
    
    print(f"  ✓ 증가 열 입력 완료: {num_rows}행")
    
    # 감소 열 채우기 (사용권자산_취득가액 그룹)
    fill_decrease_column(right_of_use_asset_sheet, num_rows, start_row)


def fill_decrease_column(right_of_use_asset_sheet, num_rows, start_row):
    """
    사용권자산 시트의 감소 열 채우기 (사용권자산_취득가액 그룹)
    
    수식: =-IF($C$4>=MIN(J9,K9),-SUMIFS(전대_Lease_Data!$AL:$AL,전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$N:$N, MIN(IF(전대_Lease_Data!$B:$B=B9, 전대_Lease_Data!$N:$N))))
    
    Args:
        right_of_use_asset_sheet: 사용권자산 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[감소 열 채우기 (취득가액 그룹)]")
    
    # 사용권자산 시트의 헤더 행 (7행)에서 필요한 열 찾기
    header_row = 7
    contract_no_col_idx = None
    lease_end_date_col_idx = None
    early_termination_date_col_idx = None
    decrease_col_idx = None
    
    # 계약번호 열 찾기 (B열, 인덱스 0)
    contract_no_col_idx = 0
    
    # 리스종료일 열 찾기
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "리스종료일" in str(header_value):
            lease_end_date_col_idx = col_idx - 2  # B열 기준 인덱스
            break
    
    # 리스변경일 열 찾기
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and ("리스변경일" in str(header_value) or "중도해지일" in str(header_value)):
            early_termination_date_col_idx = col_idx - 2  # B열 기준 인덱스
            break
    
    # 감소 열 찾기 (사용권자산_취득가액 그룹)
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "감소" in str(header_value):
            decrease_col_idx = col_idx - 2  # B열 기준 인덱스
            break
    
    if lease_end_date_col_idx is None:
        print("  경고: '리스종료일' 헤더를 찾을 수 없습니다.")
        return
    
    if decrease_col_idx is None:
        print("  경고: '감소' 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열
    lease_end_date_col_letter = get_column_letter(2 + lease_end_date_col_idx)  # J열
    early_termination_date_col_letter = None
    if early_termination_date_col_idx is not None:
        early_termination_date_col_letter = get_column_letter(2 + early_termination_date_col_idx)  # K열
    decrease_col_letter = get_column_letter(2 + decrease_col_idx)
    
    # 전대_Lease_Data 시트의 열 문자
    # 계약번호(Ref no.): B열
    # 수령회차: N열
    # 차감_사용권자산: AL열
    lease_data_contract_no_col = "B"  # 전대_Lease_Data 시트의 계약번호 열
    lease_data_receipt_count_col = "N"  # 전대_Lease_Data 시트의 수령회차 열
    lease_data_right_of_use_asset_col = "AL"  # 전대_Lease_Data 시트의 차감_사용권자산 열
    
    # 사용권자산 시트의 기준일 셀 (C4)
    base_date_cell = "C4"
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =-IF($C$4>=MIN(J9,K9),-SUMIFS(전대_Lease_Data!$AL:$AL,전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$N:$N, MIN(IF(전대_Lease_Data!$B:$B=B9, 전대_Lease_Data!$N:$N))))
        
        contract_no_cell = f"{contract_no_col_letter}{row_idx}"  # B9
        lease_end_date_cell = f"{lease_end_date_col_letter}{row_idx}"  # J9
        decrease_cell = f"{decrease_col_letter}{row_idx}"
        
        # K9 (리스변경일)이 있으면 사용, 없으면 J9만 사용
        if early_termination_date_col_letter:
            early_termination_date_cell = f"{early_termination_date_col_letter}{row_idx}"  # K9
            min_date_expr = f"MIN({lease_end_date_cell},{early_termination_date_cell})"
        else:
            min_date_expr = f"{lease_end_date_cell}"
        
        # 새로운 수식: IF(COUNTIFS(...) > 0, 0, 1) 사용
        formula = f"=SUMIFS(전대_Lease_Data!${lease_data_right_of_use_asset_col}:${lease_data_right_of_use_asset_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_receipt_count_col}:${lease_data_receipt_count_col},IF(COUNTIFS(전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_receipt_count_col}:${lease_data_receipt_count_col},0)>0,0,1))"
        
        add_formula(right_of_use_asset_sheet, decrease_cell, formula)
    
    print(f"  ✓ 감소 열 입력 완료: {num_rows}행")
    
    # 기말 열 채우기 (사용권자산_취득가액 그룹)
    fill_ending_balance_column(right_of_use_asset_sheet, num_rows, start_row)


def fill_depreciation_beginning_balance_column(right_of_use_asset_sheet, num_rows, start_row):
    """
    사용권자산 시트의 기초 열 채우기 (사용권자산_감가상각누계액 그룹)
    
    수식: =-SUMIFS(전대_Lease_Data!$AK:$AK,전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$I:$I,EOMONTH(Input_data!$C$6,-1))
    
    Args:
        right_of_use_asset_sheet: 사용권자산 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[기초 열 채우기 (감가상각누계액 그룹)]")
    
    # 사용권자산 시트의 헤더 행 (7행)에서 기초 열 찾기 (감가상각누계액 그룹)
    header_row = 7
    beginning_balance_col_idx = None
    
    # 헤더 행에서 "기초" 열 찾기 (감가상각누계액 그룹)
    # 취득가액 그룹의 기초 열을 건너뛰고 감가상각누계액 그룹의 기초 열 찾기
    found_first_beginning = False
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "기초" in str(header_value):
            if not found_first_beginning:
                # 첫 번째 "기초" 열은 취득가액 그룹
                found_first_beginning = True
            else:
                # 두 번째 "기초" 열은 감가상각누계액 그룹
                beginning_balance_col_idx = col_idx - 2  # B열 기준 인덱스
                break
    
    if beginning_balance_col_idx is None:
        print("  경고: '기초' 헤더(감가상각누계액 그룹)를 찾을 수 없습니다.")
        return
    
    # 계약번호 열 찾기 (B열, 인덱스 0)
    contract_no_col_idx = 0
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열
    
    # 기초 열 문자 계산
    beginning_balance_col_letter = get_column_letter(2 + beginning_balance_col_idx)
    
    # 전대_Lease_Data 시트의 열 문자
    # 계약번호(Ref no.): B열
    # 유효일자: I열
    # 차감_감가상각누계액: AK열
    lease_data_contract_no_col = "B"  # 전대_Lease_Data 시트의 계약번호 열
    lease_data_effective_date_col = "I"  # 전대_Lease_Data 시트의 유효일자 열
    lease_data_depreciation_accumulated_col = "AK"  # 전대_Lease_Data 시트의 차감_감가상각누계액 열
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =-SUMIFS(전대_Lease_Data!$AK:$AK,전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$I:$I,EOMONTH(Input_data!$C$6,-1))
        
        contract_no_cell = f"{contract_no_col_letter}{row_idx}"
        beginning_balance_cell = f"{beginning_balance_col_letter}{row_idx}"
        
        formula = f"=-SUMIFS(전대_Lease_Data!${lease_data_depreciation_accumulated_col}:${lease_data_depreciation_accumulated_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_effective_date_col}:${lease_data_effective_date_col},EOMONTH(Input_data!$C$6,-1))"
        
        add_formula(right_of_use_asset_sheet, beginning_balance_cell, formula)
    
    print(f"  ✓ 기초 열 입력 완료 (감가상각누계액 그룹): {num_rows}행")


def fill_depreciation_increase_column(right_of_use_asset_sheet, num_rows, start_row):
    """
    사용권자산 시트의 증가 열 채우기 (사용권자산_감가상각누계액 그룹)
    
    수식: =-IF(I9>=Input_data!$C$6,SUMIFS(전대_Lease_Data!$AK:$AK,전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$N:$N,IF(COUNTIFS(전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$N:$N,0)>0,0,1)),0)
    
    Args:
        right_of_use_asset_sheet: 사용권자산 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[증가 열 채우기 (감가상각누계액 그룹)]")
    
    # 사용권자산 시트의 헤더 행 (7행)에서 필요한 열 찾기
    header_row = 7
    contract_no_col_idx = None
    lease_start_date_col_idx = None
    increase_col_idx = None
    
    # 계약번호 열 찾기 (B열, 인덱스 0)
    contract_no_col_idx = 0
    
    # 리스개시일 열 찾기
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "리스개시일" in str(header_value):
            lease_start_date_col_idx = col_idx - 2  # B열 기준 인덱스
            break
    
    # 증가 열 찾기 (사용권자산_감가상각누계액 그룹)
    # 취득가액 그룹의 증가 열을 건너뛰고 감가상각누계액 그룹의 증가 열 찾기
    found_first_increase = False
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "증가" in str(header_value):
            if not found_first_increase:
                # 첫 번째 "증가" 열은 취득가액 그룹
                found_first_increase = True
            else:
                # 두 번째 "증가" 열은 감가상각누계액 그룹
                increase_col_idx = col_idx - 2  # B열 기준 인덱스
                break
    
    if lease_start_date_col_idx is None:
        print("  경고: '리스개시일' 헤더를 찾을 수 없습니다.")
        return
    
    if increase_col_idx is None:
        print("  경고: '증가' 헤더(감가상각누계액 그룹)를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열
    lease_start_date_col_letter = get_column_letter(2 + lease_start_date_col_idx)  # I열
    increase_col_letter = get_column_letter(2 + increase_col_idx)
    
    # 전대_Lease_Data 시트의 열 문자
    # 계약번호(Ref no.): B열
    # 수령회차: N열
    # 차감_감가상각누계액: AK열
    lease_data_contract_no_col = "B"  # 전대_Lease_Data 시트의 계약번호 열
    lease_data_receipt_count_col = "N"  # 전대_Lease_Data 시트의 수령회차 열
    lease_data_depreciation_accumulated_col = "AK"  # 전대_Lease_Data 시트의 차감_감가상각누계액 열
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =-IF(I9>=Input_data!$C$6,SUMIFS(전대_Lease_Data!$AK:$AK,전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$N:$N,IF(COUNTIFS(전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$N:$N,0)>0,0,1)),0)
        
        lease_start_date_cell = f"{lease_start_date_col_letter}{row_idx}"
        contract_no_cell = f"{contract_no_col_letter}{row_idx}"
        increase_cell = f"{increase_col_letter}{row_idx}"
        
        formula = f"=-IF({lease_start_date_cell}>=Input_data!$C$6,SUMIFS(전대_Lease_Data!${lease_data_depreciation_accumulated_col}:${lease_data_depreciation_accumulated_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_receipt_count_col}:${lease_data_receipt_count_col},IF(COUNTIFS(전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_receipt_count_col}:${lease_data_receipt_count_col},0)>0,0,1)),0)"
        
        add_formula(right_of_use_asset_sheet, increase_cell, formula)
    
    print(f"  ✓ 증가 열 입력 완료 (감가상각누계액 그룹): {num_rows}행")


def fill_depreciation_decrease_column(right_of_use_asset_sheet, num_rows, start_row):
    """
    사용권자산 시트의 감소 열 채우기 (사용권자산_감가상각누계액 그룹)
    
    수식: =IF($C$4>=MIN(J9,K9),SUMIFS(전대_Lease_Data!$AK:$AK,전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$I:$I,EOMONTH(MIN(J9,K9),0)),SUMIFS(전대_Lease_Data!$AK:$AK,전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$I:$I,EOMONTH($C$4,0)))
    
    Args:
        right_of_use_asset_sheet: 사용권자산 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[감소 열 채우기 (감가상각누계액 그룹)]")
    
    # 사용권자산 시트의 헤더 행 (7행)에서 필요한 열 찾기
    header_row = 7
    contract_no_col_idx = None
    lease_end_date_col_idx = None
    early_termination_date_col_idx = None
    decrease_col_idx = None
    
    # 계약번호 열 찾기 (B열, 인덱스 0)
    contract_no_col_idx = 0
    
    # 리스종료일 열 찾기
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "리스종료일" in str(header_value):
            lease_end_date_col_idx = col_idx - 2  # B열 기준 인덱스
            break
    
    # 리스변경일 열 찾기
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and ("리스변경일" in str(header_value) or "중도해지일" in str(header_value)):
            early_termination_date_col_idx = col_idx - 2  # B열 기준 인덱스
            break
    
    # 감소 열 찾기 (사용권자산_감가상각누계액 그룹)
    # 취득가액 그룹의 감소 열을 건너뛰고 감가상각누계액 그룹의 감소 열 찾기
    found_first_decrease = False
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "감소" in str(header_value):
            if not found_first_decrease:
                # 첫 번째 "감소" 열은 취득가액 그룹
                found_first_decrease = True
            else:
                # 두 번째 "감소" 열은 감가상각누계액 그룹
                decrease_col_idx = col_idx - 2  # B열 기준 인덱스
                break
    
    if lease_end_date_col_idx is None:
        print("  경고: '리스종료일' 헤더를 찾을 수 없습니다.")
        return
    
    if decrease_col_idx is None:
        print("  경고: '감소' 헤더(감가상각누계액 그룹)를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열
    lease_end_date_col_letter = get_column_letter(2 + lease_end_date_col_idx)  # J열
    early_termination_date_col_letter = None
    if early_termination_date_col_idx is not None:
        early_termination_date_col_letter = get_column_letter(2 + early_termination_date_col_idx)  # K열
    decrease_col_letter = get_column_letter(2 + decrease_col_idx)
    
    # 전대_Lease_Data 시트의 열 문자
    # 계약번호(Ref no.): B열
    # 유효일자: I열
    # 차감_감가상각누계액: AK열
    lease_data_contract_no_col = "B"  # 전대_Lease_Data 시트의 계약번호 열
    lease_data_effective_date_col = "I"  # 전대_Lease_Data 시트의 유효일자 열
    lease_data_depreciation_accumulated_col = "AK"  # 전대_Lease_Data 시트의 차감_감가상각누계액 열
    
    # 사용권자산 시트의 기준일 셀 (C4)
    base_date_cell = "C4"
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =IF($C$4>=MIN(J9,K9),SUMIFS(전대_Lease_Data!$AK:$AK,전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$I:$I,EOMONTH(MIN(J9,K9),0)),SUMIFS(전대_Lease_Data!$AK:$AK,전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$I:$I,EOMONTH($C$4,0)))
        
        contract_no_cell = f"{contract_no_col_letter}{row_idx}"  # B9
        lease_end_date_cell = f"{lease_end_date_col_letter}{row_idx}"  # J9
        decrease_cell = f"{decrease_col_letter}{row_idx}"
        
        # K9 (리스변경일)이 있으면 사용, 없으면 J9만 사용
        if early_termination_date_col_letter:
            early_termination_date_cell = f"{early_termination_date_col_letter}{row_idx}"  # K9
            min_date_expr = f"MIN({lease_end_date_cell},{early_termination_date_cell})"
        else:
            min_date_expr = f"{lease_end_date_cell}"
        
        formula = f"=IF(${base_date_cell}>={min_date_expr},SUMIFS(전대_Lease_Data!${lease_data_depreciation_accumulated_col}:${lease_data_depreciation_accumulated_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_effective_date_col}:${lease_data_effective_date_col},EOMONTH({min_date_expr},0)),SUMIFS(전대_Lease_Data!${lease_data_depreciation_accumulated_col}:${lease_data_depreciation_accumulated_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_effective_date_col}:${lease_data_effective_date_col},EOMONTH(${base_date_cell},0)))"
        
        add_formula(right_of_use_asset_sheet, decrease_cell, formula)
    
    print(f"  ✓ 감소 열 입력 완료 (감가상각누계액 그룹): {num_rows}행")


def fill_depreciation_ending_balance_column(right_of_use_asset_sheet, num_rows, start_row):
    """
    사용권자산 시트의 기말 열 채우기 (사용권자산_감가상각누계액 그룹)
    
    수식: =SUM(R9:V9)
    R9: 동일행의 기초 열 (감가상각누계액 그룹)
    V9: 동일행의 감소 열 (감가상각누계액 그룹)
    
    Args:
        right_of_use_asset_sheet: 사용권자산 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[기말 열 채우기 (감가상각누계액 그룹)]")
    
    # 사용권자산 시트의 헤더 행 (7행)에서 필요한 열 찾기
    header_row = 7
    beginning_balance_col_idx = None
    decrease_col_idx = None
    ending_balance_col_idx = None
    
    # 기초 열 찾기 (감가상각누계액 그룹)
    found_first_beginning = False
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "기초" in str(header_value):
            if not found_first_beginning:
                # 첫 번째 "기초" 열은 취득가액 그룹
                found_first_beginning = True
            else:
                # 두 번째 "기초" 열은 감가상각누계액 그룹
                beginning_balance_col_idx = col_idx - 2  # B열 기준 인덱스
                break
    
    # 감소 열 찾기 (사용권자산_감가상각누계액 그룹)
    found_first_decrease = False
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "감소" in str(header_value):
            if not found_first_decrease:
                # 첫 번째 "감소" 열은 취득가액 그룹
                found_first_decrease = True
            else:
                # 두 번째 "감소" 열은 감가상각누계액 그룹
                decrease_col_idx = col_idx - 2  # B열 기준 인덱스
                break
    
    # 기말 열 찾기 (사용권자산_감가상각누계액 그룹)
    found_first_ending = False
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "기말" in str(header_value):
            if not found_first_ending:
                # 첫 번째 "기말" 열은 취득가액 그룹
                found_first_ending = True
            else:
                # 두 번째 "기말" 열은 감가상각누계액 그룹
                ending_balance_col_idx = col_idx - 2  # B열 기준 인덱스
                break
    
    if beginning_balance_col_idx is None:
        print("  경고: '기초' 헤더(감가상각누계액 그룹)를 찾을 수 없습니다.")
        return
    
    if decrease_col_idx is None:
        print("  경고: '감소' 헤더(감가상각누계액 그룹)를 찾을 수 없습니다.")
        return
    
    if ending_balance_col_idx is None:
        print("  경고: '기말' 헤더(감가상각누계액 그룹)를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산
    beginning_balance_col_letter = get_column_letter(2 + beginning_balance_col_idx)  # R열
    decrease_col_letter = get_column_letter(2 + decrease_col_idx)  # V열
    ending_balance_col_letter = get_column_letter(2 + ending_balance_col_idx)
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =SUM(R9:V9)
        
        beginning_balance_cell = f"{beginning_balance_col_letter}{row_idx}"
        decrease_cell = f"{decrease_col_letter}{row_idx}"
        ending_balance_cell = f"{ending_balance_col_letter}{row_idx}"
        
        formula = f"=SUM({beginning_balance_cell}:{decrease_cell})"
        
        add_formula(right_of_use_asset_sheet, ending_balance_cell, formula)
    
    print(f"  ✓ 기말 열 입력 완료 (감가상각누계액 그룹): {num_rows}행")


def fill_depreciation_column(right_of_use_asset_sheet, num_rows, start_row):
    """
    사용권자산 시트의 상각 열 채우기 (사용권자산_감가상각누계액 그룹)
    
    수식: =-SUMIFS(전대_Lease_Data!$AJ:$AJ,전대_Lease_Data!$B:$B,$B9,전대_Lease_Data!$I:$I,"<="&EOMONTH(MIN($J9,$K9,$C$4),0),전대_Lease_Data!$I:$I,">="&Input_data!$C$6)
    
    Args:
        right_of_use_asset_sheet: 사용권자산 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[상각 열 채우기 (감가상각누계액 그룹)]")
    
    # 사용권자산 시트의 헤더 행 (7행)에서 필요한 열 찾기
    header_row = 7
    contract_no_col_idx = None
    lease_end_date_col_idx = None
    early_termination_date_col_idx = None
    depreciation_col_idx = None
    
    # 계약번호 열 찾기 (B열, 인덱스 0)
    contract_no_col_idx = 0
    
    # 리스종료일 열 찾기
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "리스종료일" in str(header_value):
            lease_end_date_col_idx = col_idx - 2  # B열 기준 인덱스
            break
    
    # 리스변경일 열 찾기
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and ("리스변경일" in str(header_value) or "중도해지일" in str(header_value)):
            early_termination_date_col_idx = col_idx - 2  # B열 기준 인덱스
            break
    
    # 상각 열 찾기 (사용권자산_감가상각누계액 그룹)
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "상각" in str(header_value):
            # 이미 찾은 상각 열이 있는지 확인 (감가상각누계액 그룹의 상각 열)
            # 취득가액 그룹의 상각 열과 구분하기 위해 위치나 컨텍스트를 확인해야 할 수도 있음
            # 일단 첫 번째 "상각" 열을 찾되, 나중에 필요하면 수정
            depreciation_col_idx = col_idx - 2  # B열 기준 인덱스
            # 감가상각누계액 그룹의 상각 열은 보통 취득가액 그룹보다 오른쪽에 위치
            # 일단 찾은 것을 사용하되, 사용자가 정확한 위치를 알려주면 수정 가능
            break
    
    if lease_end_date_col_idx is None:
        print("  경고: '리스종료일' 헤더를 찾을 수 없습니다.")
        return
    
    if depreciation_col_idx is None:
        print("  경고: '상각' 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열
    lease_end_date_col_letter = get_column_letter(2 + lease_end_date_col_idx)  # J열
    early_termination_date_col_letter = None
    if early_termination_date_col_idx is not None:
        early_termination_date_col_letter = get_column_letter(2 + early_termination_date_col_idx)  # K열
    depreciation_col_letter = get_column_letter(2 + depreciation_col_idx)
    
    # 전대_Lease_Data 시트의 열 문자
    # 계약번호(Ref no.): B열
    # 유효일자: I열
    # 차감_감가상각비: AJ열
    lease_data_contract_no_col = "B"  # 전대_Lease_Data 시트의 계약번호 열
    lease_data_effective_date_col = "I"  # 전대_Lease_Data 시트의 유효일자 열
    lease_data_depreciation_col = "AJ"  # 전대_Lease_Data 시트의 차감_감가상각비 열
    
    # 사용권자산 시트의 기준일 셀 (C4)
    base_date_cell = "C4"
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =-SUMIFS(전대_Lease_Data!$AJ:$AJ,전대_Lease_Data!$B:$B,$B9,전대_Lease_Data!$I:$I,"<="&EOMONTH(MIN($J9,$K9,$C$4),0),전대_Lease_Data!$I:$I,">="&Input_data!$C$6)
        
        contract_no_cell = f"${contract_no_col_letter}${row_idx}"  # $B9 (절대 열 참조)
        lease_end_date_cell = f"${lease_end_date_col_letter}${row_idx}"  # $J9 (절대 열 참조)
        depreciation_cell = f"{depreciation_col_letter}{row_idx}"
        
        # K9 (리스변경일)이 있으면 사용, 없으면 J9만 사용
        if early_termination_date_col_letter:
            early_termination_date_cell = f"${early_termination_date_col_letter}${row_idx}"  # $K9 (절대 열 참조)
            min_date_expr = f"MIN({lease_end_date_cell},{early_termination_date_cell},${base_date_cell})"
        else:
            min_date_expr = f"MIN({lease_end_date_cell},${base_date_cell})"
        
        formula = f"=-SUMIFS(전대_Lease_Data!${lease_data_depreciation_col}:${lease_data_depreciation_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_effective_date_col}:${lease_data_effective_date_col},\"<=\"&EOMONTH({min_date_expr},0),전대_Lease_Data!${lease_data_effective_date_col}:${lease_data_effective_date_col},\">=\"&Input_data!$C$6)"
        
        add_formula(right_of_use_asset_sheet, depreciation_cell, formula)
    
    print(f"  ✓ 상각 열 입력 완료: {num_rows}행")


def fill_rent_column(right_of_use_asset_sheet, num_rows, start_row):
    """
    전대사용권자산 시트의 임차료 열 채우기 (사용권자산_장부가액 그룹)
    
    수식: =Q9-W9-AB9
    Q9: 사용권자산_취득가액 그룹의 기말 열
    W9: 사용권자산_감가상각누계액 그룹의 기말 열
    AB9: 사용권자산_손상차손누계액 그룹의 기말 열
    
    Args:
        right_of_use_asset_sheet: 전대사용권자산 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[임차료 열 채우기]")
    
    # 사용권자산 시트의 헤더 행 (7행)에서 필요한 열 찾기
    header_row = 7
    acquisition_ending_col_idx = None
    depreciation_ending_col_idx = None
    impairment_ending_col_idx = None
    rent_col_idx = None
    
    # 사용권자산_취득가액 그룹의 기말 열 찾기 (첫 번째 "기말")
    ending_count = 0
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "기말" in str(header_value):
            ending_count += 1
            if ending_count == 1:
                # 첫 번째 "기말" 열은 취득가액 그룹
                acquisition_ending_col_idx = col_idx - 2  # B열 기준 인덱스
            elif ending_count == 2:
                # 두 번째 "기말" 열은 감가상각누계액 그룹
                depreciation_ending_col_idx = col_idx - 2  # B열 기준 인덱스
            elif ending_count == 3:
                # 세 번째 "기말" 열은 손상차손누계액 그룹
                impairment_ending_col_idx = col_idx - 2  # B열 기준 인덱스
                break
    
    # 임차료 열 찾기
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "임차료" in str(header_value):
            rent_col_idx = col_idx - 2  # B열 기준 인덱스
            break
    
    if acquisition_ending_col_idx is None:
        print("  경고: '기말' 헤더(취득가액 그룹)를 찾을 수 없습니다.")
        return
    
    if depreciation_ending_col_idx is None:
        print("  경고: '기말' 헤더(감가상각누계액 그룹)를 찾을 수 없습니다.")
        return
    
    if impairment_ending_col_idx is None:
        print("  경고: '기말' 헤더(손상차손누계액 그룹)를 찾을 수 없습니다.")
        return
    
    if rent_col_idx is None:
        print("  경고: '임차료' 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산
    acquisition_ending_col_letter = get_column_letter(2 + acquisition_ending_col_idx)  # Q열
    depreciation_ending_col_letter = get_column_letter(2 + depreciation_ending_col_idx)  # W열
    impairment_ending_col_letter = get_column_letter(2 + impairment_ending_col_idx)  # AB열
    rent_col_letter = get_column_letter(2 + rent_col_idx)
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =Q9-W9-AB9
        
        acquisition_ending_cell = f"{acquisition_ending_col_letter}{row_idx}"
        depreciation_ending_cell = f"{depreciation_ending_col_letter}{row_idx}"
        impairment_ending_cell = f"{impairment_ending_col_letter}{row_idx}"
        rent_cell = f"{rent_col_letter}{row_idx}"
        
        formula = f"={acquisition_ending_cell}-{depreciation_ending_cell}-{impairment_ending_cell}"
        
        add_formula(right_of_use_asset_sheet, rent_cell, formula)
    
    print(f"  ✓ 임차료 열 입력 완료: {num_rows}행")


def fill_monthly_depreciation_expense_columns(right_of_use_asset_sheet, num_rows, start_row):
    """
    전대사용권자산 시트의 감가상각비 그룹의 1월~12월 열 채우기
    
    수식: =-SUMIFS(전대_Lease_Data!$AJ:$AJ,전대_Lease_Data!$B:$B,$B9,전대_Lease_Data!$I:$I,"<="&EOMONTH(MIN($J9,$K9),0),전대_Lease_Data!$J:$J,YEAR($C$4),전대_Lease_Data!$K:$K,SUBSTITUTE(AF7,"월",""))
    
    Args:
        right_of_use_asset_sheet: 전대사용권자산 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[감가상각비 그룹의 1월~12월 열 채우기]")
    
    # 사용권자산 시트의 헤더 행 (7행)에서 필요한 열 찾기
    header_row = 7
    contract_no_col_idx = None
    lease_end_date_col_idx = None
    early_termination_date_col_idx = None
    monthly_col_indices = []  # 1월~12월 열 인덱스
    
    # 계약번호 열 찾기 (B열, 인덱스 0)
    contract_no_col_idx = 0
    
    # 리스종료일 열 찾기
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "리스종료일" in str(header_value):
            lease_end_date_col_idx = col_idx - 2  # B열 기준 인덱스
            break
    
    # 리스변경일 열 찾기
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "리스변경일" in str(header_value):
            early_termination_date_col_idx = col_idx - 2  # B열 기준 인덱스
            break
    
    # 1월~12월 열 찾기
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value:
            header_str = str(header_value).strip()
            # "1월", "2월", ..., "12월" 형식 확인
            if header_str.endswith("월") and header_str[:-1].isdigit():
                month_num = int(header_str[:-1])
                if 1 <= month_num <= 12:
                    monthly_col_indices.append((col_idx - 2, col_letter, month_num))  # B열 기준 인덱스, 열 문자, 월 번호
    
    if contract_no_col_idx is None:
        print("  경고: '계약번호(Ref no.)' 헤더를 찾을 수 없습니다.")
        return
    
    if lease_end_date_col_idx is None:
        print("  경고: '리스종료일' 헤더를 찾을 수 없습니다.")
        return
    
    if early_termination_date_col_idx is None:
        print("  경고: '리스변경일' 헤더를 찾을 수 없습니다.")
        return
    
    if len(monthly_col_indices) == 0:
        print("  경고: '1월'~'12월' 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열
    lease_end_date_col_letter = get_column_letter(2 + lease_end_date_col_idx)
    early_termination_date_col_letter = get_column_letter(2 + early_termination_date_col_idx)
    
    # 전대_Lease_Data 시트의 열 참조
    # 전대_Lease_Data 시트의 헤더를 확인하여 열 인덱스 찾기
    from excel_utils import get_sheet_by_name
    wb = right_of_use_asset_sheet.parent
    lease_data_sheet = get_sheet_by_name(wb, "전대_Lease_Data")
    
    if lease_data_sheet is None:
        print("  경고: '전대_Lease_Data' 시트를 찾을 수 없습니다.")
        return
    
    # 전대_Lease_Data 시트의 헤더 행 찾기 (6행)
    lease_data_header_row = 6
    lease_data_contract_no_col = None
    lease_data_effective_date_col = None
    lease_data_fiscal_year_col = None
    lease_data_fiscal_month_col = None
    lease_data_depreciation_col = None
    
    # 전대_Lease_Data 시트의 헤더 찾기
    for col_idx in range(1, lease_data_sheet.max_column + 1):  # A열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = lease_data_sheet[f"{col_letter}{lease_data_header_row}"]
        header_value = header_cell.value
        if header_value:
            header_str = str(header_value).strip()
            # 계약번호(Ref no.) 열 찾기 - "상위리스"가 포함되지 않은 것
            if "계약번호" in header_str and "Ref no" in header_str and "상위리스" not in header_str:
                lease_data_contract_no_col = col_letter
            elif "유효일자" in header_str:
                lease_data_effective_date_col = col_letter
            elif "회계연도" in header_str:
                lease_data_fiscal_year_col = col_letter
            elif "결산월" in header_str:
                lease_data_fiscal_month_col = col_letter
            elif "차감_감가상각비" in header_str:
                lease_data_depreciation_col = col_letter
    
    if lease_data_contract_no_col is None:
        print("  경고: 전대_Lease_Data 시트에서 '계약번호(Ref no.)' 헤더를 찾을 수 없습니다.")
        return
    
    if lease_data_effective_date_col is None:
        print("  경고: 전대_Lease_Data 시트에서 '유효일자' 헤더를 찾을 수 없습니다.")
        return
    
    if lease_data_fiscal_year_col is None:
        print("  경고: 전대_Lease_Data 시트에서 '회계연도' 헤더를 찾을 수 없습니다.")
        return
    
    if lease_data_fiscal_month_col is None:
        print("  경고: 전대_Lease_Data 시트에서 '결산월' 헤더를 찾을 수 없습니다.")
        return
    
    if lease_data_depreciation_col is None:
        print("  경고: 전대_Lease_Data 시트에서 '차감_감가상각비' 헤더를 찾을 수 없습니다.")
        return
    
    # 기준일 셀 (C4)
    base_date_cell = "C4"
    
    # 각 행에 대해 1월~12월 열에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        contract_no_cell = f"${contract_no_col_letter}${row_idx}"  # $B9 형식
        lease_end_date_cell = f"${lease_end_date_col_letter}${row_idx}"  # $J9 형식
        early_termination_date_cell = f"${early_termination_date_col_letter}${row_idx}"  # $K9 형식
        
        # 각 월별 열에 수식 입력
        for col_idx_offset, col_letter, month_num in monthly_col_indices:
            monthly_cell = f"{col_letter}{row_idx}"
            header_cell_ref = f"{col_letter}{header_row}"  # 헤더 행 참조 (예: AF7)
            
            # 수식 구성
            # =-SUMIFS(전대_Lease_Data!$AJ:$AJ,전대_Lease_Data!$B:$B,$B9,전대_Lease_Data!$I:$I,"<="&EOMONTH(MIN($J9,$K9),0),전대_Lease_Data!$J:$J,YEAR($C$4),전대_Lease_Data!$K:$K,SUBSTITUTE(AF7,"월",""))
            formula = f"=-SUMIFS(전대_Lease_Data!${lease_data_depreciation_col}:${lease_data_depreciation_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_effective_date_col}:${lease_data_effective_date_col},\"<=\"&EOMONTH(MIN({lease_end_date_cell},{early_termination_date_cell}),0),전대_Lease_Data!${lease_data_fiscal_year_col}:${lease_data_fiscal_year_col},YEAR(${base_date_cell}),전대_Lease_Data!${lease_data_fiscal_month_col}:${lease_data_fiscal_month_col},SUBSTITUTE({header_cell_ref},\"월\",\"\"))"
            
            add_formula(right_of_use_asset_sheet, monthly_cell, formula)
    
    print(f"  ✓ 감가상각비 그룹의 1월~12월 열 입력 완료: {num_rows}행")


def apply_accounting_format_to_right_of_use_asset(right_of_use_asset_sheet, num_rows, start_row):
    """
    전대사용권자산 시트의 기초 열부터 마지막 12월 열까지 회계 형식 적용
    
    Args:
        right_of_use_asset_sheet: 전대사용권자산 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행
    """
    from openpyxl.utils import get_column_letter
    
    print("\n[회계 형식 적용]")
    
    # 헤더 행 (7행)에서 기초 열과 마지막 12월 열 찾기
    header_row = 7
    beginning_balance_col_idx = None
    last_dec_col_idx = None
    
    # 첫 번째 "기초" 열 찾기 (사용권자산_취득가액 그룹)
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "기초" in str(header_value):
            beginning_balance_col_idx = col_idx - 2  # B열 기준 인덱스
            break
    
    if beginning_balance_col_idx is None:
        print("  경고: '기초' 헤더를 찾을 수 없습니다.")
        return
    
    # 마지막 "12월" 열 찾기 (감가상각비 그룹)
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value:
            header_str = str(header_value).strip()
            # "12월" 형식 확인
            if header_str == "12월":
                last_dec_col_idx = col_idx - 2  # B열 기준 인덱스
    
    if last_dec_col_idx is None:
        print("  경고: 마지막 '12월' 헤더를 찾을 수 없습니다.")
        return
    
    # 회계 형식 (#,##0)
    accounting_format = "#,##0"
    
    # 기초 열부터 마지막 12월 열까지 회계 형식 적용
    for col_idx in range(beginning_balance_col_idx, last_dec_col_idx + 1):
        col_letter = get_column_letter(2 + col_idx)  # B열부터 시작
        for row_idx in range(start_row, start_row + num_rows):
            cell = right_of_use_asset_sheet[f"{col_letter}{row_idx}"]
            cell.number_format = accounting_format
    
    num_cols = last_dec_col_idx - beginning_balance_col_idx + 1
    print(f"  ✓ 회계 형식 적용 완료: {num_cols}개 열 ({num_rows}행)")


def apply_border_to_right_of_use_asset_table(right_of_use_asset_sheet, num_rows, start_row):
    """
    전대사용권자산 시트의 테이블 범위에 모든 테두리 적용
    
    Args:
        right_of_use_asset_sheet: 전대사용권자산 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (헤더는 7행)
    """
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter
    
    # 테두리 스타일 정의
    thin_border = Side(style='thin')
    border = Border(
        left=thin_border,
        right=thin_border,
        top=thin_border,
        bottom=thin_border
    )
    
    # 셀병합된 행 (6행)부터 마지막 데이터 행까지
    merged_row = 6  # 셀병합된 행
    header_row = 7  # 헤더 행
    num_cols = len(HEADERS)  # 헤더 개수
    start_col = 2  # B열부터 시작
    
    # 셀병합된 행(6행)에 테두리 적용
    for col_idx in range(start_col, start_col + num_cols):
        col_letter = get_column_letter(col_idx)
        cell = right_of_use_asset_sheet[f"{col_letter}{merged_row}"]
        cell.border = border
    
    # 헤더 행(7행)부터 마지막 데이터 행까지 테두리 적용
    for row_idx in range(header_row, start_row + num_rows):
        for col_idx in range(start_col, start_col + num_cols):
            col_letter = get_column_letter(col_idx)
            cell = right_of_use_asset_sheet[f"{col_letter}{row_idx}"]
            cell.border = border
    
    print(f"  ✓ 테두리 적용 완료: {num_cols}개 열, {num_rows + 2}행 (셀병합 행 + 헤더 포함)")


def fill_ending_balance_column(right_of_use_asset_sheet, num_rows, start_row):
    """
    사용권자산 시트의 기말 열 채우기 (사용권자산_취득가액 그룹)
    
    수식: =SUM(M9:P9)
    M9: 동일행의 기초 열
    P9: 동일행의 감소 열
    
    Args:
        right_of_use_asset_sheet: 사용권자산 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[기말 열 채우기 (취득가액 그룹)]")
    
    # 사용권자산 시트의 헤더 행 (7행)에서 필요한 열 찾기
    header_row = 7
    beginning_balance_col_idx = None
    decrease_col_idx = None
    ending_balance_col_idx = None
    
    # 기초 열 찾기
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "기초" in str(header_value):
            beginning_balance_col_idx = col_idx - 2  # B열 기준 인덱스
            break
    
    # 감소 열 찾기 (사용권자산_취득가액 그룹)
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "감소" in str(header_value):
            decrease_col_idx = col_idx - 2  # B열 기준 인덱스
            break
    
    # 기말 열 찾기 (사용권자산_취득가액 그룹)
    for col_idx in range(2, right_of_use_asset_sheet.max_column + 1):  # B열부터 시작
        col_letter = get_column_letter(col_idx)
        header_cell = right_of_use_asset_sheet[f"{col_letter}{header_row}"]
        header_value = header_cell.value
        if header_value and "기말" in str(header_value):
            # 취득가액 그룹의 기말 열인지 확인 필요
            # 일단 첫 번째 "기말" 열을 찾되, 나중에 필요하면 수정
            ending_balance_col_idx = col_idx - 2  # B열 기준 인덱스
            break
    
    if beginning_balance_col_idx is None:
        print("  경고: '기초' 헤더를 찾을 수 없습니다.")
        return
    
    if decrease_col_idx is None:
        print("  경고: '감소' 헤더를 찾을 수 없습니다.")
        return
    
    if ending_balance_col_idx is None:
        print("  경고: '기말' 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산
    beginning_balance_col_letter = get_column_letter(2 + beginning_balance_col_idx)  # M열
    decrease_col_letter = get_column_letter(2 + decrease_col_idx)  # P열
    ending_balance_col_letter = get_column_letter(2 + ending_balance_col_idx)
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =SUM(M9:P9)
        
        beginning_balance_cell = f"{beginning_balance_col_letter}{row_idx}"
        decrease_cell = f"{decrease_col_letter}{row_idx}"
        ending_balance_cell = f"{ending_balance_col_letter}{row_idx}"
        
        formula = f"=SUM({beginning_balance_cell}:{decrease_cell})"
        
        add_formula(right_of_use_asset_sheet, ending_balance_cell, formula)
    
    print(f"  ✓ 기말 열 입력 완료: {num_rows}행")

