"""
리스채권 시트 생성 모듈
- 리스채권 시트 생성 및 헤더 설정
"""

from excel_utils import add_sheet, set_cell_value, add_formula
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


# ============================================================================
# 헤더 설정 (이미지 순서대로, 총 43개)
# ============================================================================

HEADERS = [
    "계약번호(Ref no.)",
    "리스명",
    "거래처",
    "자산구분",
    "비용구분",
    "내부거래여부",
    "주석구분",
    "전대_리스개시일",
    "전대_리스종료일",
    "전대_중도해지일",
    "기초",
    "취득",
    "원금상환",
    "이자수익",
    "변경",
    "중도해지",
    "기말",
    "유동",
    "비유동",
    "1월",
    "2월",
    "3월",
    "4월",
    "5월",
    "6월",
    "7월",
    "8월",
    "9월",
    "10월",
    "11월",
    "12월",
    "1월",
    "2월",
    "3월",
    "4월",
    "5월",
    "6월",
    "7월",
    "8월",
    "9월",
    "10월",
    "11월",
    "12월",
]


# ============================================================================
# 리스채권 시트 생성 함수
# ============================================================================

def create_lease_receivable_sheet(wb):
    """
    리스채권 시트 생성 및 기본 설정
    
    Args:
        wb: Workbook 객체
    
    Returns:
        Worksheet: 생성된 시트 객체
    """
    print("\n[리스채권 시트 생성]")
    
    # 시트 생성
    sheet_name = "리스채권"
    ws = add_sheet(wb, sheet_name)
    
    # B2셀: 리스채권
    set_cell_value(ws, "B2", "리스채권")
    
    # B4셀: 기준일
    set_cell_value(ws, "B4", "기준일")
    
    # C4셀: =Input_data!$C$7
    add_formula(ws, "C4", "=Input_data!$C$7")
    # C4셀에 날짜 형식 적용
    ws["C4"].number_format = "yyyy-mm-dd"
    
    # 6행: 셀 병합 및 그룹 헤더 설정 (가운데 정렬)
    # B6~K6: 기본정보 (10개 열)
    ws.merge_cells("B6:K6")
    set_cell_value(ws, "B6", "기본정보")
    ws["B6"].alignment = Alignment(horizontal='center', vertical='center')
    
    # L6~T6: 리스채권 (9개 열)
    ws.merge_cells("L6:T6")
    set_cell_value(ws, "L6", "리스채권")
    ws["L6"].alignment = Alignment(horizontal='center', vertical='center')
    
    # U6~AF6: 이자수익 (18개 열)
    ws.merge_cells("U6:AF6")
    set_cell_value(ws, "U6", "이자수익")
    ws["U6"].alignment = Alignment(horizontal='center', vertical='center')
    
    # AG6~AR6: 리스료 (12개 열)
    ws.merge_cells("AG6:AR6")
    set_cell_value(ws, "AG6", "리스료")
    ws["AG6"].alignment = Alignment(horizontal='center', vertical='center')
    
    # 7행: 헤더 행 설정 (B7부터 시작)
    for idx, header in enumerate(HEADERS):
        col_letter = get_column_letter(2 + idx)  # B열부터 시작 (2 = B)
        cell = f"{col_letter}7"
        set_cell_value(ws, cell, header)
    
    print(f"  ✓ 리스채권 시트 생성 완료")
    print(f"  ✓ 헤더 설정 완료: {len(HEADERS)}개 열")
    
    return ws


def fill_lease_receivable_sheet(wb, data_manager):
    """
    리스채권 시트에 데이터 채우기
    Input_data 시트의 전대 리스 데이터를 리스채권 시트에 채움
    
    Args:
        wb: Workbook 객체
        data_manager: DataManager 객체
    """
    from excel_utils import get_sheet_by_name, get_table_data, set_cell_value
    from openpyxl.utils import get_column_letter
    
    print("\n[리스채권 시트 데이터 채우기]")
    
    # 리스채권 시트 확인
    lease_receivable_sheet = get_sheet_by_name(wb, "리스채권")
    if lease_receivable_sheet is None:
        print("  경고: '리스채권' 시트를 찾을 수 없습니다.")
        return
    
    # Input_data 시트 확인
    input_data_sheet = get_sheet_by_name(wb, "Input_data")
    if input_data_sheet is None:
        print("  경고: 'Input_data' 시트를 찾을 수 없습니다.")
        return
    
    # input_table 테이블에서 데이터 읽기
    try:
        table_data = get_table_data(input_data_sheet, "input_table")
    except Exception as e:
        print(f"  경고: input_table 데이터 읽기 실패: {e}")
        return
    
    # 전대 리스 데이터 필터링 (data_manager의 계약번호 목록 사용)
    filtered_data = []
    contract_nos = set(data_manager.contract_data.keys())
    
    # 계약번호 컬럼 찾기
    ref_no_col = None
    if table_data:
        for col_name in table_data[0].keys():
            col_lower = str(col_name).lower().replace(' ', '').replace('_', '').replace('.', '')
            if 'refno' in col_lower or '계약번호' in str(col_name):
                ref_no_col = col_name
                break
    
    if ref_no_col is None:
        print("  경고: Input_data에서 계약번호 컬럼을 찾을 수 없습니다.")
        return
    
    # 전대 리스 데이터 필터링
    for row in table_data:
        contract_no = row.get(ref_no_col)
        if contract_no in contract_nos:
            filtered_data.append(row)
    
    if len(filtered_data) == 0:
        print("  경고: 필터링된 데이터가 없습니다.")
        return
    
    print(f"  ✓ 전대 리스 데이터 필터링 완료: {len(filtered_data)}행")
    
    # 리스채권 시트의 헤더와 Input_data의 컬럼 매핑
    # 이미지에서 보이는 열들:
    # 1. 계약번호 (Ref no.) -> 계약번호(Ref no.)
    # 2. 리스명 -> 리스명
    # 3. 거래처 -> 거래처
    # 4. 자산구분 -> 자산구분
    # 5. 비용구분 -> 비용구분
    # 6. 내부거래여부 -> 내부거래여부
    # 7. 주석구분 -> 주석구분
    # 8. 리스개시일 -> 전대_리스개시일
    # 9. 리스종료일 -> 전대_리스종료일
    # 10. 리스변경일 -> 전대_중도해지일
    
    # Input_data의 컬럼명 찾기 (유연한 매칭)
    def find_column(table_data, search_terms):
        """컬럼명을 유연하게 찾기"""
        if not table_data:
            return None
        for col_name in table_data[0].keys():
            col_str = str(col_name).strip().lower().replace(' ', '').replace('_', '').replace('.', '')
            for term in search_terms:
                if term.lower().replace(' ', '').replace('_', '').replace('.', '') in col_str:
                    return col_name
        return None
    
    # 매핑 정의
    column_mapping = [
        ("계약번호(Ref no.)", ["계약번호", "ref no", "refno"]),
        ("리스명", ["리스명"]),
        ("거래처", ["거래처"]),
        ("자산구분", ["자산구분"]),
        ("비용구분", ["비용구분"]),
        ("내부거래여부", ["내부거래", "내부거래여부"]),
        ("주석구분", ["주석구분"]),
        ("전대_리스개시일", ["리스개시일"]),
        ("전대_리스종료일", ["리스종료일"]),
        ("전대_중도해지일", ["리스변경일", "중도해지일"]),
    ]
    
    # Input_data의 컬럼 찾기
    input_columns = {}
    for lease_receivable_header, search_terms in column_mapping:
        col = find_column(table_data, search_terms)
        if col:
            input_columns[lease_receivable_header] = col
        else:
            print(f"  경고: '{lease_receivable_header}'에 해당하는 Input_data 컬럼을 찾을 수 없습니다.")
    
    # 데이터 채우기 (8행부터 시작)
    start_row = 8
    date_headers = ["전대_리스개시일", "전대_리스종료일", "전대_중도해지일"]
    
    for row_idx, row_data in enumerate(filtered_data):
        current_row = start_row + row_idx
        
        for lease_receivable_header, input_col in input_columns.items():
            # 리스채권 시트의 헤더 인덱스 찾기
            header_idx = None
            for idx, header in enumerate(HEADERS):
                if header == lease_receivable_header:
                    header_idx = idx
                    break
            
            if header_idx is None:
                continue
            
            # 셀 주소 계산 (B열부터 시작)
            col_letter = get_column_letter(2 + header_idx)  # B열 = 2
            cell = f"{col_letter}{current_row}"
            
            # 값 가져오기
            value = row_data.get(input_col)
            
            # 날짜 열인 경우 날짜 형식 적용
            if lease_receivable_header in date_headers:
                set_cell_value(lease_receivable_sheet, cell, value)
                lease_receivable_sheet[cell].number_format = "yyyy-mm-dd"
            else:
                set_cell_value(lease_receivable_sheet, cell, value)
    
    print(f"  ✓ 리스채권 시트 데이터 채우기 완료: {len(filtered_data)}행")
    
    # 기초 열 채우기
    fill_beginning_balance_column(lease_receivable_sheet, len(filtered_data), start_row)


def fill_beginning_balance_column(lease_receivable_sheet, num_rows, start_row=8):
    """
    리스채권 시트의 기초 열 채우기
    
    수식: =IF(I9>=Input_data!$C$6,0,SUMIFS(전대_Lease_Data!$Z:$Z,전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$I:$I,EOMONTH(Input_data!$C$6,-1)))
    
    Args:
        lease_receivable_sheet: 리스채권 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8)
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[기초 열 채우기]")
    
    # 리스채권 시트의 헤더 인덱스 찾기
    # 계약번호(Ref no.): B열 (인덱스 0)
    # 전대_리스개시일: I열 (인덱스 7)
    # 기초: K열 (인덱스 10)
    
    contract_no_col_idx = None
    lease_start_date_col_idx = None
    beginning_balance_col_idx = None
    
    for idx, header in enumerate(HEADERS):
        if header == "계약번호(Ref no.)":
            contract_no_col_idx = idx
        elif header == "전대_리스개시일":
            lease_start_date_col_idx = idx
        elif header == "기초":
            beginning_balance_col_idx = idx
    
    if contract_no_col_idx is None or lease_start_date_col_idx is None or beginning_balance_col_idx is None:
        print("  경고: 필요한 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산 (B열부터 시작)
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열 = 2
    lease_start_date_col_letter = get_column_letter(2 + lease_start_date_col_idx)  # I열
    beginning_balance_col_letter = get_column_letter(2 + beginning_balance_col_idx)  # K열
    
    # 전대_Lease_Data 시트의 열 문자
    # 계약번호(Ref no.): B열
    # 유효일자: I열 (전대_Lease_Data 시트 헤더 기준)
    # 리스채권: W열 (전대_Lease_Data 시트 헤더 기준)
    lease_data_contract_no_col = "B"  # 전대_Lease_Data 시트의 계약번호 열
    lease_data_effective_date_col = "I"  # 전대_Lease_Data 시트의 유효일자 열
    lease_data_receivable_col = "W"  # 전대_Lease_Data 시트의 리스채권 열
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =IF(I9>=Input_data!$C$6,0,SUMIFS(전대_Lease_Data!$Z:$Z,전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$I:$I,EOMONTH(Input_data!$C$6,-1)))
        
        lease_start_date_cell = f"{lease_start_date_col_letter}{row_idx}"
        contract_no_cell = f"{contract_no_col_letter}{row_idx}"
        beginning_balance_cell = f"{beginning_balance_col_letter}{row_idx}"
        
        formula = f"=IF({lease_start_date_cell}>=Input_data!$C$6,0,SUMIFS(전대_Lease_Data!${lease_data_receivable_col}:${lease_data_receivable_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_effective_date_col}:${lease_data_effective_date_col},EOMONTH(Input_data!$C$6,-1)))"
        
        add_formula(lease_receivable_sheet, beginning_balance_cell, formula)
    
    print(f"  ✓ 기초 열 입력 완료: {num_rows}행")
    
    # 취득 열 채우기
    fill_acquisition_column(lease_receivable_sheet, num_rows, start_row)


def fill_acquisition_column(lease_receivable_sheet, num_rows, start_row=8):
    """
    리스채권 시트의 취득 열 채우기
    
    수식: =IF(I9>=Input_data!$C$6,SUMIFS(전대_Lease_Data!$Z:$Z,전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$N:$N, MIN(IF(전대_Lease_Data!$B:$B=B9, 전대_Lease_Data!$N:$N))),0)
    
    Args:
        lease_receivable_sheet: 리스채권 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8)
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[취득 열 채우기]")
    
    # 리스채권 시트의 헤더 인덱스 찾기
    contract_no_col_idx = None
    lease_start_date_col_idx = None
    acquisition_col_idx = None
    
    for idx, header in enumerate(HEADERS):
        if header == "계약번호(Ref no.)":
            contract_no_col_idx = idx
        elif header == "전대_리스개시일":
            lease_start_date_col_idx = idx
        elif header == "취득":
            acquisition_col_idx = idx
    
    if contract_no_col_idx is None or lease_start_date_col_idx is None or acquisition_col_idx is None:
        print("  경고: 필요한 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산 (B열부터 시작)
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열 = 2
    lease_start_date_col_letter = get_column_letter(2 + lease_start_date_col_idx)  # I열
    acquisition_col_letter = get_column_letter(2 + acquisition_col_idx)  # L열
    
    # 전대_Lease_Data 시트의 열 문자
    # 계약번호(Ref no.): B열
    # 수령회차: N열 (전대_Lease_Data 시트 헤더 기준)
    # 리스채권: W열 (취득 열에서는 리스채권 열 사용, 리스채권(환산후) 아님)
    lease_data_contract_no_col = "B"  # 전대_Lease_Data 시트의 계약번호 열
    lease_data_receipt_count_col = "N"  # 전대_Lease_Data 시트의 수령회차 열
    lease_data_receivable_col = "W"  # 전대_Lease_Data 시트의 리스채권 열 (취득 열용)
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =IF(I9>=Input_data!$C$6,SUMIFS(전대_Lease_Data!$W:$W,전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$N:$N, MIN(IF(전대_Lease_Data!$B:$B=B9, 전대_Lease_Data!$N:$N))),0)
        
        lease_start_date_cell = f"{lease_start_date_col_letter}{row_idx}"
        contract_no_cell = f"{contract_no_col_letter}{row_idx}"
        acquisition_cell = f"{acquisition_col_letter}{row_idx}"
        
        # 배열 수식으로 MIN(IF(...)) 사용
        formula = f"=IF({lease_start_date_cell}>=Input_data!$C$6,SUMIFS(전대_Lease_Data!${lease_data_receivable_col}:${lease_data_receivable_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_receipt_count_col}:${lease_data_receipt_count_col},MIN(IF(전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col}={contract_no_cell},전대_Lease_Data!${lease_data_receipt_count_col}:${lease_data_receipt_count_col}))),0)"
        
        add_formula(lease_receivable_sheet, acquisition_cell, formula)
    
    print(f"  ✓ 취득 열 입력 완료: {num_rows}행")
    
    # 원금상환 열 채우기
    fill_principal_repayment_column(lease_receivable_sheet, num_rows, start_row)


def fill_principal_repayment_column(lease_receivable_sheet, num_rows, start_row=8):
    """
    리스채권 시트의 원금상환 열 채우기
    
    수식: =-SUMIFS(
      전대_Lease_Data!$T:$T,
      전대_Lease_Data!$B:$B, B9,
      전대_Lease_Data!$G:$G, ">="&Input_data!$C$6,
      전대_Lease_Data!$G:$G, "<="&IF(K9="", $C$4, MIN($C$4, EOMONTH(K9,0)))
    )
    
    Args:
        lease_receivable_sheet: 리스채권 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8)
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[원금상환 열 채우기]")
    
    # 리스채권 시트의 헤더 인덱스 찾기
    contract_no_col_idx = None
    early_termination_date_col_idx = None
    principal_repayment_col_idx = None
    
    for idx, header in enumerate(HEADERS):
        if header == "계약번호(Ref no.)":
            contract_no_col_idx = idx
        elif header == "전대_중도해지일":
            early_termination_date_col_idx = idx
        elif header == "원금상환":
            principal_repayment_col_idx = idx
    
    if contract_no_col_idx is None or early_termination_date_col_idx is None or principal_repayment_col_idx is None:
        print("  경고: 필요한 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산 (B열부터 시작)
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열 = 2
    early_termination_date_col_letter = get_column_letter(2 + early_termination_date_col_idx)  # K열
    principal_repayment_col_letter = get_column_letter(2 + principal_repayment_col_idx)  # N열
    
    # 전대_Lease_Data 시트의 열 문자
    # 계약번호(Ref no.): B열
    # 지급일자: G열
    # 리스료: S열 (원금상환 열에서는 리스료 열 사용, 리스료(환산후) 아님)
    lease_data_contract_no_col = "B"  # 전대_Lease_Data 시트의 계약번호 열
    lease_data_payment_date_col = "G"  # 전대_Lease_Data 시트의 지급일자 열
    lease_data_lease_fee_col = "S"  # 전대_Lease_Data 시트의 리스료 열
    
    # 리스채권 시트의 기준일 셀 (C4)
    base_date_cell = "C4"
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =-SUMIFS(
        #   전대_Lease_Data!$S:$S,
        #   전대_Lease_Data!$B:$B, B9,
        #   전대_Lease_Data!$G:$G, ">="&Input_data!$C$6,
        #   전대_Lease_Data!$G:$G, "<="&IF(K9="", $C$4, MIN($C$4, EOMONTH(K9,0)))
        # )
        
        contract_no_cell = f"{contract_no_col_letter}{row_idx}"
        early_termination_date_cell = f"{early_termination_date_col_letter}{row_idx}"
        principal_repayment_cell = f"{principal_repayment_col_letter}{row_idx}"
        
        formula = f"=-SUMIFS(전대_Lease_Data!${lease_data_lease_fee_col}:${lease_data_lease_fee_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_payment_date_col}:${lease_data_payment_date_col},\">=\"&Input_data!$C$6,전대_Lease_Data!${lease_data_payment_date_col}:${lease_data_payment_date_col},\"<=\"&IF({early_termination_date_cell}=\"\",{base_date_cell},MIN({base_date_cell},EOMONTH({early_termination_date_cell},0))))"
        
        add_formula(lease_receivable_sheet, principal_repayment_cell, formula)
    
    print(f"  ✓ 원금상환 열 입력 완료: {num_rows}행")
    
    # 이자수익 열 채우기
    fill_interest_income_column(lease_receivable_sheet, num_rows, start_row)


def fill_interest_income_column(lease_receivable_sheet, num_rows, start_row=8):
    """
    리스채권 시트의 이자수익 열 채우기
    
    수식: =SUMIFS(전대_Lease_Data!$X:$X,전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$I:$I,">="&Input_data!$C$6,전대_Lease_Data!$I:$I,"<="&IF(K9=0,$C$4,MIN(K9,$C$4)))
    
    Args:
        lease_receivable_sheet: 리스채권 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8)
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[이자수익 열 채우기]")
    
    # 리스채권 시트의 헤더 인덱스 찾기
    contract_no_col_idx = None
    early_termination_date_col_idx = None
    interest_income_col_idx = None
    
    for idx, header in enumerate(HEADERS):
        if header == "계약번호(Ref no.)":
            contract_no_col_idx = idx
        elif header == "전대_중도해지일":
            early_termination_date_col_idx = idx
        elif header == "이자수익":
            interest_income_col_idx = idx
    
    if contract_no_col_idx is None or early_termination_date_col_idx is None or interest_income_col_idx is None:
        print("  경고: 필요한 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산 (B열부터 시작)
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열 = 2
    early_termination_date_col_letter = get_column_letter(2 + early_termination_date_col_idx)  # K열
    interest_income_col_letter = get_column_letter(2 + interest_income_col_idx)  # O열
    
    # 전대_Lease_Data 시트의 열 문자
    # 계약번호(Ref no.): B열
    # 유효일자: I열
    # 이자수익: X열
    lease_data_contract_no_col = "B"  # 전대_Lease_Data 시트의 계약번호 열
    lease_data_effective_date_col = "I"  # 전대_Lease_Data 시트의 유효일자 열
    lease_data_interest_income_col = "X"  # 전대_Lease_Data 시트의 이자수익 열
    
    # 리스채권 시트의 기준일 셀 (C4)
    base_date_cell = "C4"
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =SUMIFS(전대_Lease_Data!$X:$X,전대_Lease_Data!$B:$B,B9,전대_Lease_Data!$I:$I,">="&Input_data!$C$6,전대_Lease_Data!$I:$I,"<="&IF(K9=0,$C$4,MIN(K9,$C$4)))
        
        contract_no_cell = f"{contract_no_col_letter}{row_idx}"
        early_termination_date_cell = f"{early_termination_date_col_letter}{row_idx}"
        interest_income_cell = f"{interest_income_col_letter}{row_idx}"
        
        formula = f"=SUMIFS(전대_Lease_Data!${lease_data_interest_income_col}:${lease_data_interest_income_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_effective_date_col}:${lease_data_effective_date_col},\">=\"&Input_data!$C$6,전대_Lease_Data!${lease_data_effective_date_col}:${lease_data_effective_date_col},\"<=\"&IF({early_termination_date_cell}=0,{base_date_cell},MIN({early_termination_date_cell},{base_date_cell})))"
        
        add_formula(lease_receivable_sheet, interest_income_cell, formula)
    
    print(f"  ✓ 이자수익 열 입력 완료: {num_rows}행")
    
    # 중도해지 열 채우기
    fill_early_termination_column(lease_receivable_sheet, num_rows, start_row)


def fill_early_termination_column(lease_receivable_sheet, num_rows, start_row=8):
    """
    리스채권 시트의 중도해지 열 채우기
    
    수식: =-SUMIFS(전대_Lease_Data!Z:Z,전대_Lease_Data!B:B,B9,전대_Lease_Data!AP:AP,"O")
    
    Args:
        lease_receivable_sheet: 리스채권 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8)
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[중도해지 열 채우기]")
    
    # 리스채권 시트의 헤더 인덱스 찾기
    contract_no_col_idx = None
    early_termination_col_idx = None
    
    for idx, header in enumerate(HEADERS):
        if header == "계약번호(Ref no.)":
            contract_no_col_idx = idx
        elif header == "중도해지":
            early_termination_col_idx = idx
    
    if contract_no_col_idx is None or early_termination_col_idx is None:
        print("  경고: 필요한 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산 (B열부터 시작)
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열 = 2
    early_termination_col_letter = get_column_letter(2 + early_termination_col_idx)  # Q열
    
    # 전대_Lease_Data 시트의 열 문자
    # 계약번호(Ref no.): B열
    # 리스채권: W열 (리스채권 열, 리스채권(환산후) 아님)
    # 중도해지: AP열
    lease_data_contract_no_col = "B"  # 전대_Lease_Data 시트의 계약번호 열
    lease_data_receivable_col = "W"  # 전대_Lease_Data 시트의 리스채권 열
    lease_data_early_termination_col = "AP"  # 전대_Lease_Data 시트의 중도해지 열
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =-SUMIFS(전대_Lease_Data!Z:Z,전대_Lease_Data!B:B,B9,전대_Lease_Data!AP:AP,"O")
        
        contract_no_cell = f"{contract_no_col_letter}{row_idx}"
        early_termination_cell = f"{early_termination_col_letter}{row_idx}"
        
        formula = f"=-SUMIFS(전대_Lease_Data!${lease_data_receivable_col}:${lease_data_receivable_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_early_termination_col}:${lease_data_early_termination_col},\"O\")"
        
        add_formula(lease_receivable_sheet, early_termination_cell, formula)
    
    print(f"  ✓ 중도해지 열 입력 완료: {num_rows}행")
    
    # 기말 열 채우기
    fill_ending_balance_column(lease_receivable_sheet, num_rows, start_row)


def fill_ending_balance_column(lease_receivable_sheet, num_rows, start_row=8):
    """
    리스채권 시트의 기말 열 채우기
    
    수식: =SUM(L9:Q9)
    L9: 동일행의 기초 열
    Q9: 동일행의 중도해지 열
    
    Args:
        lease_receivable_sheet: 리스채권 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8)
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[기말 열 채우기]")
    
    # 리스채권 시트의 헤더 인덱스 찾기
    beginning_balance_col_idx = None
    early_termination_col_idx = None
    ending_balance_col_idx = None
    
    for idx, header in enumerate(HEADERS):
        if header == "기초":
            beginning_balance_col_idx = idx
        elif header == "중도해지":
            early_termination_col_idx = idx
        elif header == "기말":
            ending_balance_col_idx = idx
    
    if beginning_balance_col_idx is None or early_termination_col_idx is None or ending_balance_col_idx is None:
        print("  경고: 필요한 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산 (B열부터 시작)
    beginning_balance_col_letter = get_column_letter(2 + beginning_balance_col_idx)  # K열
    early_termination_col_letter = get_column_letter(2 + early_termination_col_idx)  # P열
    ending_balance_col_letter = get_column_letter(2 + ending_balance_col_idx)  # Q열
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =SUM(L9:Q9) - 사용자가 L9:Q9라고 했지만, 실제로는 기초 열(K열)부터 중도해지 열(P열)까지의 합계
        # 사용자 요구사항에 따라 L9(기초 열)부터 Q9(중도해지 열)까지로 구현
        # 하지만 실제로는 기초 열이 K열이므로, K열부터 P열까지의 합계를 계산
        
        beginning_balance_cell = f"{beginning_balance_col_letter}{row_idx}"
        early_termination_cell = f"{early_termination_col_letter}{row_idx}"
        ending_balance_cell = f"{ending_balance_col_letter}{row_idx}"
        
        # 기초 열부터 중도해지 열까지의 합계
        formula = f"=SUM({beginning_balance_cell}:{early_termination_cell})"
        
        add_formula(lease_receivable_sheet, ending_balance_cell, formula)
    
    print(f"  ✓ 기말 열 입력 완료: {num_rows}행")
    
    # 유동 열 채우기
    fill_current_column(lease_receivable_sheet, num_rows, start_row)
    
    # 비유동 열 채우기
    fill_non_current_column(lease_receivable_sheet, num_rows, start_row)


def fill_current_column(lease_receivable_sheet, num_rows, start_row=8):
    """
    리스채권 시트의 유동 열 채우기
    
    수식: =IF(R9=0,0,SUMIFS(전대_Lease_Data!AQ:AQ,전대_Lease_Data!I:I,EOMONTH($C$4,0)))
    
    Args:
        lease_receivable_sheet: 리스채권 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8)
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[유동 열 채우기]")
    
    # 리스채권 시트의 헤더 인덱스 찾기
    ending_balance_col_idx = None
    current_col_idx = None
    
    for idx, header in enumerate(HEADERS):
        if header == "기말":
            ending_balance_col_idx = idx
        elif header == "유동":
            current_col_idx = idx
    
    if ending_balance_col_idx is None or current_col_idx is None:
        print("  경고: 필요한 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산 (B열부터 시작)
    ending_balance_col_letter = get_column_letter(2 + ending_balance_col_idx)  # R열
    current_col_letter = get_column_letter(2 + current_col_idx)  # S열
    
    # 전대_Lease_Data 시트의 열 문자
    # 유효일자: I열
    # 리스채권_유동: AQ열
    lease_data_effective_date_col = "I"  # 전대_Lease_Data 시트의 유효일자 열
    lease_data_current_col = "AQ"  # 전대_Lease_Data 시트의 리스채권_유동 열
    
    # 리스채권 시트의 기준일 셀 (C4)
    base_date_cell = "C4"
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =IF(R9=0,0,SUMIFS(전대_Lease_Data!AQ:AQ,전대_Lease_Data!I:I,EOMONTH($C$4,0)))
        
        ending_balance_cell = f"{ending_balance_col_letter}{row_idx}"
        current_cell = f"{current_col_letter}{row_idx}"
        
        formula = f"=IF({ending_balance_cell}=0,0,SUMIFS(전대_Lease_Data!${lease_data_current_col}:${lease_data_current_col},전대_Lease_Data!${lease_data_effective_date_col}:${lease_data_effective_date_col},EOMONTH({base_date_cell},0)))"
        
        add_formula(lease_receivable_sheet, current_cell, formula)
    
    print(f"  ✓ 유동 열 입력 완료: {num_rows}행")


def fill_non_current_column(lease_receivable_sheet, num_rows, start_row=8):
    """
    리스채권 시트의 비유동 열 채우기
    
    수식: =R9-S9
    R9: 동일행의 기말 열
    S9: 동일행의 유동 열
    
    Args:
        lease_receivable_sheet: 리스채권 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8)
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[비유동 열 채우기]")
    
    # 리스채권 시트의 헤더 인덱스 찾기
    ending_balance_col_idx = None
    current_col_idx = None
    non_current_col_idx = None
    
    for idx, header in enumerate(HEADERS):
        if header == "기말":
            ending_balance_col_idx = idx
        elif header == "유동":
            current_col_idx = idx
        elif header == "비유동":
            non_current_col_idx = idx
    
    if ending_balance_col_idx is None or current_col_idx is None or non_current_col_idx is None:
        print("  경고: 필요한 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산 (B열부터 시작)
    ending_balance_col_letter = get_column_letter(2 + ending_balance_col_idx)  # R열
    current_col_letter = get_column_letter(2 + current_col_idx)  # S열
    non_current_col_letter = get_column_letter(2 + non_current_col_idx)  # T열
    
    # 각 행에 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        # 수식 구성
        # =R9-S9
        
        ending_balance_cell = f"{ending_balance_col_letter}{row_idx}"
        current_cell = f"{current_col_letter}{row_idx}"
        non_current_cell = f"{non_current_col_letter}{row_idx}"
        
        formula = f"={ending_balance_cell}-{current_cell}"
        
        add_formula(lease_receivable_sheet, non_current_cell, formula)
    
    print(f"  ✓ 비유동 열 입력 완료: {num_rows}행")
    
    # 1월~12월 열 채우기 (이자수익 그룹)
    fill_monthly_interest_income_columns(lease_receivable_sheet, num_rows, start_row)


def fill_monthly_interest_income_columns(lease_receivable_sheet, num_rows, start_row=8):
    """
    리스채권 시트의 1월~12월 열 채우기 (이자수익 그룹)
    
    수식: =SUMIFS(전대_Lease_Data!$X:$X,전대_Lease_Data!$B:$B,$B9,전대_Lease_Data!$I:$I,"<="&EOMONTH(MIN($J9,$K9),0),전대_Lease_Data!$J:$J,YEAR($C$4),전대_Lease_Data!$K:$K,SUBSTITUTE(U7,"월",""))
    
    Args:
        lease_receivable_sheet: 리스채권 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8)
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[1월~12월 열 채우기 (이자수익 그룹)]")
    
    # 리스채권 시트의 헤더 인덱스 찾기
    contract_no_col_idx = None
    lease_end_date_col_idx = None
    early_termination_date_col_idx = None
    
    for idx, header in enumerate(HEADERS):
        if header == "계약번호(Ref no.)":
            contract_no_col_idx = idx
        elif header == "전대_리스종료일":
            lease_end_date_col_idx = idx
        elif header == "전대_중도해지일":
            early_termination_date_col_idx = idx
    
    if contract_no_col_idx is None or lease_end_date_col_idx is None or early_termination_date_col_idx is None:
        print("  경고: 필요한 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산 (B열부터 시작)
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열 = 2
    lease_end_date_col_letter = get_column_letter(2 + lease_end_date_col_idx)  # J열
    early_termination_date_col_letter = get_column_letter(2 + early_termination_date_col_idx)  # K열
    
    # 전대_Lease_Data 시트의 열 문자
    # 계약번호(Ref no.): B열
    # 유효일자: I열
    # 회계연도: J열
    # 결산월: K열
    # 이자수익: X열
    lease_data_contract_no_col = "B"  # 전대_Lease_Data 시트의 계약번호 열
    lease_data_effective_date_col = "I"  # 전대_Lease_Data 시트의 유효일자 열
    lease_data_fiscal_year_col = "J"  # 전대_Lease_Data 시트의 회계연도 열
    lease_data_fiscal_month_col = "K"  # 전대_Lease_Data 시트의 결산월 열
    lease_data_interest_income_col = "X"  # 전대_Lease_Data 시트의 이자수익 열
    
    # 리스채권 시트의 기준일 셀 (C4)
    base_date_cell = "C4"
    
    # 헤더 행 (7행)
    header_row = 7
    
    # 1월~12월 열 찾기 (첫 번째 세트, 이자수익 그룹)
    monthly_headers = ["1월", "2월", "3월", "4월", "5월", "6월", "7월", "8월", "9월", "10월", "11월", "12월"]
    monthly_col_indices = []
    
    for month_header in monthly_headers:
        for idx, header in enumerate(HEADERS):
            if header == month_header:
                monthly_col_indices.append(idx)
                break
    
    if len(monthly_col_indices) != 12:
        print("  경고: 1월~12월 헤더를 모두 찾을 수 없습니다.")
        return
    
    # 각 월별 열에 수식 입력
    for month_idx, col_idx in enumerate(monthly_col_indices):
        month_col_letter = get_column_letter(2 + col_idx)  # U열부터 시작
        header_cell = f"{month_col_letter}{header_row}"  # 헤더 셀 (예: U7)
        
        for row_idx in range(start_row, start_row + num_rows):
            # 수식 구성
            # =SUMIFS(전대_Lease_Data!$X:$X,전대_Lease_Data!$B:$B,$B9,전대_Lease_Data!$I:$I,"<="&EOMONTH(MIN($J9,$K9),0),전대_Lease_Data!$J:$J,YEAR($C$4),전대_Lease_Data!$K:$K,SUBSTITUTE(U7,"월",""))
            
            contract_no_cell = f"${contract_no_col_letter}${row_idx}"  # $B9
            lease_end_date_cell = f"${lease_end_date_col_letter}${row_idx}"  # $J9
            early_termination_date_cell = f"${early_termination_date_col_letter}${row_idx}"  # $K9
            month_cell = f"{month_col_letter}{row_idx}"
            
            formula = f"=SUMIFS(전대_Lease_Data!${lease_data_interest_income_col}:${lease_data_interest_income_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_effective_date_col}:${lease_data_effective_date_col},\"<=\"&EOMONTH(MIN({lease_end_date_cell},{early_termination_date_cell}),0),전대_Lease_Data!${lease_data_fiscal_year_col}:${lease_data_fiscal_year_col},YEAR({base_date_cell}),전대_Lease_Data!${lease_data_fiscal_month_col}:${lease_data_fiscal_month_col},SUBSTITUTE({header_cell},\"월\",\"\"))"
            
            add_formula(lease_receivable_sheet, month_cell, formula)
    
    print(f"  ✓ 1월~12월 열 입력 완료 (이자수익 그룹): {num_rows}행")
    
    # 1월~12월 열 채우기 (리스료 그룹)
    fill_monthly_lease_fee_columns(lease_receivable_sheet, num_rows, start_row)


def fill_monthly_lease_fee_columns(lease_receivable_sheet, num_rows, start_row=8):
    """
    리스채권 시트의 1월~12월 열 채우기 (리스료 그룹)
    
    수식: =SUMIFS(전대_Lease_Data!$T:$T,전대_Lease_Data!$B:$B,$B9,전대_Lease_Data!$I:$I,"<="&EOMONTH(MIN($J9,$K9),0),전대_Lease_Data!$J:$J,YEAR($C$4),전대_Lease_Data!$K:$K,SUBSTITUTE(AG7,"월",""))
    
    Args:
        lease_receivable_sheet: 리스채권 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8)
    """
    from excel_utils import add_formula
    from openpyxl.utils import get_column_letter
    
    print("\n[1월~12월 열 채우기 (리스료 그룹)]")
    
    # 리스채권 시트의 헤더 인덱스 찾기
    contract_no_col_idx = None
    lease_end_date_col_idx = None
    early_termination_date_col_idx = None
    
    for idx, header in enumerate(HEADERS):
        if header == "계약번호(Ref no.)":
            contract_no_col_idx = idx
        elif header == "전대_리스종료일":
            lease_end_date_col_idx = idx
        elif header == "전대_중도해지일":
            early_termination_date_col_idx = idx
    
    if contract_no_col_idx is None or lease_end_date_col_idx is None or early_termination_date_col_idx is None:
        print("  경고: 필요한 헤더를 찾을 수 없습니다.")
        return
    
    # 열 문자 계산 (B열부터 시작)
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx)  # B열 = 2
    lease_end_date_col_letter = get_column_letter(2 + lease_end_date_col_idx)  # J열
    early_termination_date_col_letter = get_column_letter(2 + early_termination_date_col_idx)  # K열
    
    # 전대_Lease_Data 시트의 열 문자
    # 계약번호(Ref no.): B열
    # 유효일자: I열
    # 회계연도: J열
    # 결산월: K열
    # 리스료: S열 (리스료 열, 리스료(환산후) 아님)
    lease_data_contract_no_col = "B"  # 전대_Lease_Data 시트의 계약번호 열
    lease_data_effective_date_col = "I"  # 전대_Lease_Data 시트의 유효일자 열
    lease_data_fiscal_year_col = "J"  # 전대_Lease_Data 시트의 회계연도 열
    lease_data_fiscal_month_col = "K"  # 전대_Lease_Data 시트의 결산월 열
    lease_data_lease_fee_col = "S"  # 전대_Lease_Data 시트의 리스료 열 (리스료(환산후) 아님)
    
    # 리스채권 시트의 기준일 셀 (C4)
    base_date_cell = "C4"
    
    # 헤더 행 (7행)
    header_row = 7
    
    # 1월~12월 열 찾기 (두 번째 세트, 리스료 그룹)
    # HEADERS에서 첫 번째 "1월"을 찾고, 그 다음에 나오는 두 번째 "1월"부터 12개를 가져옴
    monthly_headers = ["1월", "2월", "3월", "4월", "5월", "6월", "7월", "8월", "9월", "10월", "11월", "12월"]
    monthly_col_indices = []
    
    # 첫 번째 "1월"의 인덱스 찾기
    first_jan_idx = None
    for idx, header in enumerate(HEADERS):
        if header == "1월":
            first_jan_idx = idx
            break
    
    if first_jan_idx is None:
        print("  경고: 첫 번째 '1월' 헤더를 찾을 수 없습니다.")
        return
    
    # 두 번째 "1월"의 인덱스 찾기 (첫 번째 이후)
    second_jan_idx = None
    for idx in range(first_jan_idx + 1, len(HEADERS)):
        if HEADERS[idx] == "1월":
            second_jan_idx = idx
            break
    
    if second_jan_idx is None:
        print("  경고: 두 번째 '1월' 헤더를 찾을 수 없습니다.")
        return
    
    # 두 번째 세트의 1월~12월 인덱스 (12개)
    monthly_col_indices = list(range(second_jan_idx, second_jan_idx + 12))
    
    if len(monthly_col_indices) != 12:
        print("  경고: 리스료 그룹의 1월~12월 헤더를 모두 찾을 수 없습니다.")
        return
    
    # 각 월별 열에 수식 입력
    for month_idx, col_idx in enumerate(monthly_col_indices):
        month_col_letter = get_column_letter(2 + col_idx)  # AG열부터 시작
        header_cell = f"{month_col_letter}{header_row}"  # 헤더 셀 (예: AG7)
        
        for row_idx in range(start_row, start_row + num_rows):
            # 수식 구성
            # =SUMIFS(전대_Lease_Data!$T:$T,전대_Lease_Data!$B:$B,$B9,전대_Lease_Data!$I:$I,"<="&EOMONTH(MIN($J9,$K9),0),전대_Lease_Data!$J:$J,YEAR($C$4),전대_Lease_Data!$K:$K,SUBSTITUTE(AG7,"월",""))
            
            contract_no_cell = f"${contract_no_col_letter}${row_idx}"  # $B9
            lease_end_date_cell = f"${lease_end_date_col_letter}${row_idx}"  # $J9
            early_termination_date_cell = f"${early_termination_date_col_letter}${row_idx}"  # $K9
            month_cell = f"{month_col_letter}{row_idx}"
            
            formula = f"=SUMIFS(전대_Lease_Data!${lease_data_lease_fee_col}:${lease_data_lease_fee_col},전대_Lease_Data!${lease_data_contract_no_col}:${lease_data_contract_no_col},{contract_no_cell},전대_Lease_Data!${lease_data_effective_date_col}:${lease_data_effective_date_col},\"<=\"&EOMONTH(MIN({lease_end_date_cell},{early_termination_date_cell}),0),전대_Lease_Data!${lease_data_fiscal_year_col}:${lease_data_fiscal_year_col},YEAR({base_date_cell}),전대_Lease_Data!${lease_data_fiscal_month_col}:${lease_data_fiscal_month_col},SUBSTITUTE({header_cell},\"월\",\"\"))"
            
            add_formula(lease_receivable_sheet, month_cell, formula)
    
    print(f"  ✓ 1월~12월 열 입력 완료 (리스료 그룹): {num_rows}행")
    
    # 회계 형식 적용 (기초 열부터 마지막 12월 열까지)
    apply_accounting_format(lease_receivable_sheet, num_rows, start_row)
    
    # 테이블 범위에 테두리 적용
    apply_border_to_lease_receivable_table(lease_receivable_sheet, num_rows, start_row)


def apply_accounting_format(lease_receivable_sheet, num_rows, start_row=8):
    """
    리스채권 시트의 기초 열부터 마지막 12월 열까지 회계 형식 적용
    
    Args:
        lease_receivable_sheet: 리스채권 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8)
    """
    from openpyxl.utils import get_column_letter
    
    print("\n[회계 형식 적용]")
    
    # 기초 열의 인덱스 찾기
    beginning_balance_col_idx = None
    for idx, header in enumerate(HEADERS):
        if header == "기초":
            beginning_balance_col_idx = idx
            break
    
    if beginning_balance_col_idx is None:
        print("  경고: '기초' 헤더를 찾을 수 없습니다.")
        return
    
    # 마지막 12월 열의 인덱스 찾기 (두 번째 세트의 12월)
    last_dec_col_idx = None
    first_dec_found = False
    for idx, header in enumerate(HEADERS):
        if header == "12월":
            if not first_dec_found:
                first_dec_found = True
            else:
                last_dec_col_idx = idx
                break
    
    if last_dec_col_idx is None:
        print("  경고: 마지막 '12월' 헤더를 찾을 수 없습니다.")
        return
    
    # 회계 형식 (#,##0)
    accounting_format = "#,##0"
    
    # 기초 열부터 마지막 12월 열까지 회계 형식 적용
    for col_idx in range(beginning_balance_col_idx, last_dec_col_idx + 1):
        col_letter = get_column_letter(2 + col_idx)  # B열부터 시작
        for row_idx in range(start_row, start_row + num_rows):
            cell = lease_receivable_sheet[f"{col_letter}{row_idx}"]
            cell.number_format = accounting_format
    
    num_cols = last_dec_col_idx - beginning_balance_col_idx + 1
    print(f"  ✓ 회계 형식 적용 완료: {num_cols}개 열 ({num_rows}행)")


def apply_border_to_lease_receivable_table(lease_receivable_sheet, num_rows, start_row=8):
    """
    리스채권 시트의 테이블 범위에 모든 테두리 적용
    
    Args:
        lease_receivable_sheet: 리스채권 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 8, 헤더는 7행)
    """
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter
    
    # 테두리 스타일 정의
    thin_border = Side(style='thin')
    border = Border(
        left=thin_border,
        right=thin_border,
        top=thin_border,
        bottom=thin_border
    )
    
    # 셀병합된 행 (6행)부터 마지막 데이터 행까지
    merged_row = 6  # 셀병합된 행
    header_row = 7  # 헤더 행
    num_cols = len(HEADERS)  # 헤더 개수
    start_col = 2  # B열부터 시작
    
    # 셀병합된 행(6행)에 테두리 적용
    for col_idx in range(start_col, start_col + num_cols):
        col_letter = get_column_letter(col_idx)
        cell = lease_receivable_sheet[f"{col_letter}{merged_row}"]
        cell.border = border
    
    # 헤더 행(7행)부터 마지막 데이터 행까지 테두리 적용
    for row_idx in range(header_row, start_row + num_rows):
        for col_idx in range(start_col, start_col + num_cols):
            col_letter = get_column_letter(col_idx)
            cell = lease_receivable_sheet[f"{col_letter}{row_idx}"]
            cell.border = border
    
    print(f"  ✓ 테두리 적용 완료: {num_cols}개 열, {num_rows + 2}행 (셀병합 행 + 헤더 포함)")

