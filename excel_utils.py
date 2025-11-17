"""
엑셀 작업 유틸리티 모듈
- 시트 추가/관리
- 테이블 생성
- 수식 및 셀 작업
"""

from openpyxl.worksheet.table import Table, TableStyleInfo


def add_sheet(wb, sheet_name, position=None):
    """
    새 시트 추가
    
    Args:
        wb: openpyxl Workbook 객체
        sheet_name (str): 시트 이름
        position (int): 시트 위치 (None이면 마지막에 추가)
        
    Returns:
        Worksheet: 생성된 시트 객체
    """
    if sheet_name in wb.sheetnames:
        print(f"경고: '{sheet_name}' 시트가 이미 존재합니다.")
        return wb[sheet_name]
    
    ws = wb.create_sheet(sheet_name, position)
    print(f"시트 추가 완료: {sheet_name}")
    return ws


def create_table(ws, start_cell, end_cell, table_name, has_header=True, style_name="TableStyleMedium9"):
    """
    테이블 생성 (openpyxl의 Table 객체 생성)
    
    Args:
        ws: Worksheet 객체
        start_cell (str): 시작 셀 (예: 'A1')
        end_cell (str): 끝 셀 (예: 'D10')
        table_name (str): 테이블 이름
        has_header (bool): 헤더 행 포함 여부
        style_name (str): 테이블 스타일 이름
        
    Returns:
        Table: 생성된 테이블 객체
    """
    table = Table(displayName=table_name, ref=f"{start_cell}:{end_cell}")
    
    # 스타일 적용
    style = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    
    ws.add_table(table)
    print(f"테이블 생성 완료: {table_name} ({start_cell}:{end_cell})")
    return table


def remove_at_symbol_from_all_formulas(wb):
    """
    워크북의 모든 시트에서 모든 수식의 @ 기호 제거
    
    Args:
        wb: openpyxl Workbook 객체
    """
    import re
    
    removed_count = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    if '@' in cell.value:
                        # @ 기호 제거 - 모든 @를 완전히 제거
                        original = cell.value
                        # @ 뒤에 오는 모든 문자를 포함하여 제거
                        # @전대_Lease_Data!$B:$B -> 전대_Lease_Data!$B:$B
                        # @$ -> $
                        cleaned = re.sub(r'@([A-Za-z_가-힣][A-Za-z0-9_가-힣!$:]*|\$)', r'\1', cell.value)
                        # 나머지 모든 @ 제거 (혹시 모를 경우를 대비)
                        cleaned = cleaned.replace('@', '')
                        
                        if cleaned != original:
                            cell.value = cleaned
                            removed_count += 1
                            # 디버깅: @가 제거된 수식 출력
                            if removed_count <= 5:  # 처음 5개만 출력
                                print(f"  [@ 제거] {sheet_name}!{cell.coordinate}: {original[:50]}... -> {cleaned[:50]}...")
    
    if removed_count > 0:
        print(f"  ✓ {removed_count}개의 수식에서 @ 기호 제거 완료")
    else:
        print("  ✓ @ 기호가 포함된 수식이 없습니다.")


def add_formula(ws, cell, formula):
    """
    셀에 수식 추가
    
    Args:
        ws: Worksheet 객체
        cell (str): 셀 주소 (예: 'A1')
        formula (str): 수식 (예: '=SUM(A1:A10)')
    """
    ws[cell] = formula
    print(f"수식 추가 완료: {cell} = {formula}")


def set_cell_value(ws, cell, value):
    """
    셀에 값 설정
    
    Args:
        ws: Worksheet 객체
        cell (str): 셀 주소
        value: 설정할 값
    """
    ws[cell] = value


def get_cell_value(ws, cell):
    """
    셀 값 가져오기
    
    Args:
        ws: Worksheet 객체
        cell (str): 셀 주소
        
    Returns:
        셀의 값
    """
    return ws[cell].value


def set_range_values(ws, start_cell, values):
    """
    범위에 여러 값 설정
    
    Args:
        ws: Worksheet 객체
        start_cell (str): 시작 셀 주소 (예: 'A1')
        values (list): 설정할 값들의 리스트 (2차원 리스트 가능)
    """
    from openpyxl.utils import coordinate_from_string, column_index_from_string
    
    col_str, row = coordinate_from_string(start_cell)
    start_col = column_index_from_string(col_str)
    
    if not isinstance(values[0], (list, tuple)):
        values = [values]  # 1차원 리스트를 2차원으로 변환
    
    for i, row_values in enumerate(values):
        for j, value in enumerate(row_values):
            col_letter = chr(ord('A') + start_col - 1 + j)
            ws[f"{col_letter}{row + i}"] = value
    
    print(f"범위 값 설정 완료: {start_cell}부터 {len(values)}행")


def get_sheet_by_name(wb, sheet_name):
    """
    시트 이름으로 시트 객체 가져오기
    
    Args:
        wb: Workbook 객체
        sheet_name (str): 시트 이름
        
    Returns:
        Worksheet: 시트 객체, 없으면 None
    """
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return None


def sheet_exists(wb, sheet_name):
    """
    시트 존재 여부 확인
    
    Args:
        wb: Workbook 객체
        sheet_name (str): 시트 이름
        
    Returns:
        bool: 시트 존재 여부
    """
    return sheet_name in wb.sheetnames


def get_table_data(ws, table_name):
    """
    테이블에서 데이터를 읽어 딕셔너리 리스트로 반환
    
    Args:
        ws: Worksheet 객체
        table_name (str): 테이블 이름
        
    Returns:
        list: 각 행이 딕셔너리인 리스트 (키는 컬럼명)
    """
    # 테이블 찾기
    table = None
    for tbl in ws.tables.values():
        if tbl.displayName == table_name:
            table = tbl
            break
    
    if table is None:
        raise ValueError(f"테이블 '{table_name}'을 찾을 수 없습니다.")
    
    # 테이블 범위 파싱
    ref = table.ref
    start_cell, end_cell = ref.split(':')
    
    # openpyxl 버전에 따라 import 경로가 다를 수 있음
    try:
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
        from openpyxl.utils import get_column_letter
    except ImportError:
        # 구버전 호환
        from openpyxl.utils import get_column_letter
        from openpyxl.cell import coordinate_from_string, column_index_from_string
    
    # 시작 셀 파싱
    start_col_str, start_row = coordinate_from_string(start_cell)
    start_col = column_index_from_string(start_col_str)
    
    # 끝 셀 파싱
    end_col_str, end_row = coordinate_from_string(end_cell)
    end_col = column_index_from_string(end_col_str)
    
    # 헤더 읽기 (첫 번째 행)
    headers = []
    for col_idx in range(start_col, end_col + 1):
        col_letter = get_column_letter(col_idx)
        cell_value = ws[f"{col_letter}{start_row}"].value
        headers.append(cell_value if cell_value is not None else f"Column{col_idx}")
    
    # 데이터 읽기
    data = []
    for row_idx in range(start_row + 1, end_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers):
            col_letter = get_column_letter(start_col + col_idx)
            cell_value = ws[f"{col_letter}{row_idx}"].value
            row_data[header] = cell_value
        data.append(row_data)
    
    return data


def find_table_by_name(wb, table_name):
    """
    워크북 전체에서 테이블 찾기
    
    Args:
        wb: Workbook 객체
        table_name (str): 테이블 이름
        
    Returns:
        tuple: (Worksheet, Table) 또는 (None, None)
    """
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for table in ws.tables.values():
            if table.displayName == table_name:
                return ws, table
    return None, None

