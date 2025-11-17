def fill_currency_column(wb, lease_data_sheet, data_manager, num_rows, start_row=7):
    """
    전대_Lease_Data 시트의 통화 열 채우기
    VLOOKUP을 사용하여 Input_data 시트의 input_table에서 통화 값 가져오기
    
    Args:
        wb: Workbook 객체
        lease_data_sheet: 전대_Lease_Data 시트
        data_manager: DataManager 객체
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 7)
    """
    from excel_utils import get_sheet_by_name, get_table_data, find_table_by_name, add_formula
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    
    print(f"\n[통화 열 채우기]")
    
    # Input_data 시트 확인
    input_data_sheet = get_sheet_by_name(wb, "Input_data")
    if input_data_sheet is None:
        print("  경고: 'Input_data' 시트를 찾을 수 없습니다.")
        return
    
    # input_table 테이블 찾기
    input_ws, table = find_table_by_name(wb, "input_table")
    if table is None:
        print("  경고: 'input_table' 테이블을 찾을 수 없습니다.")
        return
    
    # 테이블 범위 파싱
    ref = table.ref
    start_cell, end_cell = ref.split(':')
    start_col_str, start_row_num = coordinate_from_string(start_cell)
    start_col_idx = column_index_from_string(start_col_str)
    
    # 테이블 데이터 읽기
    try:
        table_data = get_table_data(input_data_sheet, "input_table")
    except Exception as e:
        print(f"  경고: input_table 데이터 읽기 실패: {e}")
        return
    
    # 계약번호(Ref no.) 열과 상위리스_계약번호(Ref no.) 열 찾기
    ref_no_col = None
    parent_ref_no_col = None
    
    headers = list(table_data[0].keys())
    for col_name in headers:
        col_str = str(col_name).strip()
        if '계약번호' in col_str and 'Ref no' in col_str:
            if '상위' in col_str or 'parent' in col_str.lower():
                parent_ref_no_col = col_name
            else:
                ref_no_col = col_name
    
    if ref_no_col is None:
        print("  경고: Input_data에서 계약번호(Ref no.) 컬럼을 찾을 수 없습니다.")
        return
    
    if parent_ref_no_col is None:
        print("  경고: Input_data에서 상위리스_계약번호(Ref no.) 컬럼을 찾을 수 없습니다.")
        return
    
    # 통화 열 찾기
    currency_col = None
    for col_name in headers:
        col_str = str(col_name).strip()
        if '통화' in col_str or 'currency' in col_str.lower():
            currency_col = col_name
            break
    
    if currency_col is None:
        print("  경고: Input_data에서 통화 컬럼을 찾을 수 없습니다.")
        return
    
    # 열 인덱스 계산
    ref_no_col_idx = headers.index(ref_no_col)
    parent_ref_no_col_idx = headers.index(parent_ref_no_col)
    currency_col_idx = headers.index(currency_col)
    
    # VLOOKUP 범위: 계약번호(Ref no.) 열부터 상위리스_계약번호(Ref no.) 열까지
    # 엑셀 구버전 호환을 위해 일반 범위 참조 사용
    range_start_col = get_column_letter(start_col_idx + ref_no_col_idx)
    range_end_col = get_column_letter(start_col_idx + parent_ref_no_col_idx)
    range_start_row = start_row_num  # 테이블 시작 행 (헤더 포함)
    range_end_row = start_row_num + len(table_data)  # 테이블 끝 행
    
    # VLOOKUP 범위 (절대참조)
    vlookup_range = f"Input_data!${range_start_col}${range_start_row}:${range_end_col}${range_end_row}"
    
    # 통화 열의 컬럼 순서 (계약번호 열부터 시작하여 몇 번째 열인지)
    col_index = currency_col_idx - ref_no_col_idx + 1
    
    # 전대_Lease_Data 시트의 통화 열 찾기
    currency_col_idx_in_lease_data = None
    for idx, header in enumerate([
        "계약번호(Ref no.)",
        "리스명",
        "상위리스_계약번호(Ref no.)",
        "리스개시일",
        "리스종료일",
        "지급일자",
        "지급연도",
        "유효일자",
        "회계연도",
        "결산월",
        "기간(월)",
        "감가상각기간(월)",
        "수령회차",
        "통화",
    ], start=1):
        if header == "통화":
            currency_col_idx_in_lease_data = idx
            break
    
    if currency_col_idx_in_lease_data is None:
        print("  경고: 전대_Lease_Data 시트에서 '통화' 헤더를 찾을 수 없습니다.")
        return
    
    # 계약번호(Ref no.) 열 찾기
    contract_no_col_idx_in_lease_data = None
    for idx, header in enumerate([
        "계약번호(Ref no.)",
        "리스명",
        "상위리스_계약번호(Ref no.)",
        "리스개시일",
        "리스종료일",
        "지급일자",
        "지급연도",
        "유효일자",
        "회계연도",
        "결산월",
        "기간(월)",
        "감가상각기간(월)",
        "수령회차",
        "통화",
    ], start=1):
        if header == "계약번호(Ref no.)":
            contract_no_col_idx_in_lease_data = idx
            break
    
    if contract_no_col_idx_in_lease_data is None:
        print("  경고: 전대_Lease_Data 시트에서 '계약번호(Ref no.)' 헤더를 찾을 수 없습니다.")
        return
    
    currency_col_letter = get_column_letter(2 + currency_col_idx_in_lease_data - 1)  # B열부터 시작
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx_in_lease_data - 1)  # B열부터 시작
    
    # 각 행에 VLOOKUP 수식 입력
    for row_idx in range(start_row, start_row + num_rows):
        contract_no_cell = f"{contract_no_col_letter}{row_idx}"
        currency_cell = f"{currency_col_letter}{row_idx}"
        
        # VLOOKUP 수식: =VLOOKUP(계약번호, Input_data!범위, 통화열순서, FALSE)
        formula = f"=VLOOKUP({contract_no_cell},{vlookup_range},{col_index},FALSE)"
        add_formula(lease_data_sheet, currency_cell, formula)
    
    print(f"  ✓ 통화 열 입력 완료: {num_rows}행")


def fill_additional_columns(wb, lease_data_sheet, data_manager, num_rows, start_row=7):
    """
    전대_Lease_Data 시트의 주석구분, 내부거래, 중도해지 열 채우기
    
    Args:
        wb: Workbook 객체
        lease_data_sheet: 전대_Lease_Data 시트
        data_manager: DataManager 객체
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 7)
    """
    from excel_utils import get_sheet_by_name, get_table_data, find_table_by_name, add_formula
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    
    print(f"\n[주석구분/내부거래/중도해지 열 채우기]")
    
    # Input_data 시트 확인
    input_data_sheet = get_sheet_by_name(wb, "Input_data")
    if input_data_sheet is None:
        print("  경고: 'Input_data' 시트를 찾을 수 없습니다.")
        return
    
    # input_table 테이블 찾기
    input_ws, table = find_table_by_name(wb, "input_table")
    if table is None:
        print("  경고: 'input_table' 테이블을 찾을 수 없습니다.")
        return
    
    # 테이블 범위 파싱
    ref = table.ref
    start_cell, end_cell = ref.split(':')
    start_col_str, start_row_num = coordinate_from_string(start_cell)
    start_col_idx = column_index_from_string(start_col_str)
    
    # 테이블 데이터 읽기
    try:
        table_data = get_table_data(input_data_sheet, "input_table")
    except Exception as e:
        print(f"  경고: input_table 데이터 읽기 실패: {e}")
        return
    
    # 계약번호(Ref no.) 열 찾기
    ref_no_col = None
    headers = list(table_data[0].keys())
    for col_name in headers:
        col_str = str(col_name).strip()
        if '계약번호' in col_str and 'Ref no' in col_str and ('상위' not in col_str and 'parent' not in col_str.lower()):
            ref_no_col = col_name
            break
    
    if ref_no_col is None:
        print("  경고: Input_data에서 계약번호(Ref no.) 컬럼을 찾을 수 없습니다.")
        return
    
    ref_no_col_idx = headers.index(ref_no_col)
    
    # 전대_Lease_Data 시트의 헤더 목록
    lease_data_headers = [
        "계약번호(Ref no.)",
        "리스명",
        "상위리스_계약번호(Ref no.)",
        "리스개시일",
        "리스종료일",
        "지급일자",
        "지급연도",
        "유효일자",
        "회계연도",
        "결산월",
        "기간(월)",
        "감가상각기간(월)",
        "수령회차",
        "통화",
        "최초환율",
        "평균환율",
        "기말환율",
        "리스료",
        "리스료(환산후)",
        "할인율(연)",
        "임대보증금할인율(연)",
        "리스채권",
        "이자수익",
        "상각액",
        "리스채권(환산후)",
        "전월환산손익취소(리스채권)",
        "당월환산손익(리스채권)",
        "임대보증금",
        "임대보증금_현할차",
        "임대보증금이자",
        "전월환산손익취소(임대보증금)",
        "당월환산손익(임대보증금)",
        "임대보증금(환산후)",
        "중도해지손익",
        "차감_감가상각비",
        "차감_감가상각누계액",
        "차감_사용권자산",
        "차감_사용권자산(순)",
        "주석구분",
        "내부거래",
        "중도해지",
        "리스채권_유동",
    ]
    
    # 계약번호(Ref no.) 열 찾기
    contract_no_col_idx_in_lease_data = None
    for idx, header in enumerate(lease_data_headers, start=1):
        if header == "계약번호(Ref no.)":
            contract_no_col_idx_in_lease_data = idx
            break
    
    if contract_no_col_idx_in_lease_data is None:
        print("  경고: 전대_Lease_Data 시트에서 '계약번호(Ref no.)' 헤더를 찾을 수 없습니다.")
        return
    
    contract_no_col_letter = get_column_letter(2 + contract_no_col_idx_in_lease_data - 1)  # B열부터 시작
    
    # 1. 주석구분 열 채우기
    # =VLOOKUP(B7,Input_data!$C$10:$I$5009,7,FALSE)
    comment_col_idx = None
    for idx, header in enumerate(lease_data_headers, start=1):
        if header == "주석구분":
            comment_col_idx = idx
            break
    
    if comment_col_idx:
        comment_col_letter = get_column_letter(2 + comment_col_idx - 1)
        # Input_data!$C$10:$I$5009 범위 (C열부터 I열까지, 10행부터 5009행까지)
        # C열은 계약번호 열이므로, ref_no_col_idx를 기준으로 계산
        # 사용자가 지정한 범위: C10:I5009
        vlookup_range_comment = f"Input_data!$C$10:$I$5009"
        col_index_comment = 7  # 주석구분 열의 컬럼 순서
        
        for row_idx in range(start_row, start_row + num_rows):
            contract_no_cell = f"{contract_no_col_letter}{row_idx}"
            comment_cell = f"{comment_col_letter}{row_idx}"
            formula = f"=VLOOKUP({contract_no_cell},{vlookup_range_comment},{col_index_comment},FALSE)"
            add_formula(lease_data_sheet, comment_cell, formula)
        
        print(f"  ✓ 주석구분 열 입력 완료: {num_rows}행")
    
    # 2. 내부거래 열 채우기
    # =IF(VLOOKUP(B7,Input_data!$C$10:$H$5009,6,FALSE) = "O","O","")
    internal_col_idx = None
    for idx, header in enumerate(lease_data_headers, start=1):
        if header == "내부거래":
            internal_col_idx = idx
            break
    
    if internal_col_idx:
        internal_col_letter = get_column_letter(2 + internal_col_idx - 1)
        # Input_data!$C$10:$H$5009 범위
        vlookup_range_internal = f"Input_data!$C$10:$H$5009"
        col_index_internal = 6  # 내부거래구분 열의 컬럼 순서
        
        for row_idx in range(start_row, start_row + num_rows):
            contract_no_cell = f"{contract_no_col_letter}{row_idx}"
            internal_cell = f"{internal_col_letter}{row_idx}"
            formula = f"=IF(VLOOKUP({contract_no_cell},{vlookup_range_internal},{col_index_internal},FALSE)=\"O\",\"O\",\"\")"
            add_formula(lease_data_sheet, internal_cell, formula)
        
        print(f"  ✓ 내부거래 열 입력 완료: {num_rows}행")
    
    # 3. 중도해지 열 채우기
    # =IF(EOMONTH(VLOOKUP(B22,Input_data!$C$10:$M$5009,11,FALSE),0)=Lease_Data!H22,"O","")
    early_term_col_idx = None
    for idx, header in enumerate(lease_data_headers, start=1):
        if header == "중도해지":
            early_term_col_idx = idx
            break
    
    # 유효일자 열 찾기
    effective_date_col_idx = None
    for idx, header in enumerate(lease_data_headers, start=1):
        if header == "유효일자":
            effective_date_col_idx = idx
            break
    
    if early_term_col_idx and effective_date_col_idx:
        early_term_col_letter = get_column_letter(2 + early_term_col_idx - 1)
        effective_date_col_letter = get_column_letter(2 + effective_date_col_idx - 1)
        # Input_data!$C$10:$M$5009 범위
        vlookup_range_early_term = f"Input_data!$C$10:$M$5009"
        col_index_early_term = 11  # 리스변경일 열의 컬럼 순서
        
        # 시트명 가져오기
        sheet_name = lease_data_sheet.title
        
        for row_idx in range(start_row, start_row + num_rows):
            contract_no_cell = f"{contract_no_col_letter}{row_idx}"
            effective_date_cell = f"{effective_date_col_letter}{row_idx}"
            early_term_cell = f"{early_term_col_letter}{row_idx}"
            formula = f"=IF(EOMONTH(VLOOKUP({contract_no_cell},{vlookup_range_early_term},{col_index_early_term},FALSE),0)={sheet_name}!{effective_date_cell},\"O\",\"\")"
            add_formula(lease_data_sheet, early_term_cell, formula)
        
        print(f"  ✓ 중도해지 열 입력 완료: {num_rows}행")
    
    # 4. 중도해지손익 열 채우기
    # =(AL8-AK8)-Z8-AD8
    # AL8: 동일행의 차감_사용권자산 열
    # AK8: 동일행의 차감_감가상각누계액 열
    # Z8: 동일행의 리스채권 열
    # AD8: 동일행의 임대보증금_현할차 열
    early_term_gain_loss_col_idx = None
    for idx, header in enumerate(lease_data_headers, start=1):
        if header == "중도해지손익":
            early_term_gain_loss_col_idx = idx
            break
    
    # 필요한 열들 찾기
    deductible_asset_col_idx = None
    deductible_depr_accum_col_idx = None
    lease_receivable_col_idx = None
    deposit_pv_col_idx = None
    
    for idx, header in enumerate(lease_data_headers, start=1):
        if header == "차감_사용권자산":
            deductible_asset_col_idx = idx
        elif header == "차감_감가상각누계액":
            deductible_depr_accum_col_idx = idx
        elif header == "리스채권":
            lease_receivable_col_idx = idx
        elif header == "임대보증금_현할차":
            deposit_pv_col_idx = idx
    
    if (early_term_gain_loss_col_idx and deductible_asset_col_idx and 
        deductible_depr_accum_col_idx and lease_receivable_col_idx and deposit_pv_col_idx):
        early_term_gain_loss_col_letter = get_column_letter(2 + early_term_gain_loss_col_idx - 1)
        deductible_asset_col_letter = get_column_letter(2 + deductible_asset_col_idx - 1)
        deductible_depr_accum_col_letter = get_column_letter(2 + deductible_depr_accum_col_idx - 1)
        lease_receivable_col_letter = get_column_letter(2 + lease_receivable_col_idx - 1)
        deposit_pv_col_letter = get_column_letter(2 + deposit_pv_col_idx - 1)
        
        for row_idx in range(start_row, start_row + num_rows):
            deductible_asset_cell = f"{deductible_asset_col_letter}{row_idx}"
            deductible_depr_accum_cell = f"{deductible_depr_accum_col_letter}{row_idx}"
            lease_receivable_cell = f"{lease_receivable_col_letter}{row_idx}"
            deposit_pv_cell = f"{deposit_pv_col_letter}{row_idx}"
            early_term_gain_loss_cell = f"{early_term_gain_loss_col_letter}{row_idx}"
            
            formula = f"=({deductible_asset_cell}-{deductible_depr_accum_cell})-{lease_receivable_cell}-{deposit_pv_cell}"
            add_formula(lease_data_sheet, early_term_gain_loss_cell, formula)
        
        print(f"  ✓ 중도해지손익 열 입력 완료: {num_rows}행")
    
    # 5. 날짜 열 및 금액 열 형식 적용
    apply_lease_data_formats(lease_data_sheet, num_rows, start_row)
    
    # 6. 테이블 범위에 테두리 적용
    apply_border_to_lease_data_table(lease_data_sheet, num_rows, start_row)


def apply_lease_data_formats(lease_data_sheet, num_rows, start_row=7):
    """
    전대_Lease_Data 시트의 날짜 열과 금액 열에 엑셀 형식 적용
    
    Args:
        lease_data_sheet: 전대_Lease_Data 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 7)
    """
    from openpyxl.utils import get_column_letter
    
    # 헤더 목록
    headers = [
        "계약번호(Ref no.)",
        "리스명",
        "상위리스_계약번호(Ref no.)",
        "리스개시일",
        "리스종료일",
        "지급일자",
        "지급연도",
        "유효일자",
        "회계연도",
        "결산월",
        "기간(월)",
        "감가상각기간(월)",
        "수령회차",
        "통화",
        "최초환율",
        "평균환율",
        "기말환율",
        "리스료",
        "리스료(환산후)",
        "할인율(연)",
        "임대보증금할인율(연)",
        "리스채권",
        "이자수익",
        "상각액",
        "리스채권(환산후)",
        "전월환산손익취소(리스채권)",
        "당월환산손익(리스채권)",
        "임대보증금",
        "임대보증금_현할차",
        "임대보증금이자",
        "전월환산손익취소(임대보증금)",
        "당월환산손익(임대보증금)",
        "임대보증금(환산후)",
        "중도해지손익",
        "차감_감가상각비",
        "차감_감가상각누계액",
        "차감_사용권자산",
        "차감_사용권자산(순)",
        "주석구분",
        "내부거래",
        "중도해지",
        "리스채권_유동",
    ]
    
    # 날짜 열 목록
    date_headers = [
        "리스개시일",
        "리스종료일",
        "지급일자",
        "유효일자",
    ]
    
    # 금액 열 목록
    amount_headers = [
        "리스료",
        "리스료(환산후)",
        "리스채권",
        "이자수익",
        "상각액",
        "리스채권(환산후)",
        "전월환산손익취소(리스채권)",
        "당월환산손익(리스채권)",
        "임대보증금",
        "임대보증금_현할차",
        "임대보증금이자",
        "전월환산손익취소(임대보증금)",
        "당월환산손익(임대보증금)",
        "임대보증금(환산후)",
        "중도해지손익",
        "차감_감가상각비",
        "차감_감가상각누계액",
        "차감_사용권자산",
        "차감_사용권자산(순)",
        "리스채권_유동",
    ]
    
    # 날짜 열 인덱스 찾기
    date_col_indices = []
    for header in date_headers:
        for idx, h in enumerate(headers, start=1):
            if h == header:
                date_col_indices.append(idx)
                break
    
    # 금액 열 인덱스 찾기
    amount_col_indices = []
    for header in amount_headers:
        for idx, h in enumerate(headers, start=1):
            if h == header:
                amount_col_indices.append(idx)
                break
    
    # 날짜 형식 적용 (yyyy-mm-dd 형식)
    date_format = "yyyy-mm-dd"
    for col_idx in date_col_indices:
        col_letter = get_column_letter(2 + col_idx - 1)  # B열부터 시작
        for row_idx in range(start_row, start_row + num_rows):
            cell = lease_data_sheet[f"{col_letter}{row_idx}"]
            cell.number_format = date_format
    
    # 금액 형식 적용 (회계 형식: 천단위 구분, 소수점 없음)
    amount_format = "#,##0"
    for col_idx in amount_col_indices:
        col_letter = get_column_letter(2 + col_idx - 1)  # B열부터 시작
        for row_idx in range(start_row, start_row + num_rows):
            cell = lease_data_sheet[f"{col_letter}{row_idx}"]
            cell.number_format = amount_format
    
    if date_col_indices:
        print(f"  ✓ 날짜 형식 적용: {len(date_col_indices)}개 열")
    if amount_col_indices:
        print(f"  ✓ 금액 형식 적용: {len(amount_col_indices)}개 열")


def apply_border_to_lease_data_table(lease_data_sheet, num_rows, start_row=7):
    """
    전대_Lease_Data 시트의 테이블 범위에 모든 테두리 적용
    
    Args:
        lease_data_sheet: 전대_Lease_Data 시트
        num_rows (int): 데이터 행 수
        start_row (int): 데이터 시작 행 (기본값: 7, 헤더는 6행)
    """
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter
    
    # 테두리 스타일 정의
    thin_border = Side(style='thin')
    border = Border(
        left=thin_border,
        right=thin_border,
        top=thin_border,
        bottom=thin_border
    )
    
    # 헤더 행 (6행)부터 마지막 데이터 행까지
    header_row = 6
    num_cols = 42  # 헤더 개수
    start_col = 2  # B열부터 시작
    
    for row_idx in range(header_row, start_row + num_rows):
        for col_idx in range(start_col, start_col + num_cols):
            col_letter = get_column_letter(col_idx)
            cell = lease_data_sheet[f"{col_letter}{row_idx}"]
            cell.border = border
    
    print(f"  ✓ 테두리 적용 완료: {num_cols}개 열, {num_rows + 1}행 (헤더 포함)")


def create_sublease_lease_data_sheet(wb):
    """
    전대_Lease_Data 시트 생성 및 기본 설정
    
    Args:
        wb: Workbook 객체
    """
    from excel_utils import add_sheet, set_cell_value, add_formula
    
    # 시트 생성
    sheet_name = "전대_Lease_Data"
    ws = add_sheet(wb, sheet_name)
    
    print(f"\n[전대_Lease_Data 시트 생성]")
    
    # 1. B2셀에 "SubLease Schedule" 입력
    set_cell_value(ws, "B2", "SubLease Schedule")
    
    # 2. B4셀에 "기준일" 입력
    set_cell_value(ws, "B4", "기준일")
    
    # 3. C4셀에 =Input_data!$C$7 수식
    add_formula(ws, "C4", "=Input_data!$C$7")
    # C4셀에 간단한 날짜 형식 적용
    ws["C4"].number_format = "yyyy-mm-dd"
    
    # 4. B6셀부터 헤더행 입력 (이미지의 항목들 순서대로 - 42개)
    headers = [
        "계약번호(Ref no.)",
        "리스명",
        "상위리스_계약번호(Ref no.)",
        "리스개시일",
        "리스종료일",
        "지급일자",
        "지급연도",
        "유효일자",
        "회계연도",
        "결산월",
        "기간(월)",
        "감가상각기간(월)",
        "수령회차",
        "통화",
        "최초환율",
        "평균환율",
        "기말환율",
        "리스료",
        "리스료(환산후)",
        "할인율(연)",
        "임대보증금할인율(연)",
        "리스채권",
        "이자수익",
        "상각액",
        "리스채권(환산후)",
        "전월환산손익취소(리스채권)",
        "당월환산손익(리스채권)",
        "임대보증금",
        "임대보증금_현할차",
        "임대보증금이자",
        "전월환산손익취소(임대보증금)",
        "당월환산손익(임대보증금)",
        "임대보증금(환산후)",
        "중도해지손익",
        "차감_감가상각비",
        "차감_감가상각누계액",
        "차감_사용권자산",
        "차감_사용권자산(순)",
        "주석구분",
        "내부거래",
        "중도해지",
        "리스채권_유동",
    ]
    
    # B6셀부터 헤더 입력
    from openpyxl.utils import get_column_letter
    start_col = 2  # B열
    start_row = 6
    
    for idx, header in enumerate(headers):
        col_letter = get_column_letter(start_col + idx)
        cell_address = f"{col_letter}{start_row}"
        set_cell_value(ws, cell_address, header)
    
    print(f"  ✓ 시트 생성 완료: {sheet_name}")
    print(f"  ✓ B2셀: SubLease Schedule")
    print(f"  ✓ B4셀: 기준일")
    print(f"  ✓ C4셀: =Input_data!$C$7")
    print(f"  ✓ B6셀부터 헤더 {len(headers)}개 입력 완료")
    
    return ws


def fill_sublease_lease_data_sheet(wb, data_manager):
    """
    전대_Lease_Data 시트에 각 계약별 상각표 시트의 데이터를 수식으로 가져오기
    
    Args:
        wb: Workbook 객체
        data_manager: DataManager 객체
    """
    from excel_utils import get_sheet_by_name, get_cell_value
    from openpyxl.utils import get_column_letter
    
    # 전대_Lease_Data 시트 가져오기
    lease_data_sheet = get_sheet_by_name(wb, "전대_Lease_Data")
    if lease_data_sheet is None:
        print("  경고: '전대_Lease_Data' 시트를 찾을 수 없습니다.")
        return
    
    print(f"\n[전대_Lease_Data 시트 데이터 채우기]")
    
    # 헤더 목록 (이미지 기준)
    headers = [
        "계약번호(Ref no.)",
        "리스명",
        "상위리스_계약번호(Ref no.)",
        "리스개시일",
        "리스종료일",
        "지급일자",
        "지급연도",
        "유효일자",
        "회계연도",
        "결산월",
        "기간(월)",
        "감가상각기간(월)",
        "수령회차",
        "통화",
        "최초환율",
        "평균환율",
        "기말환율",
        "리스료",
        "리스료(환산후)",
        "할인율(연)",
        "임대보증금할인율(연)",
        "리스채권",
        "이자수익",
        "상각액",
        "리스채권(환산후)",
        "전월환산손익취소(리스채권)",
        "당월환산손익(리스채권)",
        "임대보증금",
        "임대보증금_현할차",
        "임대보증금이자",
        "전월환산손익취소(임대보증금)",
        "당월환산손익(임대보증금)",
        "임대보증금(환산후)",
        "중도해지손익",
        "차감_감가상각비",
        "차감_감가상각누계액",
        "차감_사용권자산",
        "차감_사용권자산(순)",
        "주석구분",
        "내부거래",
        "중도해지",
        "리스채권_유동",
    ]
    
    # 전대_Lease_Data 시트의 헤더 행 (6행)
    lease_data_header_row = 6
    lease_data_start_row = 7  # 데이터 시작 행
    
    # 각 계약별로 데이터 가져오기
    current_data_row = lease_data_start_row
    
    for contract_no, contract_data in data_manager.contract_data.items():
        # 계약번호로 시트 찾기
        contract_sheet = None
        sheet_name = str(contract_no)
        # Excel 시트명에 사용할 수 없는 문자 제거
        invalid_chars = ['\\', '/', '?', '*', '[', ']']
        for char in invalid_chars:
            sheet_name = sheet_name.replace(char, '_')
        sheet_name = sheet_name[:31]  # Excel 시트명 최대 31자
        
        if sheet_name in wb.sheetnames:
            contract_sheet = wb[sheet_name]
        else:
            print(f"  경고: 계약번호 {contract_no}의 시트를 찾을 수 없습니다.")
            continue
        
        # 계약 시트의 헤더 찾기 (1행)
        contract_header_row = 1
        contract_data_start_row = 2  # 계약 시트의 데이터 시작 행
        
        # 계약 시트의 헤더 읽기
        contract_headers = {}
        max_col = contract_sheet.max_column
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            header_cell = f"{col_letter}{contract_header_row}"
            header_value = get_cell_value(contract_sheet, header_cell)
            if header_value:
                contract_headers[str(header_value).strip()] = col_idx
        
        # 계약 시트의 데이터 행 수 확인
        contract_data_rows = 0
        for row_idx in range(contract_data_start_row, contract_sheet.max_row + 1):
            # 첫 번째 열(계약번호)에 값이 있는지 확인
            first_col_value = get_cell_value(contract_sheet, f"A{row_idx}")
            if first_col_value is None or first_col_value == '':
                break
            contract_data_rows += 1
        
        if contract_data_rows == 0:
            print(f"  경고: 계약번호 {contract_no}의 시트에 데이터가 없습니다.")
            continue
        
        print(f"  계약번호 {contract_no}: {contract_data_rows}행 데이터 처리 중...")
        
        # 전대_Lease_Data 시트의 각 헤더에 대해 데이터 가져오기
        for header_idx, header_name in enumerate(headers):
            lease_data_col = get_column_letter(2 + header_idx)  # B열부터 시작
            
            # 계약 시트에서 동일한 헤더 찾기 (정확한 매칭만)
            contract_col_idx = None
            
            # 정확한 매칭만 사용
            if header_name in contract_headers:
                contract_col_idx = contract_headers[header_name]
            
            # 동일한 헤더가 있으면 수식으로 데이터 가져오기
            if contract_col_idx is not None:
                contract_col_letter = get_column_letter(contract_col_idx)
                
                # 각 데이터 행에 수식 입력
                for row_offset in range(contract_data_rows):
                    lease_data_row = current_data_row + row_offset
                    contract_data_row = contract_data_start_row + row_offset
                    
                    # 수식: =시트명!열명행번호
                    formula = f"={sheet_name}!{contract_col_letter}{contract_data_row}"
                    from excel_utils import add_formula
                    add_formula(lease_data_sheet, f"{lease_data_col}{lease_data_row}", formula)
        
        # 다음 계약을 위해 현재 행 위치 업데이트
        current_data_row += contract_data_rows
    
    total_rows = current_data_row - lease_data_start_row
    print(f"  ✓ 전대_Lease_Data 시트 데이터 채우기 완료: 총 {total_rows}행")
    
    # 통화 열 채우기
    fill_currency_column(wb, lease_data_sheet, data_manager, total_rows, lease_data_start_row)
    
    # 주석구분, 내부거래, 중도해지 열 채우기
    fill_additional_columns(wb, lease_data_sheet, data_manager, total_rows, lease_data_start_row)

